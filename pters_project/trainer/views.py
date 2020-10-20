# Create your views here.
import collections
import datetime
import json
import logging
import random
import urllib

from operator import attrgetter
from urllib.parse import quote

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.core.paginator import EmptyPage
from django.core.paginator import Paginator
from django.db import IntegrityError, InternalError, transaction
from django.db.models import Q, Sum
from django.db.models.expressions import RawSQL
from django.http import HttpResponse, JsonResponse, request
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.datastructures import MultiValueDictKeyError
from django.views import View
from django.views.generic import TemplateView, RedirectView
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook

from admin_spooner.functions import func_upload_board_content_image_logic
from configs import settings
from configs.functions import func_upload_profile_image_logic, func_delete_profile_image_logic
from configs.views import AccessTestMixin, get_function_auth_type_cd, func_setting_data_update
from configs.const import ON_SCHEDULE_TYPE, OFF_SCHEDULE_TYPE, USE, UN_USE, AUTO_FINISH_OFF, \
    MEMBER_RESERVE_PROHIBITION_ON, SORT_MEMBER_NAME, SORT_REMAIN_COUNT, SORT_START_DATE, SORT_ASC, SORT_REG_COUNT, \
    GROUP_SCHEDULE, SCHEDULE_DUPLICATION_ENABLE, LECTURE_TYPE_ONE_TO_ONE, STATE_CD_IN_PROGRESS, STATE_CD_NOT_PROGRESS, \
    STATE_CD_ABSENCE, STATE_CD_FINISH, PERMISSION_STATE_CD_APPROVE, AUTH_TYPE_VIEW, AUTH_TYPE_WAIT, AUTH_TYPE_DELETE, \
    LECTURE_TYPE_NORMAL, SHOW, SORT_TICKET_TYPE, SORT_TICKET_NAME, SORT_TICKET_MEMBER_COUNT, SORT_TICKET_CREATE_DATE, \
    SORT_LECTURE_NAME, SORT_LECTURE_MEMBER_COUNT, SORT_LECTURE_CAPACITY_COUNT, SORT_LECTURE_CREATE_DATE, \
    CALENDAR_TIME_SELECTOR_BASIC, SORT_END_DATE, SORT_MEMBER_TICKET, SORT_SCHEDULE_DT, STATE_CD_REFUND, \
    SORT_SCHEDULE_MONTHLY, SHARED_PROGRAM, MY_PROGRAM, PROGRAM_SELECT, PROGRAM_LECTURE_CONNECT_DELETE, \
    PROGRAM_LECTURE_CONNECT_ACCEPT, BOARD_TYPE_CD_NOTICE, ON_SCHEDULE, ALARM_PAGINATION_COUNTER, STATE_CD_HOLDING, \
    CLOSED_SCHEDULE_TYPE, SCHEDULE_PAGINATION_COUNTER, OWN_TYPE_OWNER, OWN_TYPE_SHARE, SORT_TRAINER_NAME, \
    OWN_TYPE_EMPLOYEE, SORT_SHOP_START_DATE
from board.models import BoardTb, QATb, NoticeTb
from login.models import MemberTb, LogTb, CommonCdTb, SnsInfoTb
from schedule.functions import func_refresh_member_ticket_count, func_get_trainer_attend_schedule, \
    func_get_lecture_member_ticket_id, func_get_trainer_schedule_all, func_get_trainer_schedule_info, \
    func_get_lecture_schedule_all, func_get_member_schedule_all_by_member_ticket, \
    func_get_member_schedule_all_by_schedule_dt, func_get_member_schedule_all_by_monthly, \
    func_get_permission_wait_schedule_all, func_add_schedule, func_send_push_notice, func_get_member_ticket_id_old, \
    func_get_repeat_schedule_date_list, func_get_lecture_trainer_schedule_all_by_schedule_dt, \
    func_get_lecture_trainer_schedule_all_by_monthly
from schedule.models import ScheduleTb, RepeatScheduleTb, HolidayTb, ScheduleAlarmTb
from stats.functions import get_sales_data
from trainee.models import MemberTicketTb, MemberClosedDateHistoryTb, MemberShopTb, MemberPaymentHistoryTb
from payment.models import PaymentInfoTb, ProductFunctionAuthTb, CouponMemberTb
from .functions import func_get_trainer_setting_list, \
    func_get_member_ing_list, func_get_member_end_list, func_get_class_member_ing_list, func_get_class_member_end_list, \
    func_get_member_info, func_get_member_from_member_ticket_list, func_check_member_connection_info, \
    func_get_member_lecture_list, func_get_member_ticket_list, func_get_lecture_info, func_add_member_ticket_info, \
    func_get_ticket_info, func_delete_member_ticket_info, func_update_lecture_member_fix_status_cd, \
    update_user_setting_data, update_program_setting_data, func_get_member_ticket_info, func_get_trainer_info, \
    update_alarm_setting_data, func_add_hold_member_ticket_info, func_delete_hold_member_ticket_info, \
    func_get_trainer_ing_list, func_get_class_trainer_end_list, func_get_trainer_end_list, \
    func_get_class_trainer_ing_list, func_get_trainer_ing_list_connect
from .models import ClassMemberTicketTb, LectureTb, ClassTb, MemberClassTb, BackgroundImgTb, \
    SettingTb, TicketTb, TicketLectureTb, CenterTrainerTb, LectureMemberTb, ProgramAuthTb, ProgramBoardTb, \
    ProgramNoticeTb, BugMemberTicketPriceTb, ScheduleClosedTb, ScheduleClosedDayTb, ShopTb

logger = logging.getLogger(__name__)


class IndexView(LoginRequiredMixin, AccessTestMixin, RedirectView):
    url = '/trainer/class_select/'

    def get(self, request, **kwargs):

        class_id = request.session.get('class_id', '')
        class_auth_data = MemberClassTb.objects.select_related('class_tb'
                                                               ).filter(member_id=request.user.id,
                                                                        auth_cd=AUTH_TYPE_VIEW,
                                                                        use=USE).order_by('-class_tb_id')

        error = None
        if class_id is None or class_id == '':
            request.session['shared_program_flag'] = MY_PROGRAM
            if len(class_auth_data) == 0:
                self.url = '/trainer/add_program/'
            elif len(class_auth_data) == 1:
                self.url = '/trainer/trainer_main/'
                for class_info in class_auth_data:
                    request.session['class_id'] = class_info.class_tb_id
                    request.session['class_hour'] = class_info.class_tb.class_hour
                    request.session['class_type_code'] = class_info.class_tb.subject_cd
                    request.session['class_type_name'] = class_info.class_tb.get_class_type_cd_name()
                    if str(class_info.class_tb.member_id) != str(request.user.id):
                        request.session['class_type_name'] = class_info.class_tb.get_class_type_cd_name() \
                                                             + '-' + class_info.class_tb.member.name
                        request.session['shared_program_flag'] = SHARED_PROGRAM
                    request.session['class_center_name'] = class_info.class_tb.get_center_name()
                    request.session['trainer_name'] = class_info.class_tb.member.name

            else:
                self.url = '/trainer/trainer_main/'
                temp_class_counter = 0
                for class_info in class_auth_data:
                    class_member_ticket_counter = ClassMemberTicketTb.objects.filter(class_tb_id=class_info.class_tb_id,
                                                                                     auth_cd=AUTH_TYPE_VIEW,
                                                                                     use=USE).count()
                    if class_member_ticket_counter > temp_class_counter:
                        request.session['class_id'] = class_info.class_tb_id
                        request.session['class_hour'] = class_info.class_tb.class_hour
                        request.session['class_type_code'] = class_info.class_tb.subject_cd
                        request.session['class_type_name'] = class_info.class_tb.get_class_type_cd_name()
                        if str(class_info.class_tb.member_id) != str(request.user.id):
                            request.session['class_type_name'] = class_info.class_tb.get_class_type_cd_name() \
                                                                 + '-' + class_info.class_tb.member.name
                            request.session['shared_program_flag'] = SHARED_PROGRAM
                        request.session['class_center_name'] = class_info.class_tb.get_center_name()
                        request.session['trainer_name'] = class_info.class_tb.member.name
                        temp_class_counter = class_member_ticket_counter
                    if class_member_ticket_counter == 0 and temp_class_counter == 0:
                        request.session['class_id'] = class_info.class_tb_id
                        request.session['class_hour'] = class_info.class_tb.class_hour
                        request.session['class_type_code'] = class_info.class_tb.subject_cd
                        request.session['class_type_name'] = class_info.class_tb.get_class_type_cd_name()
                        if str(class_info.class_tb.member_id) != str(request.user.id):
                            request.session['class_type_name'] = class_info.class_tb.get_class_type_cd_name() \
                                                                 + '-' + class_info.class_tb.member.name
                            request.session['shared_program_flag'] = SHARED_PROGRAM
                        request.session['class_center_name'] = class_info.class_tb.get_center_name()
                        request.session['trainer_name'] = class_info.class_tb.member.name

                # self.url = '/trainer/class_select/'
        else:
            self.url = '/trainer/trainer_main/'
        request.session['APP_VERSION'] = settings.APP_VERSION
        # get_setting_info(request)
        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return super(IndexView, self).get(request, **kwargs)

    def get_redirect_url(self, *args, **kwargs):
        return super(IndexView, self).get_redirect_url(*args, **kwargs)


class GetErrorInfoView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'ajax/trainer_error_ajax.html'

    def get_context_data(self, **kwargs):
        context = super(GetErrorInfoView, self).get_context_data(**kwargs)
        return context


# 일정 기능 ############### ############### ############### ############### ############### ############### ##############

class GetTrainerScheduleAllView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        date = request.GET.get('date', '')
        day = request.GET.get('day', '')
        trainer_id = request.GET.get('trainer_id', 'all')
        today = datetime.date.today()

        if date != '':
            today = datetime.datetime.strptime(date, '%Y-%m-%d')
        if day == '':
            day = 31
        start_date = today - datetime.timedelta(days=int(day))
        end_date = today + datetime.timedelta(days=int(day))
        all_schedule_data = func_get_trainer_schedule_all(class_id, start_date, end_date, trainer_id)

        return JsonResponse(all_schedule_data, json_dumps_params={'ensure_ascii': True})


class GetLectureMemberScheduleListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        lecture_schedule_id = request.GET.get('lecture_schedule_id', '')
        error = None
        lecture_schedule_list = []

        if lecture_schedule_id == '':
            error = '수업 정보를 불러오지 못했습니다.'

        if error is None:
            # 그룹 수업 일정에 해당하는 회원의 일정들 불러오기
            lecture_schedule_data = ScheduleTb.objects.select_related(
                'member_ticket_tb__member',
                'reg_member').filter(class_tb_id=class_id, lecture_schedule_id=lecture_schedule_id,
                                     use=USE).order_by('start_dt')

            for lecture_schedule_info in lecture_schedule_data:
                mod_member_id = ''
                mod_member_name = ''
                if lecture_schedule_info.mod_member is not None and lecture_schedule_info.mod_member != '':
                    mod_member_id = lecture_schedule_info.mod_member_id
                    mod_member_name = lecture_schedule_info.mod_member.name
                schedule_info = {'schedule_id': lecture_schedule_info.schedule_id,
                                 'member_name': lecture_schedule_info.member_ticket_tb.member.name,
                                 'schedule_type': GROUP_SCHEDULE,
                                 'start_dt': str(lecture_schedule_info.start_dt),
                                 'end_dt': str(lecture_schedule_info.end_dt),
                                 'state_cd': lecture_schedule_info.state_cd,
                                 'note': lecture_schedule_info.note,
                                 'private_note': lecture_schedule_info.private_note,
                                 'reg_member_id': lecture_schedule_info.reg_member_id,
                                 'reg_member_name': lecture_schedule_info.reg_member.name,
                                 'mod_member_id': mod_member_id,
                                 'mod_member_name': mod_member_name,
                                 'mod_dt': lecture_schedule_info.mod_dt,
                                 'reg_dt': lecture_schedule_info.reg_dt,
                                 'daily_record_id': lecture_schedule_info.daily_record_tb_id
                                 }

                lecture_schedule_list.append(schedule_info)

        if error is not None:
            logger.error(request.user.first_name + ' ' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'lecture_member_schedule_data': lecture_schedule_list},
                            json_dumps_params={'ensure_ascii': True})


class GetTrainerScheduleInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        schedule_id = request.GET.get('schedule_id')

        date_schedule_list = func_get_trainer_schedule_info(class_id, schedule_id)
        return JsonResponse({'schedule_info': date_schedule_list}, json_dumps_params={'ensure_ascii': True})


class GetMemberScheduleAllView(LoginRequiredMixin, AccessTestMixin, View):
    # template_name = 'test.html'
    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_id = request.GET.get('member_id', '')
        page = request.GET.get('page', 1)
        sort = request.GET.get('sort_val', SORT_SCHEDULE_DT)
        error = None
        ordered_schedule_dict = collections.OrderedDict()

        if member_id == '':
            error = '회원 정보를 불러오지 못했습니다.'
        # context = {}
        if error is None:
            if str(sort) == str(SORT_SCHEDULE_DT):
                # ordered_schedule_dict = {
                #     'member_schedule': func_get_member_schedule_all_by_schedule_dt(class_id, member_id, page)
                # }
                ordered_schedule_dict = func_get_member_schedule_all_by_schedule_dt(class_id, member_id, page)
                # context['test'] = ordered_schedule_dict['member_schedule']
            elif str(sort) == str(SORT_SCHEDULE_MONTHLY):
                ordered_schedule_dict = func_get_member_schedule_all_by_monthly(class_id, member_id, page)
            elif str(sort) == str(SORT_MEMBER_TICKET):
                ordered_schedule_dict = func_get_member_schedule_all_by_member_ticket(class_id, member_id, page)
                # context['test'] = ordered_schedule_dict
        else:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        # return render(request, self.template_name, context)
        return JsonResponse(ordered_schedule_dict, json_dumps_params={'ensure_ascii': True})


class GetLectureTrainerScheduleAllView(LoginRequiredMixin, AccessTestMixin, View):
    # template_name = 'test.html'
    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        trainer_id = request.GET.get('trainer_id', '')
        page = request.GET.get('page', 1)
        sort = request.GET.get('sort_val', SORT_SCHEDULE_DT)
        error = None
        ordered_schedule_dict = collections.OrderedDict()

        if trainer_id == '':
            error = '강사 정보를 불러오지 못했습니다.'
        # context = {}
        if error is None:
            if str(sort) == str(SORT_SCHEDULE_DT):
                ordered_schedule_dict = func_get_lecture_trainer_schedule_all_by_schedule_dt(class_id, trainer_id, page)
            elif str(sort) == str(SORT_SCHEDULE_MONTHLY):
                ordered_schedule_dict = func_get_lecture_trainer_schedule_all_by_monthly(class_id, trainer_id, page)
        else:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        # return render(request, self.template_name, context)
        return JsonResponse(ordered_schedule_dict, json_dumps_params={'ensure_ascii': True})


class GetLectureScheduleListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        lecture_id = request.GET.get('lecture_id', None)
        page = request.GET.get('page', 1)
        sort = request.GET.get('sort_val', SORT_SCHEDULE_DT)
        error = None
        ordered_schedule_dict = collections.OrderedDict()

        if lecture_id is None or lecture_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            ordered_schedule_dict = func_get_lecture_schedule_all(class_id, lecture_id, page)
        else:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse(ordered_schedule_dict, json_dumps_params={'ensure_ascii': True})


class GetRepeatScheduleAllView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        today = datetime.date.today()
        off_repeat_schedule_list = []
        closed_repeat_schedule_list = []
        member_repeat_schedule_list = []
        lecture_member_repeat_schedule_ordered_dict = collections.OrderedDict()

        # try:
        #     one_to_one_lecture = LectureTb.objects.filter(class_tb_id=class_id, lecture_type_cd=LECTURE_TYPE_ONE_TO_ONE,
        #                                                   use=USE).order_by('reg_dt').latest()
        #     member_lecture_ing_color_cd = one_to_one_lecture.ing_color_cd
        #     member_lecture_end_color_cd = one_to_one_lecture.end_color_cd
        #     member_lecture_ing_font_color_cd = one_to_one_lecture.ing_font_color_cd
        #     member_lecture_end_font_color_cd = one_to_one_lecture.end_font_color_cd
        # except ObjectDoesNotExist:
        #     member_lecture_ing_color_cd = ''
        #     member_lecture_end_color_cd = ''
        #     member_lecture_ing_font_color_cd = ''
        #     member_lecture_end_font_color_cd = ''

        # OFF 반복 일정 정보 불러오기
        off_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
            'reg_member').filter(class_tb_id=class_id,
                                 en_dis_type=OFF_SCHEDULE_TYPE).exclude(end_date__lt=today).order_by('-reg_dt')

        # 휴무일 반복 일정 정보 불러오기
        closed_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
            'reg_member').filter(class_tb_id=class_id,
                                 en_dis_type=CLOSED_SCHEDULE_TYPE).exclude(end_date__lt=today).order_by('-reg_dt')

        # 회원의 반복 일정 정보 불러오기
        member_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
            'member_ticket_tb__member',
            'reg_member').filter(class_tb_id=class_id, en_dis_type=ON_SCHEDULE_TYPE,
                                 lecture_tb__lecture_type_cd=LECTURE_TYPE_ONE_TO_ONE,
                                 lecture_schedule_id__isnull=True
                                 ).exclude(end_date__lt=today).order_by('-reg_dt', 'lecture_tb', 'lecture_schedule_id')

        # 수업 반복 일정 정보 불러오기
        lecture_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
            'lecture_tb',
            'reg_member').filter(class_tb_id=class_id, en_dis_type=ON_SCHEDULE_TYPE, lecture_tb__isnull=False,
                                 lecture_tb__lecture_type_cd=LECTURE_TYPE_NORMAL,
                                 lecture_schedule_id__isnull=True
                                 ).exclude(end_date__lt=today).order_by('-reg_dt', 'lecture_tb', 'lecture_schedule_id',)

        # 수업에 속한 회원의 반복 일정 정보 불러오기
        lecture_member_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
            'lecture_tb', 'member_ticket_tb__member', 'reg_member').filter(
            class_tb_id=class_id, en_dis_type=ON_SCHEDULE_TYPE, lecture_tb__isnull=False,
            lecture_schedule_id__isnull=False).order_by('-reg_dt', 'lecture_tb', 'lecture_schedule_id')

        week_order = ['SUN', 'MON', 'TUE', 'WED', 'THS', 'FRI', 'SAT']
        week_order = {key: i for i, key in enumerate(week_order)}
        for off_repeat_schedule_info in off_repeat_schedule_data:
            week_data = off_repeat_schedule_info.week_info.split('/')
            week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))
            mod_member_id = ''
            mod_member_name = ''
            if off_repeat_schedule_info.mod_member is not None and off_repeat_schedule_info.mod_member != '':
                mod_member_id = off_repeat_schedule_info.mod_member_id
                mod_member_name = off_repeat_schedule_info.mod_member.name
            off_repeat_schedule = {
                'repeat_schedule_id': off_repeat_schedule_info.repeat_schedule_id,
                'repeat_type_cd': off_repeat_schedule_info.repeat_type_cd,
                'week_info': "/".join(week_data),
                'start_date': off_repeat_schedule_info.start_date,
                'end_date': off_repeat_schedule_info.end_date,
                'start_time': off_repeat_schedule_info.start_time,
                'end_time': off_repeat_schedule_info.end_time,
                'time_duration': off_repeat_schedule_info.time_duration,
                'state_cd': off_repeat_schedule_info.state_cd,
                'reg_member_id': off_repeat_schedule_info.reg_member_id,
                'reg_member_name': off_repeat_schedule_info.reg_member.name,
                'mod_member_id': mod_member_id,
                'mod_member_name': mod_member_name,
                'mod_dt': str(off_repeat_schedule_info.mod_dt),
                'reg_dt': str(off_repeat_schedule_info.reg_dt)
            }
            off_repeat_schedule_list.append(off_repeat_schedule)

        for closed_repeat_schedule_info in closed_repeat_schedule_data:
            week_data = closed_repeat_schedule_info.week_info.split('/')
            week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))
            mod_member_id = ''
            mod_member_name = ''
            if closed_repeat_schedule_info.mod_member is not None and closed_repeat_schedule_info.mod_member != '':
                mod_member_id = closed_repeat_schedule_info.mod_member_id
                mod_member_name = closed_repeat_schedule_info.mod_member.name
            closed_repeat_schedule = {
                'repeat_schedule_id': closed_repeat_schedule_info.repeat_schedule_id,
                'repeat_type_cd': closed_repeat_schedule_info.repeat_type_cd,
                'week_info': "/".join(week_data),
                'start_date': closed_repeat_schedule_info.start_date,
                'end_date': closed_repeat_schedule_info.end_date,
                'start_time': closed_repeat_schedule_info.start_time,
                'end_time': closed_repeat_schedule_info.end_time,
                'time_duration': closed_repeat_schedule_info.time_duration,
                'state_cd': closed_repeat_schedule_info.state_cd,
                'reg_member_id': closed_repeat_schedule_info.reg_member_id,
                'reg_member_name': closed_repeat_schedule_info.reg_member.name,
                'mod_member_id': mod_member_id,
                'mod_member_name': mod_member_name,
                'mod_dt': str(closed_repeat_schedule_info.mod_dt),
                'reg_dt': str(closed_repeat_schedule_info.reg_dt)
            }
            closed_repeat_schedule_list.append(closed_repeat_schedule)

        for member_repeat_schedule_info in member_repeat_schedule_data:
            member_profile_url = '/static/common/icon/icon_account.png'
            if member_repeat_schedule_info.member_ticket_tb.member.profile_url is not None \
                    and member_repeat_schedule_info.member_ticket_tb.member.profile_url != '':
                member_profile_url = member_repeat_schedule_info.member_ticket_tb.member.profile_url

            week_data = member_repeat_schedule_info.week_info.split('/')
            week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

            mod_member_id = ''
            mod_member_name = ''
            if member_repeat_schedule_info.mod_member is not None and member_repeat_schedule_info.mod_member != '':
                mod_member_id = member_repeat_schedule_info.mod_member_id
                mod_member_name = member_repeat_schedule_info.mod_member.name
            member_repeat_schedule = {
                'repeat_schedule_id': member_repeat_schedule_info.repeat_schedule_id,
                'repeat_type_cd': member_repeat_schedule_info.repeat_type_cd,
                'week_info': "/".join(week_data),
                'repeat_start_date': member_repeat_schedule_info.start_date,
                'repeat_end_date': member_repeat_schedule_info.end_date,
                'repeat_start_time': member_repeat_schedule_info.start_time,
                'repeat_end_time': member_repeat_schedule_info.end_time,
                'repeat_time_duration': member_repeat_schedule_info.time_duration,
                'repeat_state_cd': member_repeat_schedule_info.state_cd,
                'reg_member_id': member_repeat_schedule_info.reg_member_id,
                'reg_member_name': member_repeat_schedule_info.reg_member.name,
                'mod_member_id': mod_member_id,
                'mod_member_name': mod_member_name,
                'mod_dt': str(member_repeat_schedule_info.mod_dt),
                'reg_dt': str(member_repeat_schedule_info.reg_dt),
                'lecture_ing_color_cd': member_repeat_schedule_info.lecture_tb.ing_color_cd,
                'lecture_end_color_cd': member_repeat_schedule_info.lecture_tb.end_color_cd,
                'lecture_ing_font_color_cd': member_repeat_schedule_info.lecture_tb.ing_font_color_cd,
                'lecture_end_font_color_cd': member_repeat_schedule_info.lecture_tb.end_font_color_cd,
                'member_id': member_repeat_schedule_info.member_ticket_tb.member.member_id,
                'member_name': member_repeat_schedule_info.member_ticket_tb.member.name,
                'member_profile_url': member_profile_url

            }
            member_repeat_schedule_list.append(member_repeat_schedule)

        for lecture_repeat_schedule_info in lecture_repeat_schedule_data:
            week_data = lecture_repeat_schedule_info.week_info.split('/')
            week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

            mod_member_id = ''
            mod_member_name = ''
            if lecture_repeat_schedule_info.mod_member is not None and lecture_repeat_schedule_info.mod_member != '':
                mod_member_id = lecture_repeat_schedule_info.mod_member_id
                mod_member_name = lecture_repeat_schedule_info.mod_member.name
            lecture_member_repeat_schedule_ordered_dict[lecture_repeat_schedule_info.repeat_schedule_id] = {
                'repeat_schedule_id': lecture_repeat_schedule_info.repeat_schedule_id,
                'repeat_type_cd': lecture_repeat_schedule_info.repeat_type_cd,
                'week_info': "/".join(week_data),
                'start_date': lecture_repeat_schedule_info.start_date,
                'end_date': lecture_repeat_schedule_info.end_date,
                'start_time': lecture_repeat_schedule_info.start_time,
                'end_time': lecture_repeat_schedule_info.end_time,
                'time_duration': lecture_repeat_schedule_info.time_duration,
                'state_cd': lecture_repeat_schedule_info.state_cd,
                'reg_member_id': lecture_repeat_schedule_info.reg_member_id,
                'reg_member_name': lecture_repeat_schedule_info.reg_member.name,
                'mod_member_id': mod_member_id,
                'mod_member_name': mod_member_name,
                'mod_dt': str(lecture_repeat_schedule_info.mod_dt),
                'reg_dt': str(lecture_repeat_schedule_info.reg_dt),
                'lecture_id': lecture_repeat_schedule_info.lecture_tb.lecture_id,
                'lecture_name': lecture_repeat_schedule_info.lecture_tb.name,
                'lecture_max_member_num': lecture_repeat_schedule_info.lecture_tb.member_num,
                # 'lecture_max_member_num_view_flag': lecture_repeat_schedule_info.lecture_tb.member_num_view_flag,
                'lecture_ing_color_cd': lecture_repeat_schedule_info.lecture_tb.ing_color_cd,
                'lecture_end_color_cd': lecture_repeat_schedule_info.lecture_tb.end_color_cd,
                'lecture_ing_font_color_cd': lecture_repeat_schedule_info.lecture_tb.ing_font_color_cd,
                'lecture_end_font_color_cd': lecture_repeat_schedule_info.lecture_tb.end_font_color_cd,
                'lecture_member_repeat_schedule_list': []
            }

        for lecture_member_repeat_schedule_info in lecture_member_repeat_schedule_data:
            member_profile_url = '/static/common/icon/icon_account.png'
            if lecture_member_repeat_schedule_info.member_ticket_tb.member.profile_url is not None \
                    and lecture_member_repeat_schedule_info.member_ticket_tb.member.profile_url != '':
                member_profile_url = lecture_member_repeat_schedule_info.member_ticket_tb.member.profile_url

            week_data = lecture_member_repeat_schedule_info.week_info.split('/')
            week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

            mod_member_id = ''
            mod_member_name = ''
            if lecture_member_repeat_schedule_info.mod_member is not None and lecture_member_repeat_schedule_info.mod_member != '':
                mod_member_id = lecture_member_repeat_schedule_info.mod_member_id
                mod_member_name = lecture_member_repeat_schedule_info.mod_member.name
            lecture_member_repeat_schedule_dict = {
                'repeat_schedule_id': lecture_member_repeat_schedule_info.repeat_schedule_id,
                'repeat_type_cd': lecture_member_repeat_schedule_info.repeat_type_cd,
                'week_info': "/".join(week_data),
                'start_date': lecture_member_repeat_schedule_info.start_date,
                'end_date': lecture_member_repeat_schedule_info.end_date,
                'start_time': lecture_member_repeat_schedule_info.start_time,
                'end_time': lecture_member_repeat_schedule_info.end_time,
                'time_duration': lecture_member_repeat_schedule_info.time_duration,
                'state_cd': lecture_member_repeat_schedule_info.state_cd,
                'reg_member_id': lecture_member_repeat_schedule_info.reg_member_id,
                'reg_member_name': lecture_member_repeat_schedule_info.reg_member.name,
                'mod_member_id': mod_member_id,
                'mod_member_name': mod_member_name,
                'mod_dt': str(lecture_member_repeat_schedule_info.mod_dt),
                'reg_dt': str(lecture_member_repeat_schedule_info.reg_dt),
                'member_id': lecture_member_repeat_schedule_info.member_ticket_tb.member.member_id,
                'member_name': lecture_member_repeat_schedule_info.member_ticket_tb.member.name,
                'member_profile_url': member_profile_url
            }
            lecture_schedule_id = lecture_member_repeat_schedule_info.lecture_schedule_id
            try:
                lecture_member_repeat_schedule_ordered_dict[
                    lecture_schedule_id]['lecture_member_repeat_schedule_list'].append(lecture_member_repeat_schedule_dict)
            except KeyError:
                lecture_member_repeat_schedule_dict = None

        repeat_schedule = collections.OrderedDict()
        repeat_schedule['off_repeat_schedule_data'] = off_repeat_schedule_list
        repeat_schedule['closed_repeat_schedule_data'] = closed_repeat_schedule_list
        repeat_schedule['member_repeat_schedule_list'] = member_repeat_schedule_list
        repeat_schedule['lecture_repeat_schedule_list'] = lecture_member_repeat_schedule_ordered_dict
        return JsonResponse(repeat_schedule, json_dumps_params={'ensure_ascii': True})


class GetPermissionWaitScheduleAllView(LoginRequiredMixin, AccessTestMixin, View):
    # template_name = 'test.html'
    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        page = request.GET.get('page', 1)

        error = None
        ordered_schedule_dict = collections.OrderedDict()

        # context = {}
        if error is None:
            ordered_schedule_dict = {
                'permission_wait_schedule': func_get_permission_wait_schedule_all(class_id, page)
            }
        else:
            logger.error('[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        # return render(request, self.template_name, context)
        return JsonResponse(ordered_schedule_dict, json_dumps_params={'ensure_ascii': True})


class GetOffRepeatScheduleView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        off_repeat_schedule_list = []

        off_repeat_schedule_data = RepeatScheduleTb.objects.select_related('reg_member').filter(class_tb_id=class_id, en_dis_type=OFF_SCHEDULE_TYPE)

        for off_repeat_schedule_info in off_repeat_schedule_data:
            mod_member_id = ''
            mod_member_name = ''
            if off_repeat_schedule_info.mod_member is not None and off_repeat_schedule_info.mod_member != '':
                mod_member_id = off_repeat_schedule_info.mod_member_id
                mod_member_name = off_repeat_schedule_info.mod_member.name
            off_repeat_schedule = {'repeat_schedule_id': off_repeat_schedule_info.repeat_schedule_id,
                                   'repeat_type_cd': off_repeat_schedule_info.repeat_type_cd,
                                   'start_date': off_repeat_schedule_info.start_date,
                                   'end_date': off_repeat_schedule_info.end_date,
                                   'start_time': off_repeat_schedule_info.start_time,
                                   'end_time': off_repeat_schedule_info.end_time,
                                   'time_duration': off_repeat_schedule_info.time_duration,
                                   'state_cd': off_repeat_schedule_info.state_cd,
                                   'reg_member_id': off_repeat_schedule_info.reg_member_id,
                                   'reg_member_name': off_repeat_schedule_info.reg_member.name,
                                   'mod_member_id': mod_member_id,
                                   'mod_member_name': mod_member_name,
                                   'mod_dt': str(off_repeat_schedule_info.mod_dt),
                                   'reg_dt': str(off_repeat_schedule_info.reg_dt)}
            off_repeat_schedule_list.append(off_repeat_schedule)

        messages.error(request, '')

        return JsonResponse({'off_repeat_schedule_data': off_repeat_schedule_list},
                            json_dumps_params={'ensure_ascii': True})


class GetLectureRepeatScheduleListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = request.session.get('class_id', '')
        lecture_id = self.request.GET.get('lecture_id', '')
        error = None
        today = datetime.date.today()
        lecture_info = None
        lecture_repeat_schedule_list = []
        member_repeat_schedule_list = []
        lecture_member_repeat_schedule_ordered_dict = collections.OrderedDict()

        if lecture_id == '':
            error = '수업 정보를 불러오지 못했습니다.[1]'
        if error is None:
            try:
                lecture_info = LectureTb.objects.get(lecture_id=lecture_id)
            except ObjectDoesNotExist:
                error = '수업 정보를 불러오지 못했습니다.[2]'
        if error is None:

            week_order = ['SUN', 'MON', 'TUE', 'WED', 'THS', 'FRI', 'SAT']
            week_order = {key: i for i, key in enumerate(week_order)}

            if lecture_info.lecture_type_cd == LECTURE_TYPE_ONE_TO_ONE:
                # 1:1 수업 회원의 반복 일정 정보 불러오기
                member_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
                    'member_ticket_tb__member', 'lecture_tb',
                    'reg_member').filter(class_tb_id=class_id, en_dis_type=ON_SCHEDULE_TYPE,
                                         lecture_tb_id=lecture_id,
                                         lecture_schedule_id__isnull=True
                                         ).exclude(end_date__lt=today).order_by('-reg_dt', 'lecture_tb',
                                                                                'lecture_schedule_id')

                for member_repeat_schedule_info in member_repeat_schedule_data:
                    member_profile_url = '/static/common/icon/icon_account.png'
                    if member_repeat_schedule_info.member_ticket_tb.member.profile_url is not None \
                            and member_repeat_schedule_info.member_ticket_tb.member.profile_url != '':
                        member_profile_url = member_repeat_schedule_info.member_ticket_tb.member.profile_url

                    week_data = member_repeat_schedule_info.week_info.split('/')
                    week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

                    mod_member_id = ''
                    mod_member_name = ''
                    if member_repeat_schedule_info.mod_member is not None and member_repeat_schedule_info.mod_member != '':
                        mod_member_id = member_repeat_schedule_info.mod_member_id
                        mod_member_name = member_repeat_schedule_info.mod_member.name
                    member_repeat_schedule = {
                        'repeat_schedule_id': member_repeat_schedule_info.repeat_schedule_id,
                        'repeat_type_cd': member_repeat_schedule_info.repeat_type_cd,
                        'week_info': "/".join(week_data),
                        'start_date': member_repeat_schedule_info.start_date,
                        'end_date': member_repeat_schedule_info.end_date,
                        'start_time': member_repeat_schedule_info.start_time,
                        'end_time': member_repeat_schedule_info.end_time,
                        'time_duration': member_repeat_schedule_info.time_duration,
                        'state_cd': member_repeat_schedule_info.state_cd,
                        'reg_member_id': member_repeat_schedule_info.reg_member_id,
                        'reg_member_name': member_repeat_schedule_info.reg_member.name,
                        'mod_member_id': mod_member_id,
                        'mod_member_name': mod_member_name,
                        'mod_dt': str(member_repeat_schedule_info.mod_dt),
                        'reg_dt': str(member_repeat_schedule_info.reg_dt),
                        'lecture_id': member_repeat_schedule_info.lecture_tb.lecture_id,
                        'lecture_name': member_repeat_schedule_info.lecture_tb.name,
                        'lecture_max_member_num': member_repeat_schedule_info.lecture_tb.member_num,
                        'lecture_ing_color_cd': member_repeat_schedule_info.lecture_tb.ing_color_cd,
                        'lecture_end_color_cd': member_repeat_schedule_info.lecture_tb.end_color_cd,
                        'lecture_ing_font_color_cd': member_repeat_schedule_info.lecture_tb.ing_font_color_cd,
                        'lecture_end_font_color_cd': member_repeat_schedule_info.lecture_tb.end_font_color_cd,
                        'member_id': member_repeat_schedule_info.member_ticket_tb.member.member_id,
                        'member_name': member_repeat_schedule_info.member_ticket_tb.member.name,
                        'member_profile_url': member_profile_url

                    }
                    member_repeat_schedule_list.append(member_repeat_schedule)
                lecture_repeat_schedule_list = member_repeat_schedule_list
            else:
                # 수업 껍데기 반복 일정 정보 불러오기
                lecture_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
                    'lecture_tb',
                    'reg_member').filter(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                         en_dis_type=ON_SCHEDULE_TYPE,
                                         lecture_schedule_id__isnull=True
                                         ).exclude(end_date__lt=today).order_by('-reg_dt', 'lecture_tb',
                                                                                'lecture_schedule_id', )

                # 수업에 속한 회원의 반복 일정 정보 불러오기
                lecture_member_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
                    'lecture_tb', 'member_ticket_tb__member', 'reg_member').filter(
                    class_tb_id=class_id, en_dis_type=ON_SCHEDULE_TYPE, lecture_tb_id=lecture_id,
                    lecture_schedule_id__isnull=False).order_by('-reg_dt', 'lecture_tb',
                                                                'lecture_schedule_id')

                for lecture_repeat_schedule_info in lecture_repeat_schedule_data:
                    week_data = lecture_repeat_schedule_info.week_info.split('/')
                    week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

                    mod_member_id = ''
                    mod_member_name = ''
                    if lecture_repeat_schedule_info.mod_member is not None and lecture_repeat_schedule_info.mod_member != '':
                        mod_member_id = lecture_repeat_schedule_info.mod_member_id
                        mod_member_name = lecture_repeat_schedule_info.mod_member.name
                    lecture_member_repeat_schedule_ordered_dict[lecture_repeat_schedule_info.repeat_schedule_id] = {
                        'repeat_schedule_id': lecture_repeat_schedule_info.repeat_schedule_id,
                        'repeat_type_cd': lecture_repeat_schedule_info.repeat_type_cd,
                        'week_info': "/".join(week_data),
                        'start_date': lecture_repeat_schedule_info.start_date,
                        'end_date': lecture_repeat_schedule_info.end_date,
                        'start_time': lecture_repeat_schedule_info.start_time,
                        'end_time': lecture_repeat_schedule_info.end_time,
                        'time_duration': lecture_repeat_schedule_info.time_duration,
                        'state_cd': lecture_repeat_schedule_info.state_cd,
                        'reg_member_id': lecture_repeat_schedule_info.reg_member_id,
                        'reg_member_name': lecture_repeat_schedule_info.reg_member.name,
                        'mod_member_id': mod_member_id,
                        'mod_member_name': mod_member_name,
                        'mod_dt': str(lecture_repeat_schedule_info.mod_dt),
                        'reg_dt': str(lecture_repeat_schedule_info.reg_dt),
                        'lecture_id': lecture_repeat_schedule_info.lecture_tb.lecture_id,
                        'lecture_name': lecture_repeat_schedule_info.lecture_tb.name,
                        'lecture_max_member_num': lecture_repeat_schedule_info.lecture_tb.member_num,
                        # 'lecture_max_member_num_view_flag': lecture_repeat_schedule_info.lecture_tb.member_num_view_flag,
                        'lecture_ing_color_cd': lecture_repeat_schedule_info.lecture_tb.ing_color_cd,
                        'lecture_end_color_cd': lecture_repeat_schedule_info.lecture_tb.end_color_cd,
                        'lecture_ing_font_color_cd': lecture_repeat_schedule_info.lecture_tb.ing_font_color_cd,
                        'lecture_end_font_color_cd': lecture_repeat_schedule_info.lecture_tb.end_font_color_cd,
                        'lecture_member_repeat_schedule_list': []
                    }

                for lecture_member_repeat_schedule_info in lecture_member_repeat_schedule_data:
                    member_profile_url = '/static/common/icon/icon_account.png'
                    if lecture_member_repeat_schedule_info.member_ticket_tb.member.profile_url is not None \
                            and lecture_member_repeat_schedule_info.member_ticket_tb.member.profile_url != '':
                        member_profile_url = lecture_member_repeat_schedule_info.member_ticket_tb.member.profile_url

                    week_data = lecture_member_repeat_schedule_info.week_info.split('/')
                    week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

                    mod_member_id = ''
                    mod_member_name = ''
                    if lecture_member_repeat_schedule_info.mod_member is not None and lecture_member_repeat_schedule_info.mod_member != '':
                        mod_member_id = lecture_member_repeat_schedule_info.mod_member_id
                        mod_member_name = lecture_member_repeat_schedule_info.mod_member.name
                    lecture_member_repeat_schedule_dict = {
                        'repeat_schedule_id': lecture_member_repeat_schedule_info.repeat_schedule_id,
                        'repeat_type_cd': lecture_member_repeat_schedule_info.repeat_type_cd,
                        'week_info': "/".join(week_data),
                        'start_date': lecture_member_repeat_schedule_info.start_date,
                        'end_date': lecture_member_repeat_schedule_info.end_date,
                        'start_time': lecture_member_repeat_schedule_info.start_time,
                        'end_time': lecture_member_repeat_schedule_info.end_time,
                        'time_duration': lecture_member_repeat_schedule_info.time_duration,
                        'state_cd': lecture_member_repeat_schedule_info.state_cd,
                        'reg_member_id': lecture_member_repeat_schedule_info.reg_member_id,
                        'reg_member_name': lecture_member_repeat_schedule_info.reg_member.name,
                        'mod_member_id': mod_member_id,
                        'mod_member_name': mod_member_name,
                        'mod_dt': str(lecture_member_repeat_schedule_info.mod_dt),
                        'reg_dt': str(lecture_member_repeat_schedule_info.reg_dt),
                        'member_id': lecture_member_repeat_schedule_info.member_ticket_tb.member.member_id,
                        'member_name': lecture_member_repeat_schedule_info.member_ticket_tb.member.name,
                        'member_profile_url': member_profile_url
                    }
                    lecture_schedule_id = lecture_member_repeat_schedule_info.lecture_schedule_id
                    try:
                        lecture_member_repeat_schedule_ordered_dict[
                            lecture_schedule_id]['lecture_member_repeat_schedule_list'].append(
                            lecture_member_repeat_schedule_dict)
                    except KeyError:
                        lecture_member_repeat_schedule_dict = None

                lecture_repeat_schedule_list = lecture_member_repeat_schedule_ordered_dict
        else:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'lecture_repeat_schedule_data': lecture_repeat_schedule_list},
                            json_dumps_params={'ensure_ascii': True})


class GetMemberRepeatScheduleView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = request.session.get('class_id', '')
        member_id = self.request.GET.get('member_id', '')
        error = None
        member_repeat_schedule_list = []
        today = datetime.date.today()

        if member_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            # 회원의 반복 일정중 강사가 볼수 있는 수강정보의 일정을 불러오기 위한 query
            query_auth = "select "+ClassMemberTicketTb._meta.get_field('auth_cd').column + \
                         " from "+ClassMemberTicketTb._meta.db_table + \
                         " as B where B."+ClassMemberTicketTb._meta.get_field('member_ticket_tb').column+" = " \
                         "`"+RepeatScheduleTb._meta.db_table+"`.`" + \
                         RepeatScheduleTb._meta.get_field('member_ticket_tb').column + \
                         "` and B.CLASS_TB_ID = " + str(class_id) + \
                         " and B."+ClassMemberTicketTb._meta.get_field('use').column+"="+str(USE)
            member_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
                'member_ticket_tb__member',
                'lecture_tb').filter(
                class_tb_id=class_id, member_ticket_tb__member_id=member_id
            ).annotate(auth_cd=RawSQL(query_auth, [])).filter(auth_cd=AUTH_TYPE_VIEW).exclude(end_date__lt=today).order_by('start_date')

            # 반복일정 정보 셋팅
            week_order = ['SUN', 'MON', 'TUE', 'WED', 'THS', 'FRI', 'SAT']
            week_order = {key: i for i, key in enumerate(week_order)}
            for member_repeat_schedule_info in member_repeat_schedule_data:
                schedule_type = ON_SCHEDULE
                try:
                    lecture_id = member_repeat_schedule_info.lecture_tb.lecture_id
                    lecture_name = member_repeat_schedule_info.lecture_tb.name
                    lecture_max_member_num = member_repeat_schedule_info.lecture_tb.member_num
                    lecture_ing_color_cd = member_repeat_schedule_info.lecture_tb.ing_color_cd
                    lecture_end_color_cd = member_repeat_schedule_info.lecture_tb.end_color_cd
                    lecture_ing_font_color_cd = member_repeat_schedule_info.lecture_tb.ing_font_color_cd
                    lecture_end_font_color_cd = member_repeat_schedule_info.lecture_tb.end_font_color_cd
                    # lecture_max_member_num_view_flag = member_repeat_schedule_info.lecture_tb.member_num_view_flag
                    if member_repeat_schedule_info.lecture_tb.lecture_type_cd != LECTURE_TYPE_ONE_TO_ONE:
                        schedule_type = GROUP_SCHEDULE
                except AttributeError:
                    lecture_id = ''
                    lecture_name = ''
                    lecture_max_member_num = ''
                    lecture_ing_color_cd = ''
                    lecture_end_color_cd = ''
                    lecture_ing_font_color_cd = ''
                    lecture_end_font_color_cd = ''
                    # lecture_max_member_num_view_flag = ''

                week_data = member_repeat_schedule_info.week_info.split('/')
                week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

                mod_member_id = ''
                mod_member_name = ''
                if member_repeat_schedule_info.mod_member is not None and member_repeat_schedule_info.mod_member != '':
                    mod_member_id = member_repeat_schedule_info.mod_member_id
                    mod_member_name = member_repeat_schedule_info.mod_member.name
                member_repeat_schedule = {'repeat_schedule_id': member_repeat_schedule_info.repeat_schedule_id,
                                          'repeat_type_cd': member_repeat_schedule_info.repeat_type_cd,
                                          'week_info': "/".join(week_data),
                                          'start_date': member_repeat_schedule_info.start_date,
                                          'end_date': member_repeat_schedule_info.end_date,
                                          'start_time': member_repeat_schedule_info.start_time,
                                          'end_time': member_repeat_schedule_info.end_time,
                                          'time_duration': member_repeat_schedule_info.time_duration,
                                          'state_cd': member_repeat_schedule_info.state_cd,
                                          'reg_member_id': member_repeat_schedule_info.reg_member_id,
                                          'reg_member_name': member_repeat_schedule_info.reg_member.name,
                                          'mod_member_id': mod_member_id,
                                          'mod_member_name': mod_member_name,
                                          'mod_dt': str(member_repeat_schedule_info.mod_dt),
                                          'reg_dt': str(member_repeat_schedule_info.reg_dt),
                                          'lecture_repeat_schedule_id':
                                              member_repeat_schedule_info.lecture_schedule_id,
                                          'member_id': member_repeat_schedule_info.member_ticket_tb.member.member_id,
                                          'member_name': member_repeat_schedule_info.member_ticket_tb.member.name,
                                          'lecture_id': lecture_id,
                                          'lecture_name': lecture_name,
                                          'lecture_max_member_num': lecture_max_member_num,
                                          'lecture_ing_color_cd': lecture_ing_color_cd,
                                          'lecture_end_color_cd': lecture_end_color_cd,
                                          'lecture_ing_font_color_cd': lecture_ing_font_color_cd,
                                          'lecture_end_font_color_cd': lecture_end_font_color_cd,
                                          # 'lecture_max_member_num_view_flag': lecture_max_member_num_view_flag,
                                          'schedule_type': schedule_type}
                member_repeat_schedule_list.append(member_repeat_schedule)
        else:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'member_repeat_schedule_data': member_repeat_schedule_list},
                            json_dumps_params={'ensure_ascii': True})


class GetTrainerRepeatScheduleView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = request.session.get('class_id', '')
        trainer_id = self.request.GET.get('trainer_id', '')
        error = None
        trainer_repeat_schedule_list = []
        today = datetime.date.today()

        if trainer_id == '':
            error = '강사 정보를 불러오지 못했습니다.'

        if error is None:
            # 회원의 반복 일정중 강사가 볼수 있는 수강정보의 일정을 불러오기 위한 query
            query_auth = "select "+ClassMemberTicketTb._meta.get_field('auth_cd').column + \
                         " from "+ClassMemberTicketTb._meta.db_table + \
                         " as B where B."+ClassMemberTicketTb._meta.get_field('member_ticket_tb').column+" = " \
                         "`"+RepeatScheduleTb._meta.db_table+"`.`" + \
                         RepeatScheduleTb._meta.get_field('member_ticket_tb').column + \
                         "` and B.CLASS_TB_ID = " + str(class_id) + \
                         " and B."+ClassMemberTicketTb._meta.get_field('use').column+"="+str(USE)
            trainer_repeat_schedule_data = RepeatScheduleTb.objects.select_related(
                'member_ticket_tb__member',
                'lecture_tb').filter(
                class_tb_id=class_id, repeat_trainer_id=trainer_id
            ).annotate(auth_cd=RawSQL(query_auth, [])).filter(auth_cd=AUTH_TYPE_VIEW).exclude(end_date__lt=today).order_by('start_date')

            # 반복일정 정보 셋팅
            week_order = ['SUN', 'MON', 'TUE', 'WED', 'THS', 'FRI', 'SAT']
            week_order = {key: i for i, key in enumerate(week_order)}
            for trainer_repeat_schedule_info in trainer_repeat_schedule_data:
                schedule_type = ON_SCHEDULE
                try:
                    lecture_id = trainer_repeat_schedule_info.lecture_tb.lecture_id
                    lecture_name = trainer_repeat_schedule_info.lecture_tb.name
                    lecture_max_member_num = trainer_repeat_schedule_info.lecture_tb.member_num
                    lecture_ing_color_cd = trainer_repeat_schedule_info.lecture_tb.ing_color_cd
                    lecture_end_color_cd = trainer_repeat_schedule_info.lecture_tb.end_color_cd
                    lecture_ing_font_color_cd = trainer_repeat_schedule_info.lecture_tb.ing_font_color_cd
                    lecture_end_font_color_cd = trainer_repeat_schedule_info.lecture_tb.end_font_color_cd
                    # lecture_max_member_num_view_flag = member_repeat_schedule_info.lecture_tb.member_num_view_flag
                    if trainer_repeat_schedule_info.lecture_tb.lecture_type_cd != LECTURE_TYPE_ONE_TO_ONE:
                        schedule_type = GROUP_SCHEDULE
                except AttributeError:
                    lecture_id = ''
                    lecture_name = ''
                    lecture_max_member_num = ''
                    lecture_ing_color_cd = ''
                    lecture_end_color_cd = ''
                    lecture_ing_font_color_cd = ''
                    lecture_end_font_color_cd = ''
                    # lecture_max_member_num_view_flag = ''

                week_data = trainer_repeat_schedule_info.week_info.split('/')
                week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))

                mod_member_id = ''
                mod_member_name = ''
                if trainer_repeat_schedule_info.mod_member is not None and trainer_repeat_schedule_info.mod_member != '':
                    mod_member_id = trainer_repeat_schedule_info.mod_member_id
                    mod_member_name = trainer_repeat_schedule_info.mod_member.name
                trainer_repeat_schedule = {'repeat_schedule_id': trainer_repeat_schedule_info.repeat_schedule_id,
                                           'repeat_type_cd': trainer_repeat_schedule_info.repeat_type_cd,
                                           'week_info': "/".join(week_data),
                                           'start_date': trainer_repeat_schedule_info.start_date,
                                           'end_date': trainer_repeat_schedule_info.end_date,
                                           'start_time': trainer_repeat_schedule_info.start_time,
                                           'end_time': trainer_repeat_schedule_info.end_time,
                                           'time_duration': trainer_repeat_schedule_info.time_duration,
                                           'state_cd': trainer_repeat_schedule_info.state_cd,
                                           'reg_member_id': trainer_repeat_schedule_info.reg_member_id,
                                           'reg_member_name': trainer_repeat_schedule_info.reg_member.name,
                                           'mod_member_id': mod_member_id,
                                           'mod_member_name': mod_member_name,
                                           'mod_dt': str(trainer_repeat_schedule_info.mod_dt),
                                           'reg_dt': str(trainer_repeat_schedule_info.reg_dt),
                                           'lecture_repeat_schedule_id':
                                               trainer_repeat_schedule_info.lecture_schedule_id,
                                           'member_id': trainer_repeat_schedule_info.member_ticket_tb.member.member_id,
                                           'member_name': trainer_repeat_schedule_info.member_ticket_tb.member.name,
                                           'lecture_id': lecture_id,
                                           'lecture_name': lecture_name,
                                           'lecture_max_member_num': lecture_max_member_num,
                                           'lecture_ing_color_cd': lecture_ing_color_cd,
                                           'lecture_end_color_cd': lecture_end_color_cd,
                                           'lecture_ing_font_color_cd': lecture_ing_font_color_cd,
                                           'lecture_end_font_color_cd': lecture_end_font_color_cd,
                                           # 'lecture_max_member_num_view_flag': lecture_max_member_num_view_flag,
                                           'schedule_type': schedule_type}
                trainer_repeat_schedule_list.append(trainer_repeat_schedule)
        else:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'trainer_repeat_schedule_data': trainer_repeat_schedule_list},
                            json_dumps_params={'ensure_ascii': True})


def add_program_closed_date_logic(request):
    class_id = request.session.get('class_id', '')
    member_id = request.POST.get('member_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    repeat_week_type = request.POST.get('week_info', '')
    title = request.POST.get('title', '')
    contents = request.POST.get('contents', '')
    is_member_view = request.POST.get('is_member_view', '1')

    error = None
    schedule_closed_id = ''
    start_date_info = None
    end_date_info = None
    repeat_schedule_date_list = []

    if member_id is None:
        member_id = ''
    if repeat_week_type == '':
        error = '요일을 선택해주세요.'
    if start_date == '':
        error = '시작 날짜를 선택해 주세요.'
    elif end_date == '':
        error = '종료 날짜를 선택해 주세요.'

    if error is None:
        try:
            start_date_info = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date_info = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            if (start_date_info - end_date_info) > datetime.timedelta(days=365):
                error = '1년까지만 불가 일정을 등록할수 있습니다.'
        except ValueError:
            error = '날짜 오류가 발생했습니다.[0]'
        except IntegrityError:
            error = '날짜 오류가 발생했습니다.[1]'
        except TypeError:
            error = '날짜 오류가 발생했습니다.[2]'

    if error is None:
        if member_id != '':
            try:
                MemberTb.objects.get(member_id=member_id)
            except ObjectDoesNotExist:
                error = '회원 정보를 불러오지 못했습니다.'

    # 등록할 날짜 list 가져오기
    if error is None:
        week_order = ['SUN', 'MON', 'TUE', 'WED', 'THS', 'FRI', 'SAT']
        week_order = {key: i for i, key in enumerate(week_order)}
        week_data = repeat_week_type.split('/')
        week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))
        repeat_week_type = "/".join(week_data)
        repeat_schedule_date_list = func_get_repeat_schedule_date_list(repeat_week_type,
                                                                       start_date_info,
                                                                       end_date_info)
        if len(repeat_schedule_date_list) == 0:
            error = '등록할 수 있는 일정이 없습니다.'

    if error is None:
        try:
            with transaction.atomic():
                schedule_closed_info = ScheduleClosedTb(class_tb_id=class_id, member_id=member_id,
                                                        start_date=start_date_info, end_date=end_date_info,
                                                        title=title, contents=contents,
                                                        week_info=repeat_week_type,
                                                        is_member_view=is_member_view, use=USE)

                schedule_closed_info.save()

                schedule_closed_id = schedule_closed_info.schedule_closed_id

                # 반복일정 데이터 등록
                for repeat_schedule_date_info in repeat_schedule_date_list:
                    repeat_schedule_info_date = str(repeat_schedule_date_info).split(' ')[0]

                    closed_date = datetime.datetime.strptime(repeat_schedule_info_date, '%Y-%m-%d')

                    schedule_closed_day = ScheduleClosedDayTb(schedule_closed_tb_id=schedule_closed_id,
                                                              closed_date=closed_date, use=USE)
                    schedule_closed_day.save()

        except ValueError:
            error = '오류가 발생했습니다. [1]'
        except IntegrityError:
            error = '오류가 발생했습니다. [2]'
        except TypeError:
            error = '오류가 발생했습니다. [3]'
        except ValidationError:
            error = '오류가 발생했습니다. [4]'
        except InternalError:
            error = '오류가 발생했습니다. [5]'

    if error is not None:
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info='불가 일정', log_how='추가', use=USE)
        log_data.save()

    return JsonResponse({'schedule_closed_id': str(schedule_closed_id)}, json_dumps_params={'ensure_ascii': True})


def update_program_closed_date_logic(request):
    class_id = request.session.get('class_id', '')
    schedule_closed_id = request.POST.get('schedule_closed_id', '')
    member_id = request.POST.get('member_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    week_info = request.POST.get('week_info', '')
    title = request.POST.get('title', '')
    contents = request.POST.get('contents', '')
    is_member_view = request.POST.get('is_member_view', '')

    error = None
    schedule_closed_info = None
    start_date_info = None
    end_date_info = None

    if member_id is None:
        member_id = ''

    try:
        schedule_closed_info = ScheduleClosedTb.objects.get(schedule_closed_id=schedule_closed_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        if week_info == '':
            week_info = schedule_closed_info.week_info
        if start_date == '':
            start_date = str(schedule_closed_info.start_date)
        if end_date == '':
            end_date = str(schedule_closed_info.end_date)
        if title == '':
            title = schedule_closed_info.title
        if contents == '':
            contents = schedule_closed_info.contents
        if is_member_view == '':
            is_member_view = schedule_closed_info.is_member_view

    if error is None:
        try:
            start_date_info = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date_info = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            if (start_date_info - end_date_info) > datetime.timedelta(days=365):
                error = '1년까지만 불가 일정을 등록할수 있습니다.'
        except ValueError:
            error = '날짜 오류가 발생했습니다.[0]'
        except IntegrityError:
            error = '날짜 오류가 발생했습니다.[1]'
        except TypeError:
            error = '날짜 오류가 발생했습니다.[2]'

    if error is None:
        if member_id != '':
            try:
                MemberTb.objects.get(member_id=member_id)
            except ObjectDoesNotExist:
                error = '회원 정보를 불러오지 못했습니다.'

    try:
        with transaction.atomic():
            schedule_closed_info.week_info = week_info
            schedule_closed_info.start_date = start_date_info
            schedule_closed_info.end_date = end_date_info
            schedule_closed_info.title = title
            schedule_closed_info.contents = contents
            schedule_closed_info.is_member_view = is_member_view

            schedule_closed_info.save()
            schedule_closed_id = schedule_closed_info.schedule_closed_id
    except ValueError:
        error = '오류가 발생했습니다. [1]'
    except IntegrityError:
        error = '오류가 발생했습니다. [2]'
    except TypeError:
        error = '오류가 발생했습니다. [3]'
    except ValidationError:
        error = '오류가 발생했습니다. [4]'
    except InternalError:
        error = '오류가 발생했습니다. [5]'

    if error is not None:
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info='불가 일정', log_how='수정', use=USE)
        log_data.save()

    return JsonResponse({'schedule_closed_id': str(schedule_closed_id)}, json_dumps_params={'ensure_ascii': True})


def delete_program_closed_date_logic(request):
    class_id = request.session.get('class_id', '')
    schedule_closed_id = request.POST.get('schedule_closed_id', '')
    error = None

    if schedule_closed_id == '':
        error = '삭제할 불가 일정을 선택해주세요.'

    if error is None:
        try:
            schedule_close_info = ScheduleClosedTb.objects.get(schedule_closed_id=schedule_closed_id)
            schedule_close_info.delete()
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다. [0]'

    if error is not None:
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info='불가 일정', log_how='삭제', use=USE)
        log_data.save()

    return JsonResponse({'schedule_closed_id': str(schedule_closed_id)}, json_dumps_params={'ensure_ascii': True})


class GetMemberClosedDateListView(LoginRequiredMixin, AccessTestMixin, View):
    def get(self, request):
        class_id = request.session.get('class_id', '')
        member_id = request.GET.get('member_id', '')
        error = None
        today = datetime.date.today()
        member_closed_list = []

        if member_id is None or member_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            member_closed_data = MemberClosedDateHistoryTb.objects.select_related(
                'member_ticket_tb__ticket_tb__class_tb').filter(Q(member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS)
                                                                | Q(member_ticket_tb__state_cd=STATE_CD_HOLDING),
                                                                member_id=member_id,
                                                                member_ticket_tb__ticket_tb__class_tb_id=class_id,
                                                                end_date__gte=today,
                                                                use=USE).order_by('-start_date', '-end_date')

            for member_closed_info in member_closed_data:
                member_closed_reason_type_cd_name = '일시정지'
                member_ticket_id = ''
                if member_closed_info.reason_type_cd == 'HD':
                    if member_closed_info.member_ticket_tb is not None:
                        member_ticket_id = member_closed_info.member_ticket_tb_id
                        member_closed_reason_type_cd_name = member_closed_info.member_ticket_tb.ticket_tb.name

                elif member_closed_info.reason_type_cd == 'PROGRAM_CLOSED':
                    member_closed_reason_type_cd_name = '휴무일 - ' + member_closed_info.member_ticket_tb.ticket_tb.name

                elif member_closed_info.reason_type_cd == 'MEMBER_CLOSED':
                    member_closed_reason_type_cd_name = '회원 불가일정 - '\
                                                        + member_closed_info.member_ticket_tb.ticket_tb.name

                schedule_id = ''
                repeat_schedule_id = ''
                schedule_tb = member_closed_info.schedule_tb
                if schedule_tb is not None and schedule_tb != '':
                    schedule_id = schedule_tb.schedule_id
                    if schedule_tb.repeat_schedule_tb is not None and schedule_tb.repeat_schedule_tb != '':
                        repeat_schedule_id = schedule_tb.repeat_schedule_tb_id
                member_closed_dict = {
                    'member_closed_date_history_id': member_closed_info.member_closed_date_history_id,
                    'member_closed_date_member_ticket_id': member_ticket_id,
                    'member_closed_schedule_id': schedule_id,
                    'member_closed_repeat_schedule_id': repeat_schedule_id,
                    'member_closed_start_date': str(member_closed_info.start_date),
                    'member_closed_end_date': str(member_closed_info.end_date),
                    'member_closed_note': member_closed_info.note,
                    'member_closed_reason_type_cd': member_closed_info.reason_type_cd,
                    'member_closed_reason_type_cd_name': member_closed_reason_type_cd_name,
                    'member_closed_extension_flag': member_closed_info.extension_flag
                }
                member_closed_list.append(member_closed_dict)
        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'member_closed_list': member_closed_list}, json_dumps_params={'ensure_ascii': True})


class GetMemberClosedDateListHistoryView(LoginRequiredMixin, AccessTestMixin, View):
    def get(self, request):
        class_id = request.session.get('class_id', '')
        member_id = request.GET.get('member_id', '')
        page = request.GET.get('page', '')
        error = None
        member_closed_list = []
        # member_closed_list = {'HD': [], 'PROGRAM_CLOSED': [], 'MEMBER_CLOSED': [], 'this_page':0, 'max_page':0}
        this_page = page
        max_page = 0
        if member_id is None or member_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            member_closed_data = MemberClosedDateHistoryTb.objects.select_related(
                'member_ticket_tb__ticket_tb__class_tb').filter(member_id=member_id,
                                                                member_ticket_tb__ticket_tb__class_tb_id=class_id,
                                                                # end_date__gte=today,
                                                                use=USE).order_by('-start_date', '-end_date')
            paginator = Paginator(member_closed_data, SCHEDULE_PAGINATION_COUNTER)
            try:
                member_closed_data = paginator.page(page)
            except EmptyPage:
                member_closed_data = []

            max_page = paginator.num_pages
            member_closed_date_idx = paginator.count - SCHEDULE_PAGINATION_COUNTER*(int(page)-1)

            for member_closed_info in member_closed_data:
                member_closed_reason_type_cd_name = '일시정지'
                member_ticket_id = ''
                if member_closed_info.reason_type_cd == 'HD':
                    if member_closed_info.member_ticket_tb is not None:
                        member_ticket_id = member_closed_info.member_ticket_tb_id
                        member_closed_reason_type_cd_name = member_closed_info.member_ticket_tb.ticket_tb.name

                elif member_closed_info.reason_type_cd == 'PROGRAM_CLOSED':
                    member_closed_reason_type_cd_name = '휴무일 - ' + member_closed_info.member_ticket_tb.ticket_tb.name

                elif member_closed_info.reason_type_cd == 'MEMBER_CLOSED':
                    member_closed_reason_type_cd_name = '회원 불가일정 - '\
                                                        + member_closed_info.member_ticket_tb.ticket_tb.name

                schedule_id = ''
                repeat_schedule_id = ''
                schedule_tb = member_closed_info.schedule_tb
                if schedule_tb is not None and schedule_tb != '':
                    schedule_id = schedule_tb.schedule_id
                    if schedule_tb.repeat_schedule_tb is not None and schedule_tb.repeat_schedule_tb != '':
                        repeat_schedule_id = schedule_tb.repeat_schedule_tb_id
                member_closed_dict = {
                    'member_closed_date_idx': member_closed_date_idx,
                    'member_closed_date_history_id': member_closed_info.member_closed_date_history_id,
                    'member_closed_date_member_ticket_id': member_ticket_id,
                    'member_closed_schedule_id': schedule_id,
                    'member_closed_repeat_schedule_id': repeat_schedule_id,
                    'member_closed_start_date': str(member_closed_info.start_date),
                    'member_closed_end_date': str(member_closed_info.end_date),
                    'member_closed_note': member_closed_info.note,
                    'member_closed_reason_type_cd': member_closed_info.reason_type_cd,
                    'member_closed_reason_type_cd_name': member_closed_reason_type_cd_name,
                    'member_closed_extension_flag': member_closed_info.extension_flag
                }
                member_closed_list.append(member_closed_dict)
                member_closed_date_idx -= 1

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'member_closed_list': member_closed_list, 'this_page': this_page, 'max_page': max_page},
                            json_dumps_params={'ensure_ascii': True})


# ############### ############### ############### ############### ############### ############### ##############
# 회원 기능 ############### ############### ############### ############### ############### ############### ##############
class GetMemberInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_id = request.GET.get('member_id', '')
        error = None
        member_result = {}
        if member_id == '':
            error = '회원 ID를 입력해주세요.'

        if error is None:
            member_result = func_get_member_info(class_id, request.user.id, member_id)
            error = member_result['error']

        if error is not None:
            logger.error(request.user.first_name + ' ' + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse(member_result['member_info'], json_dumps_params={'ensure_ascii': True})


class SearchMemberInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        search_val = request.GET.get('search_val', '')
        error = None
        user_info = None
        member_list = None
        member_result_list = []
        if search_val == '':
            error = '회원 정보를 입력해 주세요.'

        if error is None:
            if len(search_val) < 3:
                error = '3글자 이상 입력해주세요.'
        if error is None:
            try:
                user_info = User.objects.get(username=search_val)
            except ObjectDoesNotExist:
                error = '회원 ID를 확인해 주세요.'

            if error is not None:
                member_list = MemberTb.objects.select_related('user').filter(phone=search_val)
                if len(member_list) == 0:
                    error = '회원 정보를 확인해 주세요.'
                else:
                    error = None

        if error is None:
            if user_info is not None:
                group_type = user_info.groups.filter(user=user_info.id, name='trainee')
                if len(group_type) > 0:
                    member_result = func_get_member_info(class_id, request.user.id, user_info.id)
                    error = member_result['error']
                    member_result_list.append(member_result['member_info'])
            else:
                for member_info in member_list:
                    group_type = member_info.user.groups.filter(user=member_info.user.id, name='trainee')
                    if len(group_type) > 0:
                        member_result = func_get_member_info(class_id, request.user.id, member_info.user.id)
                        member_result_list.append(member_result['member_info'])
        if error is not None:
            logger.error(request.user.first_name + ' ' + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'member_list': member_result_list}, json_dumps_params={'ensure_ascii': True})


class SearchTrainerInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        search_val = request.GET.get('search_val', '')
        error = None
        user_info = None
        member_list = None
        member_result_list = []
        if search_val == '':
            error = '회원 정보를 입력해 주세요.'

        if error is None:
            if len(search_val) < 3:
                error = '3글자 이상 입력해주세요.'
        if error is None:
            try:
                user_info = User.objects.get(username=search_val)
            except ObjectDoesNotExist:
                error = '회원 ID를 확인해 주세요.'

            if error is not None:
                member_list = MemberTb.objects.select_related('user').filter(phone=search_val)
                if len(member_list) == 0:
                    error = '회원 정보를 확인해 주세요.'
                else:
                    error = None

        if error is None:
            if user_info is not None:
                group_type = user_info.groups.filter(user=user_info.id, name='trainer')
                if len(group_type) > 0:
                    member_result = func_get_trainer_info(class_id, user_info.id)
                    error = member_result['error']
                    member_result_list.append(member_result['member_info'])
            else:
                for member_info in member_list:
                    group_type = member_info.user.groups.filter(user=member_info.user.id, name='trainer')
                    if len(group_type) > 0:
                        member_result = func_get_trainer_info(class_id, member_info.user.id)
                        member_result_list.append(member_result['member_info'])
        if error is not None:
            logger.error(request.user.first_name + ' ' + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'member_list': member_result_list}, json_dumps_params={'ensure_ascii': True})


class GetMemberListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        keyword = request.GET.get('keyword', '')
        current_member_data = func_get_member_ing_list(class_id, request.user.id, keyword)
        finish_member_data = func_get_member_end_list(class_id, request.user.id, keyword)
        return JsonResponse({'current_member_data': current_member_data,
                             'finish_member_data': finish_member_data},
                            json_dumps_params={'ensure_ascii': True})


class GetMemberIngListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_sort = request.GET.get('sort_val', SORT_MEMBER_NAME)
        sort_order_by = request.GET.get('sort_order_by', SORT_ASC)
        keyword = request.GET.get('keyword', '')
        current_member_data = func_get_member_ing_list(class_id, request.user.id, keyword)
        finish_member_num = len(func_get_class_member_end_list(class_id, keyword))
        sort_info = int(member_sort)

        if sort_info == SORT_MEMBER_NAME:
            current_member_data = sorted(current_member_data, key=lambda elem: elem['member_name'],
                                         reverse=int(sort_order_by))
        elif sort_info == SORT_REMAIN_COUNT:
            current_member_data = sorted(current_member_data, key=lambda elem: elem['member_ticket_rem_count'],
                                         reverse=int(sort_order_by))
        elif sort_info == SORT_START_DATE:
            current_member_data = sorted(current_member_data, key=lambda elem: elem['start_date'],
                                         reverse=int(sort_order_by))
        elif sort_info == SORT_REG_COUNT:
            current_member_data = sorted(current_member_data, key=lambda elem: elem['member_ticket_reg_count'],
                                         reverse=int(sort_order_by))
        elif sort_info == SORT_END_DATE:
            current_member_data = sorted(current_member_data, key=lambda elem: elem['end_date'],
                                         reverse=int(sort_order_by))
        # context['total_member_num'] = len(member_data)
        # if page != 0:
        #     paginator = Paginator(member_data, 20)  # Show 20 contacts per page
        #     try:
        #         member_data = paginator.page(page)
        #     except EmptyPage:
        #         member_data = None
        #
        # context['member_data'] = member_data
        # end_dt = timezone.now()

        return JsonResponse({'current_member_data': current_member_data, 'finish_member_num': finish_member_num},
                            json_dumps_params={'ensure_ascii': True})


class GetMemberEndListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        # start_dt = timezone.now()
        class_id = self.request.session.get('class_id', '')
        # page = request.GET.get('page', 0)
        member_sort = request.GET.get('sort_val', SORT_MEMBER_NAME)
        sort_order_by = request.GET.get('sort_order_by', SORT_ASC)
        keyword = request.GET.get('keyword', '')

        finish_member_data = func_get_member_end_list(class_id, request.user.id, keyword)
        current_member_num = len(func_get_class_member_ing_list(class_id, keyword))

        sort_info = int(member_sort)
        if sort_info == SORT_MEMBER_NAME:
            finish_member_data = sorted(finish_member_data, key=lambda elem: elem['member_name'],
                                        reverse=int(sort_order_by))
        elif sort_info == SORT_REMAIN_COUNT:
            finish_member_data = sorted(finish_member_data, key=lambda elem: elem['member_ticket_rem_count'],
                                        reverse=int(sort_order_by))
        elif sort_info == SORT_START_DATE:
            finish_member_data = sorted(finish_member_data, key=lambda elem: elem['start_date'],
                                        reverse=int(sort_order_by))
        elif sort_info == SORT_REG_COUNT:
            finish_member_data = sorted(finish_member_data, key=lambda elem: elem['member_ticket_reg_count'],
                                        reverse=int(sort_order_by))
        elif sort_info == SORT_END_DATE:
            finish_member_data = sorted(finish_member_data, key=lambda elem: elem['end_date'],
                                        reverse=int(sort_order_by))

        # context['total_member_num'] = len(member_data)
        # if page != 0:
        #     paginator = Paginator(member_data, 20)  # Show 20 contacts per page
        #     try:
        #         member_data = paginator.page(page)
        #     except EmptyPage:
        #         member_data = None
        #
        # end_dt = timezone.now()
        return JsonResponse({'finish_member_data': finish_member_data, 'current_member_num': current_member_num},
                            json_dumps_params={'ensure_ascii': True})


# 회원수정
def update_member_info_logic(request):
    member_id = request.POST.get('member_id')
    first_name = request.POST.get('first_name', '')
    phone = request.POST.get('phone', '')
    sex = request.POST.get('sex', '')
    birthday_dt = request.POST.get('birthday', '')
    class_id = request.session.get('class_id', '')
    error = None
    member = None

    if member_id == '':
        error = '회원 ID를 확인해 주세요.'

    if error is None:
        try:
            member = MemberTb.objects.get(member_id=member_id)
        except ObjectDoesNotExist:
            error = '회원 ID를 확인해 주세요.'
    if error is None:
        if member.user.is_active:
            error = '회원 정보를 수정할수 없습니다.'
        if str(request.user.id) != str(member.reg_info):
            error = '다른 강사가 등록한 회원은 수정할수 없습니다.'
    if error is None:
        try:
            with transaction.atomic():
                # 회원의 이름을 변경하는 경우 자동으로 ID 변경되도록 설정
                if member.user.first_name != first_name:
                    username = first_name
                    i = 0
                    count = MemberTb.objects.filter(name=username).count()
                    max_range = (100 * (10 ** len(str(count)))) - 1
                    for i in range(0, 100):
                        username += str(random.randrange(0, max_range)).zfill(len(str(max_range)))
                        try:
                            User.objects.get(username=username)
                        except ObjectDoesNotExist:
                            break

                    if i == 100:
                        error = 'ID 변경에 실패했습니다. 다시 시도해주세요.'
                        raise InternalError

                    member.user.username = username
                    member.user.first_name = first_name
                    member.user.save()

                member.name = first_name
                member.phone = phone
                member.sex = sex
                if birthday_dt is not None and birthday_dt != '':
                    member.birthday_dt = birthday_dt
                member.save()

        except ValueError:
            error = '등록 값에 문제가 있습니다.'
        except IntegrityError:
            error = '등록 값에 문제가 있습니다.'
        except TypeError:
            error = '등록 값에 문제가 있습니다.'
        except ValidationError:
            error = '등록 값에 문제가 있습니다.'
        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member.name+' 회원 정보', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 회원 삭제
def delete_member_info_logic(request):
    class_id = request.session.get('class_id', '')
    member_id = request.POST.get('member_id')
    error = None
    class_member_ticket_data = None
    member = None
    if member_id == '':
        error = '회원 ID를 확인해 주세요.'

    if error is None:
        try:
            member = MemberTb.objects.get(member_id=member_id)
        except ObjectDoesNotExist:
            error = '회원 ID를 확인해 주세요.'

    if error is None:
        class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb__ticket_tb').filter(class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW,
                                                  member_ticket_tb__member_id=member_id,
                                                  use=USE)

    if error is None:

        schedule_data = ScheduleTb.objects.filter(class_tb_id=class_id,
                                                  member_ticket_tb__member_id=member_id,
                                                  state_cd=STATE_CD_NOT_PROGRESS)

        finish_schedule_data = ScheduleTb.objects.filter(Q(state_cd=STATE_CD_FINISH) | Q(state_cd=STATE_CD_ABSENCE),
                                                         class_tb_id=class_id,
                                                         member_ticket_tb__member_id=member_id)
        repeat_schedule_data = RepeatScheduleTb.objects.filter(class_tb_id=class_id,
                                                               member_ticket_tb__member_id=member_id)

        try:
            with transaction.atomic():
                # 등록된 일정 정리
                if len(schedule_data) > 0:
                    # 예약된 일정 삭제
                    schedule_data.delete()
                # if len(finish_schedule_data) > 0:
                    # 완료된 일정 비활성화
                    # finish_schedule_data.update(use=UN_USE)
                if len(repeat_schedule_data) > 0:
                    # 반복일정 삭제
                    repeat_schedule_data.delete()

                class_member_ticket_data.update(auth_cd=AUTH_TYPE_DELETE, mod_member_id=request.user.id)

                for class_member_ticket_info in class_member_ticket_data:
                    member_ticket_info = class_member_ticket_info.member_ticket_tb
                    if member_ticket_info.state_cd == STATE_CD_IN_PROGRESS:
                        member_ticket_info.member_ticket_avail_count = 0
                        member_ticket_info.state_cd = STATE_CD_FINISH
                        member_ticket_info.save()

                if len(class_member_ticket_data.filter(member_ticket_tb__member_auth_cd=AUTH_TYPE_VIEW)) == 0:
                    if member.user.is_active == 0 and str(member.reg_info) == str(request.user.id):
                        member.contents = member.user.username + ':' + str(member_id)
                        member.use = UN_USE
                        member.save()

                lecture_member_fix_data = LectureMemberTb.objects.filter(class_tb_id=class_id,
                                                                         member_id=member_id, use=USE)
                lecture_member_fix_data.delete()
        except ValueError:
            error = '등록 값에 문제가 있습니다.'
        except IntegrityError:
            error = '등록 값에 문제가 있습니다.'
        except TypeError:
            error = '등록 값의 형태가 문제 있습니다.'
        except ValidationError:
            error = '등록 값의 형태가 문제 있습니다'
        except InternalError:
            error = '등록 값에 문제가 있습니다.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member.name+' 회원 정보', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')

# ############### ############### ############### ############### ############### ############### ##############


class TrainerMainView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'main_trainer.html'

    def get_context_data(self, **kwargs):
        context = {}
        # context = super(TrainerMainView, self).get_context_data(**kwargs)
        # class_id = self.request.session.get('class_id', '')
        error = None
        func_setting_data_update(self.request, 'trainer')
        if error is not None:
            logger.error(self.request.user.first_name + str(self.request.user.id) + ']' + error)
            messages.error(request, error)
        else:
            logger.info(self.request.user.first_name + '[' + str(self.request.user.id) + '] : login success')

        return context


class CalTotalView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'cal_total.html'

    def get_context_data(self, **kwargs):
        context = super(CalTotalView, self).get_context_data(**kwargs)
        context['holiday'] = HolidayTb.objects.filter(use=USE)
        return context


# 단순화 1:1/그룹/클래스 통합 테스트 180828
class ManageLectureView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'manage_lecture.html'

    def get_context_data(self, **kwargs):
        context = super(ManageLectureView, self).get_context_data(**kwargs)
        return context


class ManageMemberView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'manage_member.html'

    def get_context_data(self, **kwargs):
        context = super(ManageMemberView, self).get_context_data(**kwargs)
        return context


class AttendModeView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'attend_mode.html'

    def get_context_data(self, **kwargs):
        context = super(AttendModeView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        setting_admin_password = '0000'
        setting_attend_class_prev_display_time = 0
        setting_attend_class_after_display_time = 0
        setting_schedule_auto_finish = 0
        current_time = timezone.now()
        check_setting_counter = 0
        setting_data = SettingTb.objects.filter(class_tb_id=class_id, use=USE)

        for setting_info in setting_data:
            if setting_info.setting_type_cd == 'LT_ADMIN_PASSWORD':
                setting_admin_password = setting_info.setting_info
                check_setting_counter += 1
            if setting_info.setting_type_cd == 'LT_ATTEND_CLASS_PREV_DISPLAY_TIME':
                setting_attend_class_prev_display_time = int(setting_info.setting_info)
                check_setting_counter += 1
            if setting_info.setting_type_cd == 'LT_ATTEND_CLASS_AFTER_DISPLAY_TIME':
                setting_attend_class_after_display_time = int(setting_info.setting_info)
                check_setting_counter += 1
            if setting_info.setting_type_cd == 'LT_SCHEDULE_AUTO_FINISH':
                setting_schedule_auto_finish = int(setting_info.setting_info)
                check_setting_counter += 1

        start_date = current_time + datetime.timedelta(minutes=int(setting_attend_class_prev_display_time))
        end_date = current_time - datetime.timedelta(minutes=int(setting_attend_class_after_display_time))
        func_setting_data_update(self.request, 'trainer')
        context = func_get_trainer_attend_schedule(context, class_id, start_date, end_date, current_time)

        context['setting_admin_password'] = setting_admin_password
        context['setting_attend_class_prev_display_time'] = setting_attend_class_prev_display_time
        context['setting_attend_class_after_display_time'] = setting_attend_class_after_display_time
        context['setting_schedule_auto_finish'] = setting_schedule_auto_finish
        if check_setting_counter != 4:
            context['check_setting_data'] = 1

        return context


class AttendModeDetailView(LoginRequiredMixin, AccessTestMixin, View):
    template_name = 'attend_mode_detail.html'

    def post(self, request):
        context = {}
        class_id = request.session.get('class_id', '')
        phone_number = request.POST.get('phone_number', '')
        schedule_id = request.POST.get('schedule_id', '')
        member_ticket_id = request.POST.get('member_ticket_id', '')
        error = None
        if schedule_id is None or schedule_id == '':
            error = '일정 정보를 불러오지 못했습니다.'
        if error is None:
            try:
                schedule_info = ScheduleTb.objects.get(schedule_id=schedule_id)
            except ObjectDoesNotExist:
                schedule_info = error = '일정 정보를 불러오지 못했습니다.'

        if error is None:

            member_data = ClassMemberTicketTb.objects.filter(class_tb_id=class_id,
                                                             member_ticket_tb__member__phone=phone_number,
                                                             auth_cd=AUTH_TYPE_VIEW,
                                                             use=USE)
            if len(member_data) > 0:
                context['member_info'] = member_data[0].member_ticket_tb.member
                member_id = member_data[0].member_ticket_tb.member_id
                context['member_id'] = member_id
                context['schedule_info'] = schedule_info
                context['schedule_id'] = schedule_id
                if schedule_info.member_ticket_tb is None or schedule_info.member_ticket_tb == '':
                    try:
                        member_ticket_schedule_info = ScheduleTb.objects.get(lecture_schedule_id=schedule_id,
                                                                             lecture_tb_id=schedule_info.lecture_tb_id,
                                                                             member_ticket_tb__member_id=member_id)
                        context['member_ticket_info'] = member_ticket_schedule_info.member_ticket_tb
                        context['schedule_info'] = member_ticket_schedule_info
                        context['schedule_id'] = member_ticket_schedule_info.schedule_id

                    except ObjectDoesNotExist:
                        member_ticket_id = None
                        member_ticket_result = func_get_lecture_member_ticket_id(class_id, schedule_info.lecture_tb_id,
                                                                                 member_id)

                        if member_ticket_result['error'] is not None:
                            error = member_ticket_result['error']
                        else:
                            member_ticket_id = member_ticket_result['member_ticket_id']

                        # if error is None:
                        #     if member_ticket_id is None or member_ticket_id == '':
                        #         member_ticket_id = func_get_member_ticket_id_old(class_id, member_id)

                        if error is None:
                            if member_ticket_id is None or member_ticket_id == '':
                                error = '선택하신 수업의 예약 가능 횟수가 없습니다.'
                            else:
                                try:
                                    context['member_ticket_info'] =\
                                        MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
                                except ObjectDoesNotExist:
                                    error = '수강권 정보를 불러오지 못했습니다.'
                else:
                    if schedule_info.member_ticket_tb.member_id == member_id:
                        context['member_ticket_info'] = schedule_info.member_ticket_tb
                    else:
                        error = '번호와 수업이 일치하지 않습니다.'
            else:
                error = '등록되지 않은 전화번호 입니다.'

        if error is not None:
            logger.error('class_id:'+str(class_id) + '/phone_number:' + str(phone_number) + '/schedule_id:'
                         + str(schedule_id) + '/' + error)
            messages.error(request, error)

        return render(request, self.template_name, context)


class BGSettingView(LoginRequiredMixin, AccessTestMixin, View):
    template_name = 'setting_background.html'

    def get(self, request):
        # context = super(BGSettingView, self).get_context_data(**kwargs)
        context = {}
        class_id = request.session.get('class_id', '')
        error = None
        background_img_data = None

        context['common_cd_data'] = CommonCdTb.objects.filter(upper_common_cd='14', use=USE).order_by('order')

        if error is None:
            if class_id is not None and class_id != '':
                background_img_data = BackgroundImgTb.objects.filter(class_tb_id=class_id,
                                                                     use=USE).order_by('-class_tb_id')
            else:
                background_img_data = BackgroundImgTb.objects.filter(class_tb__member_id=request.user.id,
                                                                     use=USE).order_by('-class_tb_id')

        if error is None:
            for background_img_info in background_img_data:
                try:
                    background_img_type_name = \
                        CommonCdTb.objects.get(common_cd=background_img_info.background_img_type_cd)
                except ObjectDoesNotExist:
                    background_img_type_name = None
                if background_img_type_name is not None:
                    background_img_info.background_img_type_name = background_img_type_name.common_cd_nm

        context['background_img_data'] = background_img_data

        return render(request, self.template_name, context)


class ClassSelectView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'trainer_class_select.html'

    def get_context_data(self, **kwargs):
        context = super(ClassSelectView, self).get_context_data(**kwargs)
        return context


def add_program_logic(request):
    template_name = 'trainer_add_program.html'

    # def get(self, request):
    context = {}
    cancel_redirect_url = request.GET.get('cancel_redirect_url', '/login/logout/')
    # context = super(AddClassView, self).get_context_data(**kwargs)
    class_type_cd_data = CommonCdTb.objects.filter(upper_common_cd='02', use=USE).order_by('order')
    for class_type_cd_info in class_type_cd_data:
        class_type_cd_info.subject_type_cd = CommonCdTb.objects.filter(upper_common_cd='03',
                                                                       group_cd=class_type_cd_info.common_cd,
                                                                       use=USE).order_by('order')

    center_list = CenterTrainerTb.objects.filter(member_id=request.user.id, use=USE)

    context['center_list'] = center_list
    context['class_type_cd_data'] = class_type_cd_data
    context['cancel_redirect_url'] = cancel_redirect_url

    return render(request, template_name, context)


class MyPageView(LoginRequiredMixin, AccessTestMixin, View):
    template_name = 'setting_mypage.html'

    def get(self, request):
        context = {}
        # context = super(MyPageView, self).get_context_data(**kwargs)
        class_id = request.session.get('class_id', '')
        sns_id = request.session.get('social_login_id', '')
        error = None
        now = timezone.now()
        next_schedule_start_dt = ''
        next_schedule_end_dt = ''

        today = datetime.date.today()
        month_first_day = today.replace(day=1)
        next_year = int(month_first_day.strftime('%Y')) + 1
        next_month = (int(month_first_day.strftime('%m')) + 1) % 12
        if next_month == 0:
            next_month = 1
        next_month_first_day = month_first_day.replace(month=next_month)

        if next_month == 1:
            next_month_first_day = next_month_first_day.replace(year=next_year)

        end_schedule_num = 0
        new_member_num = 0

        context['total_member_num'] = 0
        context['current_total_member_num'] = 0
        context['end_schedule_num'] = 0
        context['new_member_num'] = 0

        if class_id is None or class_id == '':
            error = '지점 정보를 불러오지 못했습니다.'

        if error is None:
            try:
                class_info = ClassTb.objects.get(class_id=class_id)
                context['center_name'] = class_info.get_center_name()
            except ObjectDoesNotExist:
                error = '지점 정보를 불러오지 못했습니다.'

        if error is None:
            # all_member = MemberTb.objects.filter().order_by('name')
            all_member = func_get_class_member_ing_list(class_id, '')
            end_member = func_get_class_member_end_list(class_id, '')

            for member_info in all_member:
                # member_data = member_info

                # 강좌에 해당하는 수강/회원 정보 가져오기
                total_class_member_ticket_list = ClassMemberTicketTb.objects.select_related(
                    'member_ticket_tb').filter(class_tb_id=class_id, member_ticket_tb__member_id=member_info.member_id,
                                               member_ticket_tb__use=USE, auth_cd=AUTH_TYPE_VIEW,
                                               use=USE).order_by('-member_ticket_tb__start_date')

                if len(total_class_member_ticket_list) > 0:
                    # total_member_num += 1
                    start_date = ''
                    for class_member_ticket_info in total_class_member_ticket_list:
                        member_ticket_info = class_member_ticket_info.member_ticket_tb
                        if member_ticket_info.state_cd == STATE_CD_IN_PROGRESS:
                            # current_total_member_num += 1
                            # for lecture_info_data in class_lecture_list:
                            if start_date == '':
                                start_date = member_ticket_info.start_date
                            else:
                                if start_date > member_ticket_info.start_date:
                                    start_date = member_ticket_info.start_date
                            # break

                    if start_date != '':
                        if month_first_day <= start_date < next_month_first_day:
                            new_member_num += 1

            query_class_auth_cd \
                = "select `AUTH_CD` from CLASS_LECTURE_TB as D" \
                  " where D.LECTURE_TB_ID = `SCHEDULE_TB`.`LECTURE_TB_ID` and D.CLASS_TB_ID = " + str(class_id)
            end_schedule_num += ScheduleTb.objects.select_related(
                'member_ticket_tb', 'lecture_tb').filter(Q(state_cd=STATE_CD_FINISH), class_tb_id=class_id,
                                                         lecture_tb__isnull=True, member_ticket_tb__isnull=False,
                                                         en_dis_type=ON_SCHEDULE_TYPE, use=USE
                                                         ).annotate(class_auth_cd=RawSQL(query_class_auth_cd, [])
                                                                    ).filter(class_auth_cd=AUTH_TYPE_VIEW).count()

            end_schedule_num += ScheduleTb.objects.filter(Q(state_cd=STATE_CD_FINISH), class_tb_id=class_id,
                                                          lecture_tb__isnull=False,
                                                          member_ticket_tb__isnull=True,
                                                          en_dis_type=ON_SCHEDULE_TYPE,
                                                          use=USE).count()
            # 남은 횟수 1개 이상인 경우 - 180314 hk.kim
            context['total_member_num'] = len(all_member) + len(end_member)
            # 남은 횟수 1개 이상 3개 미만인 경우 - 180314 hk.kim
            # context['current_total_member_num'] = current_total_member_num
            context['current_total_member_num'] = len(all_member)
            context['new_member_num'] = new_member_num

        schedule_data = ScheduleTb.objects.filter(class_tb=class_id, en_dis_type=ON_SCHEDULE_TYPE,
                                                  start_dt__gte=now, use=USE).order_by('start_dt')
        if len(schedule_data) > 0:
            next_schedule_start_dt = schedule_data[0].start_dt
            next_schedule_end_dt = schedule_data[0].end_dt

        context['next_schedule_start_dt'] = str(next_schedule_start_dt)
        context['next_schedule_end_dt'] = str(next_schedule_end_dt)
        context['end_schedule_num'] = end_schedule_num
        context['check_password_changed'] = 1
        if sns_id != '' and sns_id is not None:
            sns_password_change_check = SnsInfoTb.objects.filter(member_id=request.user.id, sns_id=sns_id,
                                                                 change_password_check=1, use=USE).count()
            if sns_password_change_check == 0:
                context['check_password_changed'] = 0

        return render(request, self.template_name, context)


class TrainerSettingView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'setting.html'

    def get_context_data(self, **kwargs):
        context = super(TrainerSettingView, self).get_context_data(**kwargs)

        return context


class PushSettingView(LoginRequiredMixin, AccessTestMixin, View):
    template_name = 'setting_push.html'

    def get(self, request):
        context = {}
        # context = super(PushSettingView, self).get_context_data(**kwargs)
        class_id = request.session.get('class_id', '')
        # class_hour = request.session.get('class_hour')
        context = func_get_trainer_setting_list(context, class_id, request.user.id)

        return render(request, self.template_name, context)


class ReserveSettingView(LoginRequiredMixin, AccessTestMixin, View):
    template_name = 'setting_reserve.html'

    def get(self, request):
        context = {}

        return render(request, self.template_name, context)


class BasicSettingView(LoginRequiredMixin, AccessTestMixin, View):
    template_name = 'setting_basic.html'

    def get(self, request):
        context = {}

        return render(request, self.template_name, context)


class ManageWorkView(LoginRequiredMixin, AccessTestMixin, View):
    template_name = 'manage_work.html'

    def get(self, request):
        context = {}
        class_id = request.session.get('class_id', '')
        start_date = request.session.get('sales_start_date', '')
        end_date = request.session.get('sales_end_date', '')

        error = None
        finish_date = None
        month_first_day = None
        if end_date == '' or end_date is None:
            finish_date = timezone.now()
            this_year = int(finish_date.strftime('%Y'))
            next_month = (int(finish_date.strftime('%m'))+1) % 12
            if next_month == 0:
                next_month = 12
            elif next_month == 1:
                this_year += 1
            finish_date = finish_date.replace(month=next_month, day=1)
            finish_date = finish_date - datetime.timedelta(days=1)
            finish_date = finish_date.replace(year=this_year)
        else:
            try:
                finish_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            except TypeError:
                error = '날짜 형식에 문제 있습니다.'
            except ValueError:
                error = '날짜 형식에 문제 있습니다.'

        if start_date == '' or start_date is None:
            month_first_day = finish_date.replace(day=1)
            # default 통계값
            # for i in range(1, 3):
            #     before_year = int(month_first_day.strftime('%Y')) - 1
            #     before_month = (int(month_first_day.strftime('%m')) - 1) % 13
            #     if before_month == 0:
            #         before_month = 12
            #     before_month_first_day = month_first_day.replace(month=before_month)
            #     if before_month == 12:
            #         before_month_first_day = before_month_first_day.replace(year=before_year)
            #     month_first_day = before_month_first_day
        else:
            try:
                month_first_day = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                month_first_day = month_first_day.replace(day=1)
            except TypeError:
                error = '날짜 형식에 문제 있습니다.'
            except ValueError:
                error = '날짜 형식에 문제 있습니다.'

        # if error is None:
        #     context = get_stats_member_data(class_id, month_first_day, finish_date)
        #     error = context['error']

        if error is None:
            sales_data_result = get_sales_data(class_id, month_first_day, finish_date)
            if sales_data_result['error'] is None:
                context['month_price_data'] = sales_data_result['month_price_data']
            else:
                error = sales_data_result['error']

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return render(request, self.template_name, context)


class AlarmView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        page = request.GET.get('this_page', 1)
        class_id = self.request.session.get('class_id', '')
        error = None

        today = datetime.date.today()
        # three_days_ago = today - datetime.timedelta(days=3)
        ordered_alarm_dict = collections.OrderedDict()
        ordered_alarm_data_dict = collections.OrderedDict()
        alarm_data = LogTb.objects.select_related('auth_member').filter(class_tb_id=class_id,
                                                                        # reg_dt__gte=three_days_ago,
                                                                        use=USE).order_by('-reg_dt')

        query = "select count(B.ID) from QA_COMMENT_TB as B where B.QA_TB_ID = `QA_TB`.`ID` and B.READ=0 and B.USE=1"

        ordered_alarm_dict['check_qa_comment'] = QATb.objects.filter(
            member_id=request.user.id, status_type_cd='QA_COMPLETE',
            use=USE).annotate(qa_comment=RawSQL(query, [])).filter(qa_comment__gt=0).count()
        paginator = Paginator(alarm_data, ALARM_PAGINATION_COUNTER)
        try:
            alarm_data = paginator.page(page)
        except EmptyPage:
            alarm_data = []

        ordered_alarm_dict['max_page'] = paginator.num_pages
        ordered_alarm_dict['this_page'] = page

        if error is None:
            temp_alarm_date = None
            date_alarm_list = []
            for alarm_info in alarm_data:
                # 날짜 셋팅
                alarm_date = str(alarm_info.reg_dt).split(' ')[0]
                # 새로운 날짜로 넘어간 경우 array 비워주고 값 셋팅
                if temp_alarm_date != alarm_date:
                    temp_alarm_date = alarm_date
                    date_alarm_list = []

                if alarm_info.read == 0:
                    alarm_info.log_read = 0
                    alarm_info.read = 1
                    alarm_info.save()
                elif alarm_info.read == 1:
                    alarm_info.log_read = 1
                else:
                    alarm_info.log_read = 2
                alarm_info.time_ago = timezone.now() - alarm_info.reg_dt
                alarm_info.reg_dt = str(alarm_info.reg_dt).split('.')[0]

                if alarm_info.log_detail is not None and alarm_info.log_detail != '':
                    log_detail_split = str(alarm_info.log_detail).split('/')
                    before_day = log_detail_split[0]
                    after_day = log_detail_split[1]
                    if '반복 일정' in alarm_info.log_how or '반복 일정' in alarm_info.log_info:
                        alarm_info.log_detail = before_day + '~' + after_day
                    else:
                        after_day_split = after_day.split(' ')
                        if len(after_day_split) > 1:
                            alarm_info.log_detail = before_day + '~' + after_day.split(' ')[1]
                        else:
                            alarm_info.log_detail = before_day + '~' + after_day
                        # log_info.log_detail = before_day + '~' + after_day.split(' ')[1]
                        if len(log_detail_split) > 2:
                            alarm_info.log_detail = before_day + '~' + after_day + '~' + log_detail_split[2]

                day = int(alarm_info.time_ago.days)
                hour = int(alarm_info.time_ago.seconds / 3600)
                minute = int(alarm_info.time_ago.seconds / 60)
                sec = int(alarm_info.time_ago.seconds)

                if day > 0:
                    alarm_info.time_ago = str(day) + '일 전'
                else:
                    if hour > 0:
                        alarm_info.time_ago = str(hour) + '시간 전'
                    else:
                        if minute > 0:
                            alarm_info.time_ago = str(minute) + '분 전'
                        else:
                            alarm_info.time_ago = str(sec) + '초 전'

                # array 에 값을 추가후 dictionary 에 추가
                date_alarm_list.append({'alarm_id': str(alarm_info.log_id),
                                        'alarm_info': alarm_info.log_info,
                                        'alarm_from_member_name': alarm_info.from_member_name,
                                        'alarm_to_member_name': alarm_info.to_member_name,
                                        'alarm_how': alarm_info.log_how,
                                        'alarm_detail': alarm_info.log_detail,
                                        'time_ago': alarm_info.time_ago,
                                        'read_check': alarm_info.log_read,
                                        'reg_dt': alarm_info.reg_dt,
                                        'reg_member_name': alarm_info.auth_member.name})
                ordered_alarm_data_dict[alarm_date] = date_alarm_list
            ordered_alarm_dict['alarm_data'] = ordered_alarm_data_dict
        return JsonResponse(ordered_alarm_dict, json_dumps_params={'ensure_ascii': True})


# ############### ############### ############### ############### ############### ############### ##############

def export_excel_member_list_logic(request):
    class_id = request.session.get('class_id', '')
    finish_flag = request.GET.get('finish_flag', '0')
    member_sort = request.GET.get('sort_val', SORT_MEMBER_NAME)
    sort_order_by = request.GET.get('sort_order_by', SORT_ASC)
    keyword = request.GET.get('keyword', '')
    sort_info = int(member_sort)

    error = None
    member_list = []

    wb = Workbook()
    ws1 = wb.active
    start_raw = 3

    ws1['A2'] = '회원명'
    ws1['B2'] = '연락처'
    ws1['C2'] = '회원 ID'
    # ws1['D2'] = '등록 횟수'
    # ws1['E2'] = '남은 횟수'
    # ws1['F2'] = '시작 일자'
    # ws1['G2'] = '종료 일자'
    # ws1['H2'] = '연락처'
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    # ws1.column_dimensions['D'].width = 10
    # ws1.column_dimensions['E'].width = 10
    # ws1.column_dimensions['F'].width = 15
    # ws1.column_dimensions['G'].width = 15
    # ws1.column_dimensions['H'].width = 20
    filename_temp = request.user.first_name + '님_'
    if finish_flag == '0':
        member_list = func_get_member_ing_list(class_id, request.user.id, keyword)

        # if sort_info == SORT_MEMBER_NAME:
        #     member_list = sorted(member_list, key=attrgetter('name'), reverse=int(sort_order_by))
        # elif sort_info == SORT_REMAIN_COUNT:
        #     member_list = sorted(member_list, key=attrgetter('member_ticket_rem_count'), reverse=int(sort_order_by))
        # elif sort_info == SORT_START_DATE:
        #     member_list = sorted(member_list, key=attrgetter('start_date'), reverse=int(sort_order_by))
        # elif sort_info == SORT_REG_COUNT:
        #     member_list = sorted(member_list, key=attrgetter('member_ticket_reg_count'), reverse=int(sort_order_by))

        filename_temp += '진행중_회원목록'
        ws1.title = "진행중 회원"
        ws1['A1'] = '진행중 회원정보'
        ws1['A1'].font = Font(bold=True, size=15)
        for member_info in member_list:

            ws1['A' + str(start_raw)] = member_info['member_name']
            # ws1['B' + str(start_raw)] = member_info.lecture_info
            ws1['B' + str(start_raw)] = member_info['member_phone']
            ws1['C' + str(start_raw)] = member_info['member_user_id']
            # ws1['D' + str(start_raw)] = member_info.member_ticket_reg_count
            # ws1['E' + str(start_raw)] = member_info.member_ticket_rem_count
            # ws1['F' + str(start_raw)] = member_info.start_date
            # if member_info.end_date == '9999-12-31':
            #     ws1['G' + str(start_raw)] = '소진시까지'
            # else:
            #     ws1['G' + str(start_raw)] = member_info.end_date
            #
            # if member_info.phone is None:
            #     ws1['H' + str(start_raw)] = '---'
            # else:
            #     ws1['H' + str(start_raw)] = member_info.phone[0:3] + '-' + member_info.phone[3:7]\
            #                                 + '-' + member_info.phone[7:]
            start_raw += 1
    else:
        member_finish_list = func_get_member_end_list(class_id, request.user.id, keyword)

        if sort_info == SORT_MEMBER_NAME:
            member_finish_list = sorted(member_finish_list, key=attrgetter('name'), reverse=int(sort_order_by))
        elif sort_info == SORT_REMAIN_COUNT:
            member_finish_list = sorted(member_finish_list, key=attrgetter('member_ticket_rem_count'),
                                        reverse=int(sort_order_by))
        elif sort_info == SORT_START_DATE:
            member_finish_list = sorted(member_finish_list, key=attrgetter('start_date'), reverse=int(sort_order_by))
        elif sort_info == SORT_REG_COUNT:
            member_finish_list = sorted(member_list, key=attrgetter('member_ticket_reg_count'),
                                        reverse=int(sort_order_by))

        ws1.title = "종료된 회원"
        filename_temp += '종료된_회원목록'
        ws1['A1'] = '종료된 회원정보'
        ws1['A1'].font = Font(bold=True, size=15)
        for member_info in member_finish_list:
            ws1['A' + str(start_raw)] = member_info.name
            ws1['B' + str(start_raw)] = member_info.lecture_info
            ws1['C' + str(start_raw)] = member_info.user.username
            ws1['D' + str(start_raw)] = member_info.member_ticket_reg_count
            ws1['E' + str(start_raw)] = member_info.member_ticket_rem_count
            ws1['F' + str(start_raw)] = member_info.start_date
            if member_info.end_date == '9999-12-31':
                ws1['G' + str(start_raw)] = '소진시까지'
            else:
                ws1['G' + str(start_raw)] = member_info.end_date
            if member_info.phone is None:
                ws1['H' + str(start_raw)] = '---'
            else:
                ws1['H' + str(start_raw)] = member_info.phone[0:3] + '-' + member_info.phone[3:7]\
                                            + '-' + member_info.phone[7:]
            start_raw += 1

    user_agent = request.META['HTTP_USER_AGENT']
    filename_temp += '.xlsx'
    filename = filename_temp.encode('utf-8')
    response = HttpResponse(save_virtual_workbook(wb),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if 'chrome' in str(user_agent) or 'Chrome' in str(user_agent):
        response['Content-Disposition'] = 'attachment; filename="' + quote(filename) + '"'
    elif 'safari' in str(user_agent) or 'Safari' in str(user_agent):
        response['Content-Disposition'] = 'attachment; filename="' + quote(filename) + '"'
    elif 'firefox' in str(user_agent) or 'Firefox' in str(user_agent):
        response['Content-Disposition'] = 'attachment; filename*="' + quote(filename) + '"'
    else:
        response['Content-Disposition'] = 'attachment; filename="' + quote(filename) + '"'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
    return response


def export_excel_member_info_logic(request):
    class_id = request.session.get('class_id', '')
    member_id = request.GET.get('member_id', '')

    error = None
    member_info = None
    member_ticket_counts = 0
    np_member_ticket_counts = 0
    class_member_ticket_data = None

    if class_id is None or class_id == '':
        error = '오류가 발생했습니다.'

    if member_id is None or member_id == '':
        error = '회원 정보를 불러오지 못했습니다.'

    wb = Workbook()
    ws1 = wb.active

    if error is None:
        try:
            member_info = MemberTb.objects.get(member_id=member_id)
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다.'

    # 수강 정보 불러 오기
    if error is None:
        class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb__ticket_tb'
        ).filter(class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW, member_ticket_tb__member_id=member_id,
                 member_ticket_tb__use=USE,
                 use=USE).order_by('-member_ticket_tb__start_date', 'member_ticket_tb__reg_dt')

    if error is None:
        # 강사 클래스의 반복일정 불러오기
        if len(class_member_ticket_data) > 0:

            for idx, class_member_ticket_info in enumerate(class_member_ticket_data):
                member_ticket_info = class_member_ticket_info.member_ticket_tb
                member_ticket_info.start_date = str(member_ticket_info.start_date)
                member_ticket_info.end_date = str(member_ticket_info.end_date)
                member_ticket_info.mod_dt = str(member_ticket_info.mod_dt)
                member_ticket_info.reg_dt = str(member_ticket_info.reg_dt)

                start_raw = 7
                ws1.title = member_ticket_info.start_date + ' 수강권 정보'
                ws1['A1'] = '수강권 정보'
                ws1['A1'].font = Font(bold=True, size=15)
                ws1['A2'] = '수강권'
                ws1['B2'] = '시작일자'
                ws1['C2'] = '종료일자'
                ws1['D2'] = '등록횟수'
                ws1['E2'] = '남은횟수'
                ws1['F2'] = '등록금액'
                ws1['G2'] = '진행상태'
                ws1['H2'] = '회원님과 연결상태'
                ws1['I2'] = '특이사항'

                ws1.column_dimensions['A'].width = 15
                ws1.column_dimensions['B'].width = 20
                ws1.column_dimensions['C'].width = 20
                ws1.column_dimensions['D'].width = 25
                ws1.column_dimensions['E'].width = 25
                ws1.column_dimensions['F'].width = 15
                ws1.column_dimensions['G'].width = 10
                ws1.column_dimensions['H'].width = 10
                ws1.column_dimensions['I'].width = 20

                try:
                    member_ticket_info.state_cd_name = CommonCdTb.objects.get(common_cd=member_ticket_info.state_cd)
                except ObjectDoesNotExist:
                    error = '수강권 정보를 불러오지 못했습니다.'
                try:
                    member_ticket_test = MemberTicketTb.objects.get(
                        member_ticket_id=member_ticket_info.member_ticket_id)
                    member_ticket_info.auth_cd = member_ticket_test.member_auth_cd
                except ObjectDoesNotExist:
                    error = '수강권 정보를 불러오지 못했습니다.'

                try:
                    member_ticket_info.auth_cd_name = CommonCdTb.objects.get(common_cd=member_ticket_info.auth_cd)
                except ObjectDoesNotExist:
                    error = '수강권 정보를 불러오지 못했습니다.'

                if member_ticket_info.auth_cd == AUTH_TYPE_WAIT:
                    np_member_ticket_counts += 1
                member_ticket_counts += 1

                class_member_ticket_info.lecture_info = ''

                if '\r\n' in member_ticket_info.note:
                    member_ticket_info.note = member_ticket_info.note.replace('\r\n', ' ')

                ws1['A3'] = class_member_ticket_info.lecture_info
                ws1['B3'] = member_ticket_info.start_date
                ws1['C3'] = member_ticket_info.end_date
                ws1['D3'] = member_ticket_info.member_ticket_reg_count
                ws1['E3'] = member_ticket_info.member_ticket_rem_count
                ws1['F3'] = member_ticket_info.price
                ws1['G3'] = member_ticket_info.state_cd_name.common_cd_nm
                ws1['H3'] = member_ticket_info.auth_cd_name.common_cd_nm
                ws1['I3'] = member_ticket_info.note

                ws1['A5'] = '반복 일정'
                ws1['A5'].font = Font(bold=True, size=15)
                ws1['A6'] = '빈도'
                ws1['B6'] = '요일'
                ws1['C6'] = '시작시각 ~ 종료시각'
                ws1['D6'] = '시작일 ~ 종료일'
                repeat_schedule_data = RepeatScheduleTb.objects.filter(
                    member_ticket_tb_id=member_ticket_info.member_ticket_id,
                    en_dis_type=ON_SCHEDULE_TYPE).order_by('-start_date')

                # if repeat_schedule_data is not None and len(repeat_schedule_data) > 0:
                for repeat_schedule_info in repeat_schedule_data:
                    if repeat_schedule_info.repeat_type_cd == 'WW':
                        ws1['A' + str(start_raw)] = '매주'
                    else:
                        ws1['A' + str(start_raw)] = '격주'
                    week_info = repeat_schedule_info.week_info.replace('MON', '월')
                    week_info = week_info.replace('TUE', '화')
                    week_info = week_info.replace('WED', '수')
                    week_info = week_info.replace('THS', '목')
                    week_info = week_info.replace('FRI', '금')
                    week_info = week_info.replace('SAT', '토')
                    week_info = week_info.replace('SUN', '일')
                    repeat_start_date = str(repeat_schedule_info.start_date)
                    repeat_end_date = str(repeat_schedule_info.start_date)
                    ws1['B' + str(start_raw)] = week_info

                    ws1['C' + str(start_raw)] = repeat_schedule_info.start_time + '~' + repeat_schedule_info.end_time

                    ws1['D' + str(start_raw)] = repeat_start_date + '~' + repeat_end_date

                    start_raw += 1

                start_raw += 1
                ws1['A' + str(start_raw)] = '레슨 이력'
                ws1['A' + str(start_raw)].font = Font(bold=True, size=15)
                start_raw += 1
                ws1['A' + str(start_raw)] = '회차'
                ws1['B' + str(start_raw)] = '수행 일시'
                ws1['C' + str(start_raw)] = '진행시간'
                ws1['D' + str(start_raw)] = '구분'
                ws1['E' + str(start_raw)] = '메모'

                start_raw += 1
                pt_schedule_data = ScheduleTb.objects.filter(member_ticket_tb_id=member_ticket_info.member_ticket_id,
                                                             en_dis_type=ON_SCHEDULE_TYPE,
                                                             use=USE).order_by('-start_dt')

                if pt_schedule_data is not None and len(pt_schedule_data) > 0:
                    schedule_idx = len(pt_schedule_data)
                    for pt_schedule_info in pt_schedule_data:

                        ws1['A' + str(start_raw)] = str(schedule_idx)
                        start_date_temp = str(pt_schedule_info.start_dt).split(':')
                        ws1['B' + str(start_raw)] = start_date_temp[0] + ':' + start_date_temp[1]

                        time_duration_temp = pt_schedule_info.end_dt - pt_schedule_info.start_dt
                        time_duration = str(time_duration_temp).split(':')
                        time_duration_str = ''
                        if time_duration[0] != '00' and time_duration[0] != '0':
                            time_duration_str += time_duration[0] + '시간'
                        if time_duration[1] != '00' and time_duration[1] != '0':
                            time_duration_str += time_duration[1] + '분'

                        ws1['C' + str(start_raw)] = time_duration_str
                        if pt_schedule_info.state_cd == STATE_CD_FINISH:
                            ws1['D' + str(start_raw)] = '완료'
                        elif pt_schedule_info.state_cd == STATE_CD_ABSENCE:
                            ws1['D' + str(start_raw)] = '결석'
                        else:
                            ws1['D' + str(start_raw)] = '시작전'

                        if pt_schedule_info.note is None:
                            ws1['E' + str(start_raw)] = ''
                        else:
                            ws1['E' + str(start_raw)] = pt_schedule_info.note
                        start_raw += 1
                        schedule_idx -= 1

                ws1 = wb.create_sheet()

    user_agent = request.META['HTTP_USER_AGENT']
    filename = str(member_info.name + '_회원님_회원정보.xlsx').encode('utf-8')
    # test_str = urllib.parse.unquote('한글')
    response = HttpResponse(save_virtual_workbook(wb),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if 'chrome' in str(user_agent) or 'Chrome' in str(user_agent):
        response['Content-Disposition'] = 'attachment; filename="' + urllib.parse.quote(filename) + '"'
    elif 'safari' in str(user_agent) or 'Safari' in str(user_agent):
        response['Content-Disposition'] = 'attachment; filename="' + urllib.parse.quote(filename) + '"'
    elif 'firefox' in str(user_agent) or 'Firefox' in str(user_agent):
        response['Content-Disposition'] = 'attachment; filename*="' + urllib.parse.quote(filename) + '"'
    else:
        response['Content-Disposition'] = 'attachment; filename="' + urllib.parse.quote(filename) + '"'
    # filename="'+test_str+'.xlsx"'
    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename=.xlsx'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
    return response


# ############### ############### ############### ############### ############### ############### ##############
# 수강권 기능 ############### ############### ############### ############### ############### ############### ########

class GetMemberTicketInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_ticket_id = request.GET.get('member_ticket_id', '')
        error = None
        member_ticket_list = collections.OrderedDict()

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if member_ticket_id is None or member_ticket_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            member_ticket_list = func_get_member_ticket_info(class_id, member_ticket_id)

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        return JsonResponse(member_ticket_list, json_dumps_params={'ensure_ascii': True})


class GetMemberTicketListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_id = request.GET.get('member_id', '')
        error = None
        member_ticket_list = collections.OrderedDict()

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if member_id is None or member_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            member_ticket_list = func_get_member_ticket_list(class_id, member_id)

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        return JsonResponse(member_ticket_list, json_dumps_params={'ensure_ascii': True})


class GetMemberLectureListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_id = request.GET.get('member_id', '')
        error = None
        member_lecture_list = {}

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if member_id is None or member_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            member_lecture_list = func_get_member_lecture_list(class_id, member_id)

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse(member_lecture_list, json_dumps_params={'ensure_ascii': True})


# 수강정보 추가
def add_member_ticket_info_logic(request):
    member_id = request.POST.get('member_id')
    contents = request.POST.get('contents', '')
    counts = request.POST.get('counts')
    price = request.POST.get('price', '')
    payment_price = request.POST.get('payment_price', '')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    ticket_id = request.POST.get('ticket_id', '')
    # NONE : 선택 안함 / CASH : 현금 , CARD : 카드 , TRANS : 계좌 이체 , CASH+CARD : 현금 + 카드, CARD + TRANS : 카드 + 계좌 이체, CASH + TRANS : 현금 + 계좌 이체
    pay_method = request.POST.get('pay_method', 'NONE')
    class_id = request.session.get('class_id', '')
    member_name = ''
    ticket_name= ''
    error = None

    if member_id is None or member_id == '':
        error = '오류가 발생했습니다.[0]'

    if pay_method == '':
        pay_method = 'NONE'

    if price == '':
        price = 0

    if payment_price == '':
        payment_price = 0

    if int(payment_price) > int(price):
        error = '수강권 가격보다 납부 금액이 많습니다.'

    if end_date == '':
        end_date = '9999-12-31'

    # if start_date > end_date:
    #     error = '종료일은 시작일보다 빠를수 없습니다.'
    if counts == '':
        error = '등록 횟수를 입력해 주세요.'
    elif start_date == '':
        error = '시작 일자를 입력해 주세요.'

    if ticket_id == '':
        error = '수강권을 선택해 주세요.'

    if error is None:
        try:
            member_info = MemberTb.objects.get(member_id=member_id)
            member_name = member_info.name
        except ObjectDoesNotExist:
            error = '가입되지 않은 회원입니다.'
    if error is None:
        try:
            ticket_info = TicketTb.objects.get(ticket_id=ticket_id)
            ticket_name = ticket_info.name
        except ObjectDoesNotExist:
            error = '수강권 정보를 확인해주세요.'
    if error is None:
        error = func_add_member_ticket_info(request.user.id, class_id, ticket_id, counts, price, payment_price,
                                            pay_method, start_date, end_date, contents, member_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_name + ' 회원님에게 '
                                  + ticket_name + ' 수강권',  log_how='추가', use=USE)
        log_data.save()
    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 수정
def update_member_ticket_info_logic(request):
    member_ticket_id = request.POST.get('member_ticket_id', '')
    note = request.POST.get('note', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    price = request.POST.get('price', '')
    payment_price = request.POST.get('payment_price', '')
    refund_price = request.POST.get('refund_price', '')
    refund_date = request.POST.get('refund_date', '')
    member_ticket_reg_count = request.POST.get('member_ticket_reg_count', '')
    # NONE : 선택 안함 / CASH : 현금 , CARD : 카드 , TRANS : 계좌 이체 , CASH+CARD : 현금 + 카드,
    # CARD + TRANS : 카드 + 계좌 이체, CASH + TRANS : 현금 + 계좌 이체
    pay_method = request.POST.get('pay_method', 'NONE')
    class_id = request.session.get('class_id', '')
    error = None
    member_ticket_info = None

    if member_ticket_id is None or member_ticket_id == '':
        error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            member_ticket_info = MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        if note is None or note == '':
            note = member_ticket_info.note
        if start_date is None or start_date == '':
            start_date = member_ticket_info.start_date
        if end_date is None or end_date == '':
            end_date = member_ticket_info.end_date
        if price is None or price == '':
            price = member_ticket_info.price
        if payment_price is None or payment_price == '':
            payment_price = member_ticket_info.payment_price
        if pay_method is None or pay_method == '':
            pay_method = member_ticket_info.pay_method
        if refund_price is None or refund_price == '':
            refund_price = member_ticket_info.refund_price
        if refund_date is None or refund_date == '':
            refund_date = member_ticket_info.refund_date
        if member_ticket_reg_count is None or member_ticket_reg_count == '':
            member_ticket_reg_count = member_ticket_info.member_ticket_reg_count
        #
        # if start_date > end_date:
        #     error = '종료일은 시작일보다 빠를수 없습니다.'

        try:
            price = int(price)
        except ValueError:
            error = '수강권 금액은 숫자만 입력 가능합니다.'

        try:
            payment_price = int(payment_price)
        except ValueError:
            error = '수강권 금액은 숫자만 입력 가능합니다.'

        if error is None:
            if payment_price > price:
                error = '수강권 가격보다 납부 금액이 많습니다.'

        try:
            member_ticket_reg_count = int(member_ticket_reg_count)
        except ValueError:
            error = '등록 횟수는 숫자만 입력 가능합니다.'

        if refund_price != '' and refund_price is not None:
            try:
                refund_price = int(refund_price)
            except ValueError:
                error = '환불 금액은 숫자만 입력 가능합니다.'

        if refund_date != '' and refund_date is not None:
            try:
                refund_date = datetime.datetime.strptime(str(refund_date), '%Y-%m-%d')
            except ValueError:
                error = '환불 날짜 오류가 발생했습니다.'
            except TypeError:
                error = '환불 날짜 오류가 발생했습니다.'

    finish_schedule_count = 0
    reserve_schedule_count = 0

    if error is None:
        schedule_list = ScheduleTb.objects.filter(member_ticket_tb=member_ticket_id, use=USE)
        if len(schedule_list) > 0:
            reserve_schedule_count = schedule_list.count()
            finish_schedule_count = schedule_list.filter(Q(state_cd=STATE_CD_FINISH) |
                                                         Q(state_cd=STATE_CD_ABSENCE)).count()

        if member_ticket_reg_count < reserve_schedule_count:
            error = '수정할 등록 횟수가 이미 등록한 일정의 횟수보다 적습니다.'

    if error is None:
        member_ticket_info.start_date = start_date
        member_ticket_info.end_date = end_date
        member_ticket_info.price = price
        member_ticket_info.payment_price = payment_price
        member_ticket_info.pay_method = pay_method
        member_ticket_info.refund_price = refund_price
        member_ticket_info.refund_date = refund_date
        member_ticket_info.note = note
        member_ticket_info.member_ticket_reg_count = member_ticket_reg_count
        member_ticket_info.member_ticket_rem_count = member_ticket_reg_count - finish_schedule_count
        member_ticket_info.member_ticket_avail_count = member_ticket_reg_count - reserve_schedule_count
        if member_ticket_info.state_cd == STATE_CD_FINISH:
            member_ticket_info.member_ticket_rem_count = 0
            member_ticket_info.member_ticket_avail_count = 0
        if member_ticket_info.state_cd == STATE_CD_REFUND:
            member_ticket_info.member_ticket_avail_count = 0
        member_ticket_info.save()
    if error is None:
        member_payment_history_data = MemberPaymentHistoryTb.objects.filter(member_ticket_tb_id=member_ticket_id,
                                                                            payment_price__gt=0)
        member_payment_history_data.update(pay_method=pay_method)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    else:
        log_data = LogTb(log_type='LC02', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_ticket_info.member.name + ' 회원님의 '
                                  + member_ticket_info.ticket_tb.name + ' 수강권 정보', log_how='변경', use=USE)
        log_data.save()
    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 삭제
def delete_member_ticket_info_logic(request):
    member_ticket_id = request.POST.get('member_ticket_id', '')
    class_id = request.session.get('class_id', '')
    error = None
    member_ticket_info = None
    if member_ticket_id is None or member_ticket_id == '':
        error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            member_ticket_info = MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        error = func_delete_member_ticket_info(request.user.id, class_id, member_ticket_id)

    if error is None:
        # 회원의 고정 수업 정리
        func_update_lecture_member_fix_status_cd(class_id, member_ticket_info.member_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_ticket_info.member.name + ' 회원님의 '
                                  + member_ticket_info.ticket_tb.name + ' 수강권', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetHoldMemberTicketListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        member_ticket_id = request.GET.get('member_ticket_id', '')
        error = None
        hold_member_ticket_list = []

        if member_ticket_id is None or member_ticket_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            hold_member_ticket_data = MemberClosedDateHistoryTb.objects.filter(member_ticket_tb_id=member_ticket_id,
                                                                               reason_type_cd='HD',
                                                                               use=USE)
            for hold_member_ticket_info in hold_member_ticket_data:
                hold_member_ticket_dict = {
                    'member_closed_date_history_id': hold_member_ticket_info.member_closed_date_history_id,
                    'member_ticket_hold_start_date': hold_member_ticket_info.start_date,
                    'member_ticket_hold_end_date': hold_member_ticket_info.end_date,
                    'member_ticket_hold_note': hold_member_ticket_info.note,
                    'member_ticket_hold_extension_flag': hold_member_ticket_info.extension_flag
                    }
                hold_member_ticket_list.append(hold_member_ticket_dict)

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'hold_member_ticket_list':hold_member_ticket_list},
                            json_dumps_params={'ensure_ascii': True})


# 수강권 일시정지
def add_hold_member_ticket_info_logic(request):
    member_ticket_id = request.POST.get('member_ticket_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    note = request.POST.get('note', '')
    extension_flag = request.POST.get('extension_flag', USE)
    class_id = request.session.get('class_id', '')
    error = None
    member_ticket_info = None
    if member_ticket_id is None or member_ticket_id == '':
        error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            member_ticket_info = MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        error = func_add_hold_member_ticket_info(request.user.id, class_id, member_ticket_id,
                                                 start_date, end_date, note, extension_flag)

    if error is None:
        # 회원의 고정 수업 정리
        func_update_lecture_member_fix_status_cd(class_id, member_ticket_info.member_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_ticket_info.member.name + ' 회원님의 '
                                  + member_ticket_info.ticket_tb.name + ' 수강권', log_how='일시정지', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 일시정지
def update_hold_member_ticket_info_logic(request):
    member_closed_date_history_id = request.POST.get('member_closed_date_history_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    note = request.POST.get('note', '')
    extension_flag = request.POST.get('extension_flag', '')
    class_id = request.session.get('class_id', '')
    error = None
    member_ticket_hold_history_info = None
    if member_closed_date_history_id is None or member_closed_date_history_id == '':
        error = '수강권 혿딩 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            member_ticket_hold_history_info = MemberClosedDateHistoryTb.objects.get(
                member_closed_date_history_id=member_closed_date_history_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        error = func_add_hold_member_ticket_info(request.user.id, class_id,
                                                 member_ticket_hold_history_info.member_ticket_tb_id,
                                                 start_date, end_date, note, extension_flag)

    if error is None:
        error = func_delete_hold_member_ticket_info(member_closed_date_history_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 일시정지
def delete_hold_member_ticket_info_logic(request):
    member_closed_date_history_id = request.POST.get('member_closed_date_history_id', '')
    error = None
    if member_closed_date_history_id is None or member_closed_date_history_id == '':
        error = '수강권 혿딩 정보를 불러오지 못했습니다.'

    if error is None:
        error = func_delete_hold_member_ticket_info(member_closed_date_history_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 일시정지
def add_member_closed_date_logic(request):
    member_ticket_id = request.POST.get('member_ticket_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    note = request.POST.get('note', '')
    extension_flag = request.POST.get('extension_flag', USE)
    class_id = request.session.get('class_id', '')
    error = None
    member_ticket_info = None
    if member_ticket_id is None or member_ticket_id == '':
        error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            member_ticket_info = MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        error = func_add_hold_member_ticket_info(request.user.id, class_id, member_ticket_id,
                                                 start_date, end_date, note, extension_flag)

    if error is None:
        # 회원의 고정 수업 정리
        func_update_lecture_member_fix_status_cd(class_id, member_ticket_info.member_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_ticket_info.member.name + ' 회원님의 '
                                  + member_ticket_info.ticket_tb.name + ' 수강권', log_how='일시정지', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 일시정지
def update_member_closed_date_logic(request):
    member_closed_date_history_id = request.POST.get('member_closed_date_history_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    note = request.POST.get('note', '')
    extension_flag = request.POST.get('extension_flag', '')
    class_id = request.session.get('class_id', '')
    error = None
    member_ticket_hold_history_info = None
    if member_closed_date_history_id is None or member_closed_date_history_id == '':
        error = '수강권 혿딩 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            member_ticket_hold_history_info = MemberClosedDateHistoryTb.objects.get(
                member_closed_date_history_id=member_closed_date_history_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        error = func_add_hold_member_ticket_info(request.user.id, class_id,
                                                 member_ticket_hold_history_info.member_ticket_tb_id,
                                                 start_date, end_date, note, extension_flag)

    if error is None:
        error = func_delete_hold_member_ticket_info(member_closed_date_history_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 일시정지
def delete_member_closed_date_logic(request):
    member_closed_date_history_id = request.POST.get('member_closed_date_history_id', '')
    error = None
    if member_closed_date_history_id is None or member_closed_date_history_id == '':
        error = '수강권 혿딩 정보를 불러오지 못했습니다.'

    if error is None:
        error = func_delete_hold_member_ticket_info(member_closed_date_history_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


def update_member_ticket_status_info_logic(request):
    class_id = request.session.get('class_id', '')
    member_ticket_id = request.POST.get('member_ticket_id', '')
    state_cd = request.POST.get('state_cd', '')
    refund_price = request.POST.get('refund_price', 0)
    refund_date = request.POST.get('refund_date', None)
    member_ticket_info = None
    error = None

    if member_ticket_id is None or member_ticket_id == '':
        error = '수강권 정보를 불러오지 못했습니다.'
    if refund_price == '':
        refund_price = 0
    if refund_date == '':
        refund_date = None

    if error is None:
        try:
            member_ticket_info = MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'
    if error is None:
        if state_cd == 'RF':
            if refund_price is None or refund_price == 0:
                error = '환불 금액을 입력해 주세요.'
            if error is None:
                try:
                    refund_price = int(refund_price)
                except ValueError:
                    error = '환불 금액은 숫자만 입력 가능합니다.'

            if error is None:
                if refund_price > member_ticket_info.price:
                    error = '환불 금액이 등록 금액보다 많습니다.'

            if error is None:
                try:
                    refund_date = datetime.datetime.strptime(str(refund_date), '%Y-%m-%d')
                except ValueError:
                    error = '환불 날짜 오류가 발생했습니다.[0]'
                except TypeError:
                    error = '환불 날짜 오류가 발생했습니다.[1]'

        if state_cd == STATE_CD_IN_PROGRESS:
            refund_price = 0
            refund_date = None
            if member_ticket_info.ticket_tb.use == UN_USE \
                    or member_ticket_info.ticket_tb.state_cd != STATE_CD_IN_PROGRESS:
                error = '해당 수강권은 진행중 상태가 아닙니다.'

        if state_cd == STATE_CD_FINISH:
            refund_price = 0
            refund_date = None

    if error is None:
        now = timezone.now()
        schedule_data = ScheduleTb.objects.filter(member_ticket_tb_id=member_ticket_id,
                                                  end_dt__lte=now,
                                                  use=USE).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                   | Q(state_cd=STATE_CD_ABSENCE))
        schedule_data_delete = ScheduleTb.objects.filter(member_ticket_tb_id=member_ticket_id,
                                                         end_dt__gt=now,
                                                         use=USE).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                          | Q(state_cd=STATE_CD_ABSENCE))
        repeat_schedule_data = RepeatScheduleTb.objects.filter(member_ticket_tb_id=member_ticket_id)

        schedule_data_count = ScheduleTb.objects.filter(member_ticket_tb_id=member_ticket_id).count()
        schedule_data_finish_count = ScheduleTb.objects.filter(Q(state_cd=STATE_CD_FINISH)
                                                               | Q(state_cd=STATE_CD_ABSENCE),
                                                               member_ticket_tb_id=member_ticket_id).count()

        if len(schedule_data) > 0:
            schedule_data.update(state_cd=STATE_CD_FINISH)
        if len(schedule_data_delete) > 0:
            schedule_data_delete.delete()
        if len(repeat_schedule_data) > 0:
            repeat_schedule_data.delete()

        member_ticket_avail_count = 0
        member_ticket_rem_count = 0
        if state_cd == STATE_CD_IN_PROGRESS:
            if member_ticket_info.member_ticket_reg_count >= schedule_data_count:
                member_ticket_avail_count = member_ticket_info.member_ticket_reg_count - schedule_data_count

        if state_cd != STATE_CD_FINISH:
            if member_ticket_info.member_ticket_reg_count >= schedule_data_finish_count:
                member_ticket_rem_count = member_ticket_info.member_ticket_reg_count - schedule_data_finish_count

        member_ticket_info.member_ticket_avail_count = member_ticket_avail_count
        member_ticket_info.member_ticket_rem_count = member_ticket_rem_count
        member_ticket_info.state_cd = state_cd
        member_ticket_info.refund_price = refund_price
        member_ticket_info.refund_date = refund_date
        member_ticket_info.save()
        if state_cd == STATE_CD_REFUND:
            member_payment_history_info = MemberPaymentHistoryTb(class_tb_id=class_id,
                                                                 member_id=member_ticket_info.member_id,
                                                                 member_ticket_tb_id=member_ticket_id,
                                                                 member_shop_tb_id=None,
                                                                 payment_price=0,
                                                                 refund_price=refund_price,
                                                                 pay_method='NONE',
                                                                 pay_date=refund_date, note='', use=USE)
            member_payment_history_info.save()
        if state_cd == STATE_CD_IN_PROGRESS:
            member_payment_history_data = MemberPaymentHistoryTb.objects.filter(refund_price__gt=0,
                                                                                member_ticket_tb_id=member_ticket_id,
                                                                                use=USE)
            member_payment_history_data.delete()
        if state_cd != STATE_CD_IN_PROGRESS:
            func_update_lecture_member_fix_status_cd(class_id, member_ticket_info.member_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_ticket_info.member.name + ' 회원님의 '
                                  + member_ticket_info.ticket_tb.name + ' 수강권 상태', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 회원 연결 관계 업데이트
def update_member_connection_info_logic(request):
    member_id = request.POST.get('member_id', '')
    member_auth_cd = request.POST.get('member_auth_cd', '')
    class_id = request.session.get('class_id', '')
    error = None
    member = None

    if member_id is None or member_id == '':
        error = '회원 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            member = MemberTb.objects.get(member_id=member_id)
        except ObjectDoesNotExist:
            error = '회원 정보를 불러오지 못했습니다.'

    if error is None:
        if member_auth_cd != AUTH_TYPE_VIEW and member_auth_cd != AUTH_TYPE_WAIT and member_auth_cd != AUTH_TYPE_DELETE:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:

        connection_check = func_check_member_connection_info(class_id, member_id)

        class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb').filter(class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW,
                                       member_ticket_tb__member_id=member_id, use=USE)

        if member_auth_cd == AUTH_TYPE_VIEW:
            # 연결이 안되어있는 경우 연결 요청 상태
            if not connection_check:
                member_auth_cd = AUTH_TYPE_WAIT
            # 회원이 인증을 안하고 내가 등록한 회원이면 무조건 연결상태로
            if not member.user.is_active and str(member.reg_info) == str(request.user.id):
                member_auth_cd = AUTH_TYPE_VIEW

        for class_member_ticket_info in class_member_ticket_data:
            class_member_ticket_info.member_ticket_tb.member_auth_cd = member_auth_cd
            class_member_ticket_info.member_ticket_tb.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


def add_lecture_info_logic(request):
    class_id = request.session.get('class_id', '')
    name = request.POST.get('name', '')
    member_num = request.POST.get('member_num', '')
    # member_num_view_flag = request.POST.get('member_num_view_flag', LECTURE_MEMBER_NUM_VIEW_ENABLE)
    note = request.POST.get('note', '')
    ing_color_cd = request.POST.get('ing_color_cd', '#ffd3d9')
    end_color_cd = request.POST.get('end_color_cd', '#d2d1cf')
    ing_font_color_cd = request.POST.get('ing_font_color_cd', '#282828')
    end_font_color_cd = request.POST.get('end_font_color_cd', '#282828')
    lecture_minute = request.POST.get('lecture_minute', 60)
    start_time = request.POST.get('start_time', 'A-0')
    main_trainer_id = request.POST.get('main_trainer_id', request.user.id)
    error = None
    lecture_id = None
    lecture_type_cd = LECTURE_TYPE_NORMAL
    if main_trainer_id is None or main_trainer_id == '':
        main_trainer_id = request.user.id
    if member_num == 1 or member_num == '1':
        lecture_type_cd = LECTURE_TYPE_ONE_TO_ONE

    check_class_auth = MemberClassTb.objects.filter(class_tb_id=class_id, member_id=main_trainer_id,
                                                    auth_cd=AUTH_TYPE_VIEW, use=USE).count()
    if check_class_auth == 0:
        error = '선택한 강사가 지점 연결 수락후 등록 가능 합니다.'

    try:
        with transaction.atomic():
            lecture_info = LectureTb(class_tb_id=class_id, member_num=member_num, lecture_type_cd=lecture_type_cd,
                                     # member_num_view_flag=member_num_view_flag,
                                     name=name, note=note, ing_color_cd=ing_color_cd, end_color_cd=end_color_cd,
                                     ing_font_color_cd=ing_font_color_cd, end_font_color_cd=end_font_color_cd,
                                     lecture_minute=lecture_minute, state_cd=STATE_CD_IN_PROGRESS,
                                     main_trainer_id=main_trainer_id,
                                     start_time=start_time, use=USE)

            lecture_info.save()
            lecture_id = lecture_info.lecture_id
    except ValueError:
        error = '오류가 발생했습니다. [1]'
    except IntegrityError:
        error = '오류가 발생했습니다. [2]'
    except TypeError:
        error = '오류가 발생했습니다. [3]'
    except ValidationError:
        error = '오류가 발생했습니다. [4]'
    except InternalError:
        error = '오류가 발생했습니다. [5]'

    if error is not None:
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=name+' 수업', log_how='추가', use=USE)
        log_data.save()
    return JsonResponse({'lecture_id': str(lecture_id)}, json_dumps_params={'ensure_ascii': True})


def delete_lecture_info_logic(request):
    class_id = request.session.get('class_id', '')
    lecture_id = request.POST.get('lecture_id', '')
    error = None
    lecture_info = None

    try:
        lecture_info = LectureTb.objects.get(lecture_id=lecture_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다.'
    if error is None:
        if lecture_info.lecture_type_cd == LECTURE_TYPE_ONE_TO_ONE:
            lecture_counter = LectureTb.objects.filter(class_tb_id=class_id,
                                                       lecture_type_cd=LECTURE_TYPE_ONE_TO_ONE, use=USE).count()
            if lecture_counter == 1:
                error = '개인 수업을 삭제 할 수 없습니다.'

    ticket_lecture_data = TicketLectureTb.objects.filter(lecture_tb_id=lecture_id, use=USE)

    query_ticket_info = Q()
    query_ticket_counter = 0
    for ticket_lecture_info in ticket_lecture_data:
        query_ticket_info |= Q(member_ticket_tb__ticket_tb_id=ticket_lecture_info.ticket_tb.ticket_id)
        query_ticket_counter += 1

    try:
        with transaction.atomic():

            if error is None:
                # schedule_data = ScheduleTb.objects.filter(class_tb_id=class_id,
                #                                           lecture_tb_id=lecture_id,
                #                                           end_dt__lte=timezone.now(),
                #                                           en_dis_type=ON_SCHEDULE_TYPE
                #                                           ).exclude(Q(state_cd=STATE_CD_FINISH)
                #                                                     | Q(state_cd=STATE_CD_ABSENCE))
                schedule_data_delete = ScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                                                 # end_dt__gt=timezone.now(),
                                                                 en_dis_type=ON_SCHEDULE_TYPE
                                                                 ).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                           | Q(state_cd=STATE_CD_ABSENCE))
                repeat_schedule_data = RepeatScheduleTb.objects.filter(class_tb_id=class_id,
                                                                       lecture_tb_id=lecture_id)
                # schedule_data_finish = ScheduleTb.objects.filter(Q(state_cd=STATE_CD_FINISH)
                #                                                  | Q(state_cd=STATE_CD_ABSENCE),
                #                                                  class_tb_id=class_id,
                #                                                  lecture_tb_id=lecture_id,
                #                                                  en_dis_type=ON_SCHEDULE_TYPE)
                # if len(schedule_data) > 0:
                #     schedule_data.update(state_cd=STATE_CD_FINISH)
                if len(schedule_data_delete) > 0:
                    schedule_data_delete.delete()
                if len(repeat_schedule_data) > 0:
                    repeat_schedule_data.delete()
                # if len(schedule_data_finish) > 0:
                #     schedule_data_finish.update(use=UN_USE)

                # 관련 수간권 회원들 수강정보 업데이트
                if len(ticket_lecture_data) > 0:
                    if query_ticket_counter > 0:
                        class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
                            'class_tb',
                            'member_ticket_tb__ticket_tb').filter(query_ticket_info, class_tb_id=class_id,
                                                                  auth_cd=AUTH_TYPE_VIEW,
                                                                  member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS,
                                                                  member_ticket_tb__use=USE, use=USE)

                        for class_member_ticket_info in class_member_ticket_data:
                            error = func_refresh_member_ticket_count(class_member_ticket_info.class_tb_id,
                                                                     class_member_ticket_info.member_ticket_tb_id)

            if error is None:
                lecture_info.state_cd = STATE_CD_FINISH
                lecture_info.use = UN_USE
                lecture_info.save()
                if len(ticket_lecture_data) > 0:
                    ticket_lecture_data.delete()

                lecture_member_fix_data = LectureMemberTb.objects.filter(class_tb_id=class_id,
                                                                         lecture_tb_id=lecture_id, use=USE)
                lecture_member_fix_data.delete()

            if error is not None:
                raise InternalError

    except ValueError:
        error = '오류가 발생했습니다. [1]'
    except IntegrityError:
        error = '오류가 발생했습니다. [2]'
    except TypeError:
        error = '오류가 발생했습니다. [3]'
    except ValidationError:
        error = '오류가 발생했습니다. [4]'
    except InternalError:
        error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=lecture_info.name+' 수업', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


def update_lecture_info_logic(request):
    class_id = request.session.get('class_id', '')
    lecture_id = request.POST.get('lecture_id', '')
    member_num = request.POST.get('member_num', '')
    # member_num_view_flag = request.POST.get('member_num_view_flag', LECTURE_MEMBER_NUM_VIEW_ENABLE)
    name = request.POST.get('name', '')
    note = request.POST.get('note', '')
    ing_color_cd = request.POST.get('ing_color_cd', '')
    end_color_cd = request.POST.get('end_color_cd', '')
    ing_font_color_cd = request.POST.get('ing_font_color_cd', '')
    end_font_color_cd = request.POST.get('end_font_color_cd', '')
    lecture_minute = request.POST.get('lecture_minute', 60)
    start_time = request.POST.get('start_time', '')
    update_this_to_all_plans = request.POST.get('update_this_to_all_plans', UN_USE)
    main_trainer_id = request.POST.get('main_trainer_id', '')
    lecture_info = None
    error = None

    try:
        lecture_info = LectureTb.objects.get(lecture_id=lecture_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다.'

    if error is None:
        if member_num == '' or member_num is None:
            member_num = lecture_info.member_num
        #
        # if member_num_view_flag == '' or member_num_view_flag is None:
        #     member_num_view_flag = lecture_info.member_num_view_flag

        if name == '' or name is None:
            name = lecture_info.name

        if note == '' or note is None:
            note = lecture_info.note

        if ing_color_cd == '' or ing_color_cd is None:
            ing_color_cd = lecture_info.ing_color_cd

        if end_color_cd == '' or end_color_cd is None:
            end_color_cd = lecture_info.end_color_cd

        if ing_font_color_cd == '' or ing_font_color_cd is None:
            ing_font_color_cd = lecture_info.ing_font_color_cd

        if end_font_color_cd == '' or end_font_color_cd is None:
            end_font_color_cd = lecture_info.end_font_color_cd

        if lecture_minute == '' or lecture_minute is None:
            lecture_minute = lecture_info.lecture_minute

        if start_time == '' or start_time is None:
            start_time = lecture_info.start_time

        if main_trainer_id == '' or main_trainer_id is None:
            main_trainer_id = lecture_info.main_trainer_id

    if error is None:
        try:
            if int(member_num) <= 0:
                error = '정원은 1명 이상이어야 합니다.'
        except ValueError:
            error = '정원은 숫자만 입력 가능합니다.'
        except TypeError:
            error = '정원은 숫자만 입력 가능합니다.'

    if error is None:

        lecture_member_fix_data = LectureMemberTb.objects.filter(class_tb_id=class_id,
                                                                 lecture_tb_id=lecture_id, use=USE)
        # 수업에 고정회원 가능 여부 체크
        if len(lecture_member_fix_data) > int(member_num):
            error = '수정하려는 정원보다 고정 회원이 많습니다.'
    if error is None:
        if main_trainer_id != lecture_info.main_trainer_id:
            check_class_auth = MemberClassTb.objects.filter(class_tb_id=class_id, member_id=main_trainer_id,
                                                            auth_cd=AUTH_TYPE_VIEW, use=USE).count()
            if check_class_auth == 0:
                error = '선택한 강사가 지점 연결 수락후 선택 가능 합니다.'

    if error is None:
        lecture_info.member_num = member_num
        # lecture_info.member_num_view_flag = member_num_view_flag
        lecture_info.name = name
        lecture_info.note = note
        lecture_info.ing_color_cd = ing_color_cd
        lecture_info.end_color_cd = end_color_cd
        lecture_info.ing_font_color_cd = ing_font_color_cd
        lecture_info.end_font_color_cd = end_font_color_cd
        lecture_info.lecture_minute = lecture_minute
        lecture_info.start_time = start_time
        lecture_info.main_trainer_id = main_trainer_id
        lecture_info.save()

    if error is None:
        # if str(lecture_info.lecture_type_cd) == str(LECTURE_TYPE_ONE_TO_ONE):
        #     # 오늘 이전의 일정
        #     schedule_data_past = ScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb__isnull=True,
        #                                                    end_dt__lte=timezone.now(), en_dis_type=ON_SCHEDULE_TYPE)
        #     # 오늘 이후의 일정
        #     schedule_data_future = ScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb__isnull=True,
        #                                                      end_dt__gt=timezone.now(), en_dis_type=ON_SCHEDULE_TYPE)
        # else:
        # 오늘 이전의 일정
        schedule_data_past = ScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                                       end_dt__lte=timezone.now(), en_dis_type=ON_SCHEDULE_TYPE)
        # 오늘 이후의 일정
        schedule_data_future = ScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                                         end_dt__gt=timezone.now(), en_dis_type=ON_SCHEDULE_TYPE)
        if str(update_this_to_all_plans) == str(USE):
            schedule_data_past.update(ing_color_cd=ing_color_cd, end_color_cd=end_color_cd,
                                      ing_font_color_cd=ing_font_color_cd, end_font_color_cd=end_font_color_cd,
                                      max_mem_count=member_num, trainer_id=main_trainer_id)

        schedule_data_future.update(ing_color_cd=ing_color_cd, end_color_cd=end_color_cd,
                                    ing_font_color_cd=ing_font_color_cd, end_font_color_cd=end_font_color_cd,
                                    max_mem_count=member_num, trainer_id=main_trainer_id)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=lecture_info.name+' 수업', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


def update_lecture_status_info_logic(request):
    lecture_id = request.POST.get('lecture_id', '')
    state_cd = request.POST.get('state_cd', '')
    class_id = request.session.get('class_id')

    error = None
    lecture_info = None
    now = timezone.now()
    try:
        lecture_info = LectureTb.objects.get(class_tb_id=class_id, lecture_id=lecture_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다.'

    if error is None:
        if lecture_info.lecture_type_cd == LECTURE_TYPE_ONE_TO_ONE:
            if state_cd == STATE_CD_FINISH:
                lecture_counter = LectureTb.objects.filter(class_tb_id=class_id,
                                                           lecture_type_cd=LECTURE_TYPE_ONE_TO_ONE,
                                                           state_cd=STATE_CD_IN_PROGRESS,
                                                           use=USE).count()
                if lecture_counter == 1:
                    error = '개인 수업을 비활성화 할 수 없습니다.'
            # error = '개인 수업은 상태 변경이 불가합니다.'

    if error is None:
        ticket_lecture_data = TicketLectureTb.objects.filter(class_tb_id=class_id, lecture_tb_id=lecture_id, use=USE)
        query_ticket_info = Q()
        query_ticket_counter = 0
        for ticket_lecture_info in ticket_lecture_data:
            query_ticket_info |= Q(member_ticket_tb__ticket_tb_id=ticket_lecture_info.ticket_tb.ticket_id)
            query_ticket_counter += 1

        if state_cd == STATE_CD_FINISH:
            schedule_data = ScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                                      end_dt__lte=now, use=USE).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                                        | Q(state_cd=STATE_CD_ABSENCE))
            schedule_data_delete = ScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                                             end_dt__gt=now,
                                                             use=USE).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                              | Q(state_cd=STATE_CD_ABSENCE))
            repeat_schedule_data = RepeatScheduleTb.objects.filter(class_tb_id=class_id, lecture_tb_id=lecture_id)

            if len(schedule_data) > 0:
                schedule_data.update(state_cd=STATE_CD_FINISH)
            if len(schedule_data_delete) > 0:
                schedule_data_delete.delete()
            if len(repeat_schedule_data) > 0:
                repeat_schedule_data.delete()

            # 관련 수간권 회원들 수강정보 업데이트
            if len(ticket_lecture_data) > 0:
                if query_ticket_counter > 0:
                    class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
                        'member_ticket_tb__ticket_tb').filter(query_ticket_info, class_tb_id=class_id,
                                                              auth_cd=AUTH_TYPE_VIEW,
                                                              member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS,
                                                              member_ticket_tb__use=USE, use=USE)

                    for class_member_ticket_info in class_member_ticket_data:
                        error = func_refresh_member_ticket_count(class_member_ticket_info.class_tb_id,
                                                                 class_member_ticket_info.member_ticket_tb_id)

            lecture_member_fix_data = LectureMemberTb.objects.filter(class_tb_id=class_id,
                                                                     lecture_tb_id=lecture_id, use=USE)
            lecture_member_fix_data.delete()
        lecture_info.state_cd = state_cd
        lecture_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=lecture_info.name+' 수업', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 그룹 고정 회원 기능
def update_fix_lecture_member_logic(request):
    class_id = request.session.get('class_id', '')
    lecture_id = request.POST.get('lecture_id', '')
    member_ids = request.POST.getlist('member_ids[]', '')
    error = None
    lecture_info = None

    if lecture_id == '':
        error = '수업 정보를 불러오지 못했습니다.'

    if error is None:
        # 수업 정보 가져오기
        try:
            lecture_info = LectureTb.objects.get(lecture_id=lecture_id, use=USE)
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다.'

    if error is None:
        if str(lecture_info.state_cd) == str(STATE_CD_IN_PROGRESS):
            for member_id in member_ids:
                if error is None:
                    member_lecture_list = func_get_member_lecture_list(class_id, member_id)
                    try:
                        member_lecture_list[lecture_id]
                    except KeyError:
                        error = '해당 회원님은 수업에 참여할 수 있는 수강권이 없습니다.'

                if error is None:
                    try:
                        with transaction.atomic():
                            lecture_member_fix_data = LectureMemberTb.objects.filter(class_tb_id=class_id,
                                                                                     lecture_tb_id=lecture_id, use=USE)
                            # 이미 고정 회원의 경우 제거하기
                            try:
                                lecture_member_fix = LectureMemberTb.objects.get(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                                                                 member_id=member_id, use=USE)
                                lecture_member_fix.delete()
                            except ObjectDoesNotExist:

                                # 수업에 고정회원 가능 여부 체크
                                if len(lecture_member_fix_data) + 1 > lecture_info.member_num:
                                    error = '정원보다 고정 회원이 많습니다.'
                                    raise InternalError()

                                # 수업에 고정회원 추가하기
                                lecture_member_fix = LectureMemberTb(class_tb_id=class_id, lecture_tb_id=lecture_id,
                                                                     member_id=member_id, fix_state_cd='FIX', use=USE)
                                lecture_member_fix.save()

                    except ValueError:
                        error = '오류가 발생했습니다. [4]'
                    except IntegrityError:
                        error = '오류가 발생했습니다. [5]'
                    except TypeError:
                        error = '오류가 발생했습니다. [6]'
                    except ValidationError:
                        error = '오류가 발생했습니다. [7]'
                    except InternalError:
                        error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


class GetLectureInfoViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        off_minute = self.request.session.get('setting_calendar_basic_select_time', 60)
        lecture_id = request.GET.get('lecture_id', '')
        lecture_info = {}
        if lecture_id is not None and lecture_id != '':
            lecture_info = func_get_lecture_info(class_id, lecture_id, request.user.id)
        else:
            lecture_info = {'lecture_id': '', 'lecture_name': 'OFF',
                            'lecture_note': '',
                            'lecture_state_cd': STATE_CD_IN_PROGRESS, 'lecture_max_num': 0,
                            # 'lecture_max_member_num_view_flag': LECTURE_MEMBER_NUM_VIEW_ENABLE,
                            'lecture_reg_dt': '', 'lecture_mod_dt': '',
                            'lecture_ticket_list': [],
                            'lecture_ticket_state_cd_list': [],
                            'lecture_ticket_id_list': [],
                            'lecture_ing_member_num': 0,
                            'lecture_ing_color_cd': '',
                            'lecture_ing_font_color_cd': '',
                            'lecture_end_color_cd': '',
                            'lecture_end_font_color_cd': '',
                            'lecture_type_cd': 'OFF',
                            'lecture_minute': off_minute,
                            'lecture_main_trainer_id': '',
                            'lecture_main_trainer_name': '',
                            'lecture_member_list': []}
        return JsonResponse(lecture_info, json_dumps_params={'ensure_ascii': True})


class GetLectureIngListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        lecture_sort = self.request.GET.get('sort_val', SORT_TICKET_NAME)
        sort_order_by = self.request.GET.get('sort_order_by', SORT_ASC)
        keyword = self.request.GET.get('keyword', '')
        sort_info = int(lecture_sort)
        error = None

        lecture_data = LectureTb.objects.select_related(
            'main_trainer', 'class_tb__member').filter(class_tb_id=class_id, state_cd=STATE_CD_IN_PROGRESS,
                                                       name__contains=keyword, use=USE).order_by('lecture_id')

        lecture_ticket_data = TicketLectureTb.objects.select_related(
            'lecture_tb__main_trainer', 'class_tb__member',
            'ticket_tb').filter(class_tb_id=class_id, lecture_tb__state_cd=STATE_CD_IN_PROGRESS,
                                lecture_tb__name__contains=keyword, lecture_tb__use=USE,
                                use=USE).order_by('lecture_tb_id', 'ticket_tb_id')

        lecture_data_dict = {}
        # 수업과 연관되어있는 수강권 정보 셋팅
        for lecture_ticket_info in lecture_ticket_data:
            lecture_tb = lecture_ticket_info.lecture_tb
            ticket_tb = lecture_ticket_info.ticket_tb
            lecture_id = str(lecture_tb.lecture_id)
            main_trainer_id = lecture_ticket_info.class_tb.member_id
            main_trainer_name = lecture_ticket_info.class_tb.member.name
            if lecture_tb.main_trainer is None or lecture_tb.main_trainer == '':
                lecture_tb.main_trainer_id = lecture_ticket_info.class_tb.member_id
                lecture_tb.save()
            else:
                main_trainer_id = lecture_tb.main_trainer_id
                main_trainer_name = lecture_tb.main_trainer.name
            try:
                lecture_data_dict[lecture_id]
            except KeyError:
                lecture_data_dict[lecture_id] = {'lecture_id': lecture_id,
                                                 'lecture_name': lecture_tb.name,
                                                 'lecture_note': lecture_tb.note,
                                                 'lecture_max_num': lecture_tb.member_num,
                                                 # 'lecture_max_member_num_view_flag': lecture_tb.member_num_view_flag,
                                                 'lecture_ing_color_cd': lecture_tb.ing_color_cd,
                                                 'lecture_ing_font_color_cd': lecture_tb.ing_font_color_cd,
                                                 'lecture_end_color_cd': lecture_tb.end_color_cd,
                                                 'lecture_end_font_color_cd': lecture_tb.end_font_color_cd,
                                                 'lecture_minute': lecture_tb.lecture_minute,
                                                 'lecture_type_cd': lecture_tb.lecture_type_cd,
                                                 'lecture_reg_dt': lecture_tb.reg_dt,
                                                 'main_trainer_id': main_trainer_id,
                                                 'main_trainer_name': main_trainer_name,
                                                 'lecture_ticket_list': [],
                                                 'lecture_ticket_state_cd_list': [],
                                                 'lecture_ticket_id_list': []}
            if ticket_tb.use == USE:
                lecture_data_dict[lecture_id]['lecture_ticket_list'].append(ticket_tb.name)
                lecture_data_dict[lecture_id]['lecture_ticket_state_cd_list'].append(ticket_tb.state_cd)
                lecture_data_dict[lecture_id]['lecture_ticket_id_list'].append(ticket_tb.ticket_id)

        # 수업에 수강권이 연결되어있지 않은 경우 처리
        if len(lecture_data) != len(lecture_data_dict):
            for lecture_info in lecture_data:
                lecture_id = str(lecture_info.lecture_id)
                main_trainer_id = lecture_info.class_tb.member_id
                main_trainer_name = lecture_info.class_tb.member.name
                if lecture_info.main_trainer is None or lecture_info.main_trainer == '':
                    lecture_info.main_trainer_id = lecture_info.class_tb.member_id
                    lecture_info.save()
                else:
                    main_trainer_id = lecture_info.main_trainer_id
                    main_trainer_name = lecture_info.main_trainer.name

                try:
                    lecture_data_dict[lecture_id]
                except KeyError:
                    lecture_data_dict[lecture_id] = {'lecture_id': lecture_id,
                                                     'lecture_name': lecture_info.name,
                                                     'lecture_note': lecture_info.note,
                                                     'lecture_max_num': lecture_info.member_num,
                                                     # 'lecture_max_member_num_view_flag': lecture_info.member_num_view_flag,
                                                     'lecture_ing_color_cd': lecture_info.ing_color_cd,
                                                     'lecture_ing_font_color_cd': lecture_info.ing_font_color_cd,
                                                     'lecture_end_color_cd': lecture_info.end_color_cd,
                                                     'lecture_end_font_color_cd': lecture_info.end_font_color_cd,
                                                     'lecture_minute': lecture_info.lecture_minute,
                                                     'lecture_type_cd': lecture_info.lecture_type_cd,
                                                     'lecture_reg_dt': lecture_info.reg_dt,
                                                     'main_trainer_id': main_trainer_id,
                                                     'main_trainer_name': main_trainer_name,
                                                     'lecture_ticket_list': [],
                                                     'lecture_ticket_state_cd_list': [],
                                                     'lecture_ticket_id_list': []}

        # lecture_data_dict = sorted(lecture_data_dict.items(), key=lambda x: (x[1]['lecture_type_cd']), reverse=True)
        #
        # if len(lecture_data_dict) > 0:
        #     if lecture_data_dict[0][1]['lecture_type_cd'] == 'ONE_TO_ONE':
        #         lecture_data_dict = collections.OrderedDict(
        #             lecture_data_dict[0:1]+sorted(lecture_data_dict[1:], key=lambda x: (x[1]['lecture_name'])))
        #     else:
        #         lecture_data_dict = collections.OrderedDict(sorted(lecture_data_dict,
        #                                                            key=lambda x: (x[1]['lecture_name'])))

        lecture_list = []

        class_member_ticket_list = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb__ticket_tb',
            'member_ticket_tb__member').filter(Q(member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS)
                                               | Q(member_ticket_tb__state_cd=STATE_CD_HOLDING),
                                               class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW,
                                               member_ticket_tb__ticket_tb__state_cd=STATE_CD_IN_PROGRESS,
                                               member_ticket_tb__ticket_tb__use=USE,
                                               member_ticket_tb__use=USE,
                                               use=USE).order_by('member_ticket_tb__ticket_tb',
                                                                 'member_ticket_tb__member')

        for lecture_info in lecture_data_dict:
            member_list = {}
            for class_member_ticket_info in class_member_ticket_list:
                if len(lecture_data_dict[lecture_info]['lecture_ticket_id_list']) > 0:
                    for ticket_info in lecture_data_dict[lecture_info]['lecture_ticket_id_list']:
                        if class_member_ticket_info.member_ticket_tb.ticket_tb.ticket_id == ticket_info:
                            member_list[class_member_ticket_info.member_ticket_tb.member_id] =\
                                class_member_ticket_info.member_ticket_tb.member_id

            lecture_data_dict[lecture_info]['lecture_ing_member_num'] = len(member_list)
            lecture_list.append(lecture_data_dict[lecture_info])

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        else:
            if sort_info == SORT_LECTURE_NAME:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_name'],
                                      reverse=int(sort_order_by))
            elif sort_info == SORT_LECTURE_MEMBER_COUNT:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_ing_member_num'],
                                      reverse=int(sort_order_by))
            elif sort_info == SORT_LECTURE_CAPACITY_COUNT:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_max_num'],
                                      reverse=int(sort_order_by))
            elif sort_info == SORT_LECTURE_CREATE_DATE:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_reg_dt'],
                                      reverse=int(sort_order_by))

        return JsonResponse({'current_lecture_data': lecture_list}, json_dumps_params={'ensure_ascii': True})


class GetLectureEndListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        lecture_sort = self.request.GET.get('sort_val', SORT_TICKET_NAME)
        sort_order_by = self.request.GET.get('sort_order_by', SORT_ASC)
        keyword = self.request.GET.get('keyword', '')
        sort_info = int(lecture_sort)
        error = None

        lecture_data = LectureTb.objects.select_related(
            'main_trainer',  'class_tb__member').filter(class_tb_id=class_id, state_cd=STATE_CD_FINISH,
                                                        name__contains=keyword, use=USE).order_by('lecture_id')

        lecture_ticket_data = TicketLectureTb.objects.select_related(
            'ticket_tb', 'class_tb__member',
            'lecture_tb__main_trainer').filter(class_tb_id=class_id, lecture_tb__state_cd=STATE_CD_FINISH,
                                               lecture_tb__name__contains=keyword, lecture_tb__use=USE,
                                               use=USE).order_by('lecture_tb_id', 'ticket_tb_id')

        lecture_data_dict = {}
        # 수업과 연관되어있는 수강권 정보 셋팅
        for lecture_ticket_info in lecture_ticket_data:
            lecture_tb = lecture_ticket_info.lecture_tb
            ticket_tb = lecture_ticket_info.ticket_tb
            lecture_id = str(lecture_tb.lecture_id)
            main_trainer_id = lecture_ticket_info.class_tb.member_id
            main_trainer_name = lecture_ticket_info.class_tb.member.name
            if lecture_tb.main_trainer is None or lecture_tb.main_trainer == '':
                lecture_tb.main_trainer_id = lecture_ticket_info.class_tb.member_id
                lecture_tb.save()
            else:
                main_trainer_id = lecture_tb.main_trainer_id
                main_trainer_name = lecture_tb.main_trainer.name

            try:
                lecture_data_dict[lecture_id]
            except KeyError:
                lecture_data_dict[lecture_id] = {'lecture_id': lecture_id,
                                                 'lecture_name': lecture_tb.name,
                                                 'lecture_note': lecture_tb.note,
                                                 'lecture_max_num': lecture_tb.member_num,
                                                 # 'lecture_max_member_num_view_flag': lecture_tb.member_num_view_flag,
                                                 'lecture_ing_color_cd': lecture_tb.ing_color_cd,
                                                 'lecture_ing_font_color_cd': lecture_tb.ing_font_color_cd,
                                                 'lecture_end_color_cd': lecture_tb.end_color_cd,
                                                 'lecture_end_font_color_cd': lecture_tb.end_font_color_cd,
                                                 'lecture_minute': lecture_tb.lecture_minute,
                                                 'lecture_type_cd': lecture_tb.lecture_type_cd,
                                                 'lecture_reg_dt': lecture_tb.reg_dt,
                                                 'lecture_main_trainer_id': main_trainer_id,
                                                 'lecture_main_trainer_name': main_trainer_name,
                                                 'lecture_ticket_list': [],
                                                 'lecture_ticket_state_cd_list': [],
                                                 'lecture_ticket_id_list': []}
            if ticket_tb.use == USE:
                lecture_data_dict[lecture_id]['lecture_ticket_list'].append(ticket_tb.name)
                lecture_data_dict[lecture_id]['lecture_ticket_state_cd_list'].append(ticket_tb.state_cd)
                lecture_data_dict[lecture_id]['lecture_ticket_id_list'].append(ticket_tb.ticket_id)

        # 수업에 수강권이 연결되어있지 않은 경우 처리
        if len(lecture_data) != len(lecture_data_dict):
            for lecture_info in lecture_data:
                lecture_id = str(lecture_info.lecture_id)
                main_trainer_id = lecture_info.class_tb.member_id
                main_trainer_name = lecture_info.class_tb.member.name
                if lecture_info.main_trainer is None or lecture_info.main_trainer == '':
                    lecture_info.main_trainer_id = lecture_info.class_tb.member_id
                    lecture_info.save()
                else:
                    main_trainer_id = lecture_info.main_trainer_id
                    main_trainer_name = lecture_info.main_trainer.name
                try:
                    lecture_data_dict[lecture_id]
                except KeyError:
                    lecture_data_dict[lecture_id] = {'lecture_id': lecture_id,
                                                     'lecture_name': lecture_info.name,
                                                     'lecture_note': lecture_info.note,
                                                     'lecture_max_num': lecture_info.member_num,
                                                     # 'lecture_max_member_num_view_flag': lecture_info.member_num_view_flag,
                                                     'lecture_ing_color_cd': lecture_info.ing_color_cd,
                                                     'lecture_ing_font_color_cd': lecture_info.ing_font_color_cd,
                                                     'lecture_end_color_cd': lecture_info.end_color_cd,
                                                     'lecture_end_font_color_cd': lecture_info.end_font_color_cd,
                                                     'lecture_minute': lecture_info.lecture_minute,
                                                     'lecture_type_cd': lecture_info.lecture_type_cd,
                                                     'lecture_reg_dt': lecture_info.reg_dt,
                                                     'lecture_main_trainer_id': main_trainer_id,
                                                     'lecture_main_trainer_name': main_trainer_name,
                                                     'lecture_ticket_list': [],
                                                     'lecture_ticket_state_cd_list': [],
                                                     'lecture_ticket_id_list': []}
        # lecture_data_dict = sorted(lecture_data_dict.items(), key=lambda x: (x[1]['lecture_type_cd']), reverse=True)
        # if len(lecture_data_dict) > 0:
        #     if lecture_data_dict[0][1]['lecture_type_cd'] == 'ONE_TO_ONE':
        #         lecture_data_dict = collections.OrderedDict(
        #             lecture_data_dict[0:1]+sorted(lecture_data_dict[1:], key=lambda x: (x[1]['lecture_name'])))
        #     else:
        #         lecture_data_dict = collections.OrderedDict(sorted(lecture_data_dict,
        #                                                            key=lambda x: (x[1]['lecture_name'])))

        lecture_list = []

        for lecture_info in lecture_data_dict:
            lecture_list.append(lecture_data_dict[lecture_info])

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        else:
            if sort_info == SORT_LECTURE_NAME:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_name'],
                                      reverse=int(sort_order_by))
            elif sort_info == SORT_LECTURE_MEMBER_COUNT:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_name'],
                                      reverse=int(sort_order_by))
            elif sort_info == SORT_LECTURE_CAPACITY_COUNT:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_max_num'],
                                      reverse=int(sort_order_by))
            elif sort_info == SORT_LECTURE_CREATE_DATE:
                lecture_list = sorted(lecture_list, key=lambda elem: elem['lecture_reg_dt'],
                                      reverse=int(sort_order_by))

        return JsonResponse({'finish_lecture_data': lecture_list}, json_dumps_params={'ensure_ascii': True})


class GetLectureIngMemberListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        lecture_id = request.GET.get('lecture_id', '')
        error = None
        member_list = []
        lecture_ticket_data = TicketLectureTb.objects.select_related(
            'ticket_tb', 'lecture_tb').filter(class_tb_id=class_id, lecture_tb_id=lecture_id, lecture_tb__use=USE,
                                              ticket_tb__state_cd=STATE_CD_IN_PROGRESS, ticket_tb__use=USE,
                                              use=USE).order_by('lecture_tb_id', 'ticket_tb_id')

        if len(lecture_ticket_data) > 0:
            # 수업에 속한 수강권을 가지고 있는 회원들을 가지고 오기 위한 작업
            query_ticket_list = Q()
            query_ticket_counter = 0
            for lecture_ticket_info in lecture_ticket_data:
                query_ticket_list |= Q(member_ticket_tb__ticket_tb_id=lecture_ticket_info.ticket_tb_id)
                query_ticket_counter += 1

            if query_ticket_counter > 0:
                all_class_member_ticket_list = ClassMemberTicketTb.objects.select_related(
                    'member_ticket_tb__ticket_tb',
                    'member_ticket_tb__member').filter(query_ticket_list,
                                                       Q(member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS)
                                                       | Q(member_ticket_tb__state_cd=STATE_CD_HOLDING),
                                                       class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW,
                                                       member_ticket_tb__ticket_tb__state_cd=STATE_CD_IN_PROGRESS,
                                                       member_ticket_tb__ticket_tb__use=USE,
                                                       member_ticket_tb__use=USE,
                                                       use=USE).order_by('member_ticket_tb__member__name',
                                                                         'member_ticket_tb__end_date')

                member_list = func_get_member_from_member_ticket_list(all_class_member_ticket_list, lecture_id,
                                                                      request.user.id)

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'lecture_ing_member_list': member_list}, json_dumps_params={'ensure_ascii': True})


# 패키지 추가
def add_ticket_info_logic(request):
    class_id = request.session.get('class_id', '')
    ticket_name = request.POST.get('ticket_name')
    ticket_note = request.POST.get('ticket_note')
    ticket_effective_days = request.POST.get('ticket_effective_days', 0)
    ticket_price = request.POST.get('ticket_price', 0)
    ticket_month_schedule_enable = request.POST.get('ticket_month_schedule_enable', 99999)
    ticket_week_schedule_enable = request.POST.get('ticket_week_schedule_enable', 99999)
    ticket_day_schedule_enable = request.POST.get('ticket_day_schedule_enable', 99999)
    ticket_reg_count = request.POST.get('ticket_reg_count', 0)
    lecture_id_list = request.POST.getlist('lecture_id_list[]', '')
    error = None

    try:
        ticket_effective_days = int(ticket_effective_days)
    except ValueError:
        error = '유효 기간은 숫자만 입력 가능합니다.'

    try:
        ticket_price = int(ticket_price)
    except ValueError:
        error = '수강권 금액은 숫자만 입력 가능합니다.'

    try:
        ticket_month_schedule_enable = int(ticket_month_schedule_enable)
    except ValueError:
        error = '월간 일정 등록 제한 횟수는 숫자만 입력 가능합니다.'

    try:
        ticket_week_schedule_enable = int(ticket_week_schedule_enable)
    except ValueError:
        error = '주간 일정 등록 제한 횟수는 숫자만 입력 가능합니다.'

    try:
        ticket_day_schedule_enable = int(ticket_day_schedule_enable)
    except ValueError:
        error = '일일 일정 등록 제한 횟수는 숫자만 입력 가능합니다.'

    try:
        ticket_reg_count = int(ticket_reg_count)
    except ValueError:
        error = '등록 횟수는 숫자만 입력 가능합니다.'

    if error is None:
        try:
            with transaction.atomic():

                ticket_info = TicketTb(class_tb_id=class_id, name=ticket_name, state_cd=STATE_CD_IN_PROGRESS,
                                       effective_days=ticket_effective_days, price=ticket_price,
                                       month_schedule_enable=ticket_month_schedule_enable,
                                       week_schedule_enable=ticket_week_schedule_enable,
                                       day_schedule_enable=ticket_day_schedule_enable,
                                       reg_count=ticket_reg_count, note=ticket_note, use=USE)
                ticket_info.save()

                for lecture_id in lecture_id_list:
                    if lecture_id is not None and lecture_id != '':
                        ticket_lecture_counter = TicketLectureTb.objects.filter(class_tb_id=class_id,
                                                                                ticket_tb_id=ticket_info.ticket_id,
                                                                                lecture_tb_id=lecture_id,
                                                                                use=USE).count()
                        if ticket_lecture_counter == 0:
                            ticket_lecture_info = TicketLectureTb(class_tb_id=class_id,
                                                                  ticket_tb_id=ticket_info.ticket_id,
                                                                  lecture_tb_id=lecture_id, use=USE)
                            ticket_lecture_info.save()
        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=ticket_name+' 수강권', log_how='추가', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 삭제
def delete_ticket_info_logic(request):
    class_id = request.session.get('class_id', '')
    ticket_id = request.POST.get('ticket_id', '')
    error = None
    ticket_info = None
    now = timezone.now()

    try:
        ticket_info = TicketTb.objects.get(class_tb_id=class_id, ticket_id=ticket_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        try:
            with transaction.atomic():
                ticket_lecture_data = TicketLectureTb.objects.filter(class_tb_id=class_id, ticket_tb_id=ticket_id,
                                                                     use=USE)
                ticket_lecture_data.delete()

                ticket_info.state_cd = STATE_CD_FINISH
                ticket_info.use = UN_USE
                ticket_info.save()

                class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
                    'member_ticket_tb').filter(class_tb_id=class_id, member_ticket_tb__ticket_tb_id=ticket_id,
                                               auth_cd=AUTH_TYPE_VIEW, use=USE)

                if class_member_ticket_data is not None:
                    for class_member_ticket_info in class_member_ticket_data:
                        member_ticket_info = class_member_ticket_info.member_ticket_tb

                        schedule_data = ScheduleTb.objects.filter(
                            member_ticket_tb_id=member_ticket_info.member_ticket_id, end_dt__lte=now,
                            use=USE).exclude(Q(state_cd=STATE_CD_FINISH) | Q(state_cd=STATE_CD_ABSENCE))

                        schedule_data_delete = ScheduleTb.objects.filter(
                            member_ticket_tb_id=member_ticket_info.member_ticket_id, end_dt__gt=now,
                            use=USE).exclude(Q(state_cd=STATE_CD_FINISH) | Q(state_cd=STATE_CD_ABSENCE))

                        repeat_schedule_data = RepeatScheduleTb.objects.filter(
                            member_ticket_tb_id=member_ticket_info.member_ticket_id)

                        if len(schedule_data) > 0:
                            schedule_data.update(state_cd=STATE_CD_FINISH)
                        if len(schedule_data_delete) > 0:
                            schedule_data_delete.delete()
                        if len(repeat_schedule_data) > 0:
                            repeat_schedule_data.delete()

                        member_ticket_info.member_ticket_avail_count = 0
                        if member_ticket_info.state_cd == STATE_CD_IN_PROGRESS:
                            member_ticket_info.member_ticket_rem_count = 0
                            member_ticket_info.state_cd = STATE_CD_FINISH
                            member_ticket_info.use = UN_USE
                            member_ticket_info.save()

                        member_payment_history = MemberPaymentHistoryTb.objects.filter(
                            class_tb_id=class_id, member_ticket_tb_id=member_ticket_info.member_ticket_id, use=USE)
                        member_payment_history.update(use=UN_USE)

        except ValueError:
            error = '오류가 발생했습니다. [1]'
        except IntegrityError:
            error = '오류가 발생했습니다. [2]'
        except TypeError:
            error = '오류가 발생했습니다. [3]'
        except ValidationError:
            error = '오류가 발생했습니다. [4]'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=ticket_info.name+' 수강권', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권 수정
def update_ticket_info_logic(request):
    class_id = request.session.get('class_id', '')
    ticket_id = request.POST.get('ticket_id', '')
    ticket_name = request.POST.get('ticket_name', '')
    ticket_note = request.POST.get('ticket_note', '')
    ticket_effective_days = request.POST.get('ticket_effective_days', '')
    ticket_price = request.POST.get('ticket_price', '')
    ticket_month_schedule_enable = request.POST.get('ticket_month_schedule_enable', '')
    ticket_week_schedule_enable = request.POST.get('ticket_week_schedule_enable', '')
    ticket_day_schedule_enable = request.POST.get('ticket_day_schedule_enable', '')
    ticket_reg_count = request.POST.get('ticket_reg_count', '')
    error = None
    ticket_info = None

    try:
        ticket_info = TicketTb.objects.get(class_tb_id=class_id, ticket_id=ticket_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        if ticket_name != '' and ticket_name is not None:
            ticket_info.name = ticket_name
        if ticket_note != '' and ticket_note is not None:
            ticket_info.note = ticket_note
        if ticket_effective_days != '' and ticket_effective_days is not None:
            ticket_info.effective_days = ticket_effective_days
        if ticket_price != '' and ticket_price is not None:
            ticket_info.price = ticket_price
        if ticket_month_schedule_enable != '' and ticket_month_schedule_enable is not None:
            ticket_info.month_schedule_enable = ticket_month_schedule_enable
        if ticket_week_schedule_enable != '' and ticket_week_schedule_enable is not None:
            ticket_info.week_schedule_enable = ticket_week_schedule_enable
        if ticket_day_schedule_enable != '' and ticket_day_schedule_enable is not None:
            ticket_info.day_schedule_enable = ticket_day_schedule_enable
        if ticket_reg_count != '' and ticket_reg_count is not None:
            ticket_info.reg_count = ticket_reg_count
        ticket_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=ticket_info.name+' 수강권', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권에 그룹 추가
def add_ticket_lecture_info_logic(request):
    class_id = request.session.get('class_id', '')
    ticket_id = request.POST.get('ticket_id', '')
    lecture_id = request.POST.get('lecture_id', '')
    error = None
    ticket_lecture_counter = TicketLectureTb.objects.filter(ticket_tb_id=ticket_id, lecture_tb_id=lecture_id,
                                                            use=USE).count()
    if ticket_lecture_counter > 0:
        error = '이미 포함된 그룹입니다.'

    if error is None:
        try:
            with transaction.atomic():
                ticket_lecture_info = TicketLectureTb(class_tb_id=class_id, ticket_tb_id=ticket_id,
                                                      lecture_tb_id=lecture_id, use=USE)
                ticket_lecture_info.save()

        except ValueError:
            error = '오류가 발생했습니다.'
        except IntegrityError:
            error = '오류가 발생했습니다.'
        except TypeError:
            error = '오류가 발생했습니다.'
        except ValidationError:
            error = '오류가 발생했습니다.'
        except InternalError:
            error = '오류가 발생했습니다.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=ticket_lecture_info.ticket_tb.name + ' 수강권에 '
                                  + ticket_lecture_info.lecture_tb.name + '수업', log_how='추가', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 패키지에서 그룹 삭제
def delete_ticket_lecture_info_logic(request):
    class_id = request.session.get('class_id', '')
    ticket_id = request.POST.get('ticket_id', '')
    lecture_id = request.POST.get('lecture_id', '')
    ticket_name = ''
    lecture_name = ''
    error = None
    if error is None:
        try:
            with transaction.atomic():
                # 그룹에 패키지에서 제외되도록 설정
                try:
                    ticket_info = TicketTb.objects.get(ticket_id=ticket_id)
                    ticket_name = ticket_info.name
                except ObjectDoesNotExist:
                    ticket_name = ''
                try:
                    lecture_info = LectureTb.objects.get(lecture_id=lecture_id)
                    lecture_name = lecture_info.name
                except ObjectDoesNotExist:
                    lecture_name = ''
                ticket_lecture_data = TicketLectureTb.objects.filter(class_tb_id=class_id,
                                                                     ticket_tb_id=ticket_id, lecture_tb_id=lecture_id,
                                                                     use=USE)
                ticket_lecture_data.delete()

                schedule_data = ScheduleTb.objects.filter(class_tb_id=class_id,
                                                          member_ticket_tb__ticket_tb_id=ticket_id,
                                                          lecture_tb_id=lecture_id,
                                                          end_dt__lte=timezone.now(),
                                                          en_dis_type=ON_SCHEDULE_TYPE
                                                          ).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                    | Q(state_cd=STATE_CD_ABSENCE))

                schedule_data_delete = ScheduleTb.objects.filter(class_tb_id=class_id,
                                                                 member_ticket_tb__ticket_tb_id=ticket_id,
                                                                 lecture_tb_id=lecture_id,
                                                                 end_dt__gt=timezone.now(),
                                                                 en_dis_type=ON_SCHEDULE_TYPE
                                                                 ).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                           | Q(state_cd=STATE_CD_ABSENCE))

                repeat_schedule_data = RepeatScheduleTb.objects.filter(class_tb_id=class_id,
                                                                       member_ticket_tb__ticket_tb_id=ticket_id,
                                                                       lecture_tb_id=lecture_id)
                if len(schedule_data) > 0:
                    schedule_data.update(state_cd=STATE_CD_FINISH)
                if len(schedule_data_delete) > 0:
                    schedule_data_delete.delete()
                if len(repeat_schedule_data) > 0:
                    repeat_schedule_data.delete()

                # 해당 패키지의 회원들 수강정보 업데이트
                class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
                    'member_ticket_tb__ticket_tb').filter(class_tb_id=class_id,
                                                          member_ticket_tb__ticket_tb_id=ticket_id,
                                                          auth_cd=AUTH_TYPE_VIEW,
                                                          member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS,
                                                          member_ticket_tb__use=USE, use=USE)

                for class_member_ticket_info in class_member_ticket_data:
                    error = func_refresh_member_ticket_count(class_member_ticket_info.class_tb_id,
                                                             class_member_ticket_info.member_ticket_tb_id)

                if error is not None:
                    raise InternalError

        except ValueError:
            error = '오류가 발생했습니다.'
        except IntegrityError:
            error = '오류가 발생했습니다.'
        except TypeError:
            error = '오류가 발생했습니다.'
        except ValidationError:
            error = '오류가 발생했습니다.'
        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=ticket_name + ' 수강권에 '
                                  + lecture_name + '수업', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


def update_ticket_status_info_logic(request):
    class_id = request.session.get('class_id', '')
    ticket_id = request.POST.get('ticket_id', '')
    state_cd = request.POST.get('state_cd')

    error = None
    ticket_info = None
    ticket_lecture_data = None
    now = timezone.now()

    if ticket_id is None or ticket_id == '':
        error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            ticket_info = TicketTb.objects.get(ticket_id=ticket_id)
        except ObjectDoesNotExist:
            error = '수강권 정보를 불러오지 못했습니다.'

    if error is None:

        class_member_ticket_data = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb__member').filter(class_tb_id=class_id, member_ticket_tb__ticket_tb_id=ticket_id,
                                               auth_cd=AUTH_TYPE_VIEW,
                                               use=USE).order_by('member_ticket_tb__member__member_id')

        if class_member_ticket_data is not None and state_cd == STATE_CD_FINISH:
            member_id = None
            for class_member_ticket_info in class_member_ticket_data:
                member_ticket_info = class_member_ticket_info.member_ticket_tb
                schedule_data = ScheduleTb.objects.filter(member_ticket_tb_id=member_ticket_info.member_ticket_id,
                                                          end_dt__lte=now,
                                                          use=USE).exclude(Q(state_cd=STATE_CD_FINISH)
                                                                           | Q(state_cd=STATE_CD_ABSENCE))

                schedule_data_delete = ScheduleTb.objects.filter(
                    member_ticket_tb_id=member_ticket_info.member_ticket_id, end_dt__gt=now,
                    use=USE).exclude(Q(state_cd=STATE_CD_FINISH) | Q(state_cd=STATE_CD_ABSENCE))

                repeat_schedule_data = RepeatScheduleTb.objects.filter(
                    member_ticket_tb_id=member_ticket_info.member_ticket_id)

                if len(schedule_data) > 0:
                    schedule_data.update(state_cd=STATE_CD_FINISH)
                if len(schedule_data_delete) > 0:
                    schedule_data_delete.delete()
                if len(repeat_schedule_data) > 0:
                    repeat_schedule_data.delete()

                member_ticket_info.member_ticket_avail_count = 0
                if member_ticket_info.state_cd == STATE_CD_IN_PROGRESS:
                    member_ticket_info.member_ticket_rem_count = 0
                    member_ticket_info.state_cd = STATE_CD_FINISH
                member_ticket_info.save()
                if member_id != member_ticket_info.member_id:
                    member_id = member_ticket_info.member_id
                    func_update_lecture_member_fix_status_cd(class_id, member_id)

        ticket_info.state_cd = state_cd
        ticket_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=ticket_info.name + ' 수강권', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetTicketInfoViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id')
        ticket_id = request.GET.get('ticket_id')

        ticket_info = func_get_ticket_info(class_id, ticket_id, self.request.user.id)

        return JsonResponse({'ticket_info': ticket_info}, json_dumps_params={'ensure_ascii': True})


class GetTicketIngListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        # page = self.request.GET.get('page', 0)
        ticket_sort = self.request.GET.get('sort_val', SORT_TICKET_NAME)
        sort_order_by = self.request.GET.get('sort_order_by', SORT_ASC)
        keyword = self.request.GET.get('keyword', '')
        sort_info = int(ticket_sort)
        error = None
        # start_time = timezone.now()
        ticket_data = TicketTb.objects.filter(class_tb_id=class_id, state_cd=STATE_CD_IN_PROGRESS,
                                              name__contains=keyword,
                                              use=USE).order_by('ticket_id')
        ticket_lecture_data = TicketLectureTb.objects.select_related(
            'ticket_tb', 'lecture_tb').filter(class_tb_id=class_id, ticket_tb__state_cd=STATE_CD_IN_PROGRESS,
                                              ticket_tb__name__contains=keyword,
                                              ticket_tb__use=USE,
                                              use=USE).order_by('ticket_tb_id', 'lecture_tb__state_cd', 'lecture_tb_id')

        ticket_data_dict = {}
        for ticket_lecture_info in ticket_lecture_data:
            ticket_tb = ticket_lecture_info.ticket_tb
            lecture_tb = ticket_lecture_info.lecture_tb
            ticket_id = str(ticket_tb.ticket_id)
            try:
                ticket_data_dict[ticket_id]
            except KeyError:
                ticket_data_dict[ticket_id] = {'ticket_id': ticket_id,
                                               'ticket_name': ticket_tb.name,
                                               'ticket_note': ticket_tb.note,
                                               'ticket_effective_days': ticket_tb.effective_days,
                                               'ticket_price': ticket_tb.price,
                                               'ticket_reg_count': ticket_tb.reg_count,
                                               'ticket_month_schedule_enable': ticket_tb.month_schedule_enable,
                                               'ticket_week_schedule_enable': ticket_tb.week_schedule_enable,
                                               'ticket_day_schedule_enable': ticket_tb.day_schedule_enable,
                                               # 'ticket_type_cd': ticket_tb.ticket_type_cd,
                                               'ticket_reg_dt': ticket_tb.reg_dt,
                                               'ticket_lecture_list': [],
                                               'ticket_lecture_state_cd_list': [],
                                               'ticket_lecture_id_list': [],
                                               'ticket_lecture_ing_color_cd_list': [],
                                               'ticket_lecture_ing_font_color_cd_list': [],
                                               'ticket_lecture_end_color_cd_list': [],
                                               'ticket_lecture_end_font_color_cd_list': []}
            if lecture_tb.use == USE:
                ticket_data_dict[ticket_id]['ticket_lecture_list'].append(lecture_tb.name)
                ticket_data_dict[ticket_id]['ticket_lecture_state_cd_list'].append(lecture_tb.state_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_id_list'].append(lecture_tb.lecture_id)
                ticket_data_dict[ticket_id]['ticket_lecture_ing_color_cd_list'].append(lecture_tb.ing_color_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_ing_font_color_cd_list'].append(lecture_tb.ing_font_color_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_end_color_cd_list'].append(lecture_tb.end_color_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_end_font_color_cd_list'].append(lecture_tb.end_font_color_cd)

        if len(ticket_data) != len(ticket_data_dict):
            for ticket_info in ticket_data:
                ticket_id = str(ticket_info.ticket_id)
                try:
                    ticket_data_dict[ticket_id]
                except KeyError:
                    ticket_data_dict[ticket_id] = {'ticket_id': ticket_id,
                                                   'ticket_name': ticket_info.name,
                                                   'ticket_note': ticket_info.note,
                                                   'ticket_effective_days': ticket_info.effective_days,
                                                   'ticket_price': ticket_info.price,
                                                   'ticket_reg_count': ticket_info.reg_count,
                                                   'ticket_month_schedule_enable': ticket_info.month_schedule_enable,
                                                   'ticket_week_schedule_enable': ticket_info.week_schedule_enable,
                                                   'ticket_day_schedule_enable': ticket_info.day_schedule_enable,
                                                   # 'ticket_type_cd': ticket_info.ticket_type_cd,
                                                   'ticket_reg_dt': ticket_info.reg_dt,
                                                   'ticket_lecture_list': [],
                                                   'ticket_lecture_state_cd_list': [],
                                                   'ticket_lecture_id_list': [],
                                                   'ticket_lecture_ing_color_cd_list': [],
                                                   'ticket_lecture_ing_font_color_cd_list': [],
                                                   'ticket_lecture_end_color_cd_list': [],
                                                   'ticket_lecture_end_font_color_cd_list': []}

        ticket_list = []
        class_member_ticket_list = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb__ticket_tb',
            'member_ticket_tb__member').filter(Q(member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS)
                                               | Q(member_ticket_tb__state_cd=STATE_CD_HOLDING),
                                               class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW,
                                               member_ticket_tb__ticket_tb__state_cd=STATE_CD_IN_PROGRESS,
                                               member_ticket_tb__ticket_tb__use=USE,
                                               member_ticket_tb__use=USE,
                                               use=USE).order_by('member_ticket_tb__ticket_tb',
                                                                 'member_ticket_tb__member')

        for ticket_info in ticket_data_dict:
            member_list = {}
            for class_member_ticket_info in class_member_ticket_list:
                ticket_id = class_member_ticket_info.member_ticket_tb.ticket_tb_id
                if ticket_id == ticket_info:
                    member_id = class_member_ticket_info.member_ticket_tb.member_id
                    member_list[member_id] = member_id

            ticket_data_dict[ticket_info]['ticket_ing_member_num'] = len(member_list)
            ticket_list.append(ticket_data_dict[ticket_info])

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        else:
            # if sort_info == SORT_TICKET_TYPE:
            #     ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_type_cd'],
            #                          reverse=int(sort_order_by))
            if sort_info == SORT_TICKET_NAME:
                ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_name'],
                                     reverse=int(sort_order_by))
            elif sort_info == SORT_TICKET_MEMBER_COUNT:
                ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_ing_member_num'],
                                     reverse=int(sort_order_by))
            elif sort_info == SORT_TICKET_CREATE_DATE:
                ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_reg_dt'],
                                     reverse=int(sort_order_by))

        # end_time = timezone.now()
        return JsonResponse({'current_ticket_data': ticket_list}, json_dumps_params={'ensure_ascii': True})


class GetTicketEndListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')

        # page = self.request.GET.get('page', 0)
        ticket_sort = self.request.GET.get('sort_val', SORT_TICKET_TYPE)
        sort_order_by = self.request.GET.get('sort_order_by', SORT_ASC)
        keyword = self.request.GET.get('keyword', '')
        sort_info = int(ticket_sort)

        error = None

        ticket_data = TicketTb.objects.filter(class_tb_id=class_id, state_cd=STATE_CD_FINISH,
                                              name__contains=keyword,
                                              use=USE).order_by('ticket_id')
        ticket_lecture_data = TicketLectureTb.objects.select_related(
            'ticket_tb', 'lecture_tb').filter(class_tb_id=class_id, ticket_tb__state_cd=STATE_CD_FINISH,
                                              ticket_tb__name__contains=keyword,
                                              ticket_tb__use=USE, use=USE).order_by('ticket_tb_id',
                                                                                    'lecture_tb__state_cd',
                                                                                    'lecture_tb_id')

        ticket_data_dict = {}
        for ticket_lecture_info in ticket_lecture_data:
            ticket_tb = ticket_lecture_info.ticket_tb
            lecture_tb = ticket_lecture_info.lecture_tb
            ticket_id = str(ticket_tb.ticket_id)
            try:
                ticket_data_dict[ticket_id]
            except KeyError:
                ticket_data_dict[ticket_id] = {'ticket_id': ticket_id,
                                               'ticket_name': ticket_tb.name,
                                               'ticket_note': ticket_tb.note,
                                               'ticket_effective_days': ticket_tb.effective_days,
                                               'ticket_price': ticket_tb.price,
                                               'ticket_reg_count': ticket_tb.reg_count,
                                               'ticket_month_schedule_enable': ticket_tb.month_schedule_enable,
                                               'ticket_week_schedule_enable': ticket_tb.week_schedule_enable,
                                               'ticket_day_schedule_enable': ticket_tb.day_schedule_enable,
                                               'ticket_reg_dt': ticket_tb.reg_dt,
                                               'ticket_lecture_list': [],
                                               'ticket_lecture_state_cd_list': [],
                                               'ticket_lecture_id_list': [],
                                               'ticket_lecture_ing_color_cd_list': [],
                                               'ticket_lecture_ing_font_color_cd_list': [],
                                               'ticket_lecture_end_color_cd_list': [],
                                               'ticket_lecture_end_font_color_cd_list': []}
            if lecture_tb.use == USE:
                ticket_data_dict[ticket_id]['ticket_lecture_list'].append(lecture_tb.name)
                ticket_data_dict[ticket_id]['ticket_lecture_state_cd_list'].append(lecture_tb.state_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_id_list'].append(lecture_tb.lecture_id)
                ticket_data_dict[ticket_id]['ticket_lecture_ing_color_cd_list'].append(lecture_tb.ing_color_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_ing_font_color_cd_list'].append(lecture_tb.ing_font_color_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_end_color_cd_list'].append(lecture_tb.end_color_cd)
                ticket_data_dict[ticket_id]['ticket_lecture_end_font_color_cd_list'].append(lecture_tb.end_font_color_cd)

        if len(ticket_data) != len(ticket_data_dict):
            for ticket_info in ticket_data:
                ticket_id = str(ticket_info.ticket_id)
                try:
                    ticket_data_dict[ticket_id]
                except KeyError:
                    ticket_data_dict[ticket_id] = {'ticket_id': ticket_id,
                                                   'ticket_name': ticket_info.name,
                                                   'ticket_note': ticket_info.note,
                                                   'ticket_effective_days': ticket_info.effective_days,
                                                   'ticket_price': ticket_info.price,
                                                   'ticket_reg_count': ticket_info.reg_count,
                                                   'ticket_month_schedule_enable': ticket_info.month_schedule_enable,
                                                   'ticket_week_schedule_enable': ticket_info.week_schedule_enable,
                                                   'ticket_day_schedule_enable': ticket_info.day_schedule_enable,
                                                   'ticket_reg_dt': ticket_info.reg_dt,
                                                   'ticket_lecture_list': [],
                                                   'ticket_lecture_state_cd_list': [],
                                                   'ticket_lecture_id_list': [],
                                                   'ticket_lecture_ing_color_cd_list': [],
                                                   'ticket_lecture_ing_font_color_cd_list': [],
                                                   'ticket_lecture_end_color_cd_list': [],
                                                   'ticket_lecture_end_font_color_cd_list': []}

        ticket_list = []
        # class_member_ticket_list = ClassMemberTicketTb.objects.select_related(
        #     'member_ticket_tb__ticket_tb',
        #     'member_ticket_tb__member').filter(class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW,
        #                                        member_ticket_tb__ticket_tb__state_cd=STATE_CD_FINISH,
        #                                        member_ticket_tb__ticket_tb__use=USE,
        #                                        member_ticket_tb__use=USE,
        #                                        use=USE).order_by('member_ticket_tb__ticket_tb',
        #                                                          'member_ticket_tb__member')

        for ticket_info in ticket_data_dict:
            # member_list = {}
            # for class_member_ticket_info in class_member_ticket_list:
            #     ticket_id = class_member_ticket_info.member_ticket_tb.ticket_tb_id
            #     if ticket_id == ticket_info:
            #         member_id = class_member_ticket_info.member_ticket_tb.member_id
            #         member_list[member_id] = member_id
            # ticket_data_dict[ticket_info]['ticket_end_member_num'] = len(member_list)
            # ticket_data_dict[ticket_info]['ticket_ing_member_num'] = 0
            ticket_data_dict[ticket_info]['ticket_end_member_num'] = 0
            ticket_list.append(ticket_data_dict[ticket_info])

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        else:
            # if sort_info == SORT_TICKET_TYPE:
            #     ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_type_cd'],
            #                          reverse=int(sort_order_by))
            if sort_info == SORT_TICKET_NAME:
                ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_name'],
                                     reverse=int(sort_order_by))
            elif sort_info == SORT_TICKET_MEMBER_COUNT:
                ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_name'],
                                     reverse=int(sort_order_by))
            elif sort_info == SORT_TICKET_CREATE_DATE:
                ticket_list = sorted(ticket_list, key=lambda elem: elem['ticket_reg_dt'],
                                     reverse=int(sort_order_by))

        return JsonResponse({'finish_ticket_data': ticket_list}, json_dumps_params={'ensure_ascii': True})


class GetTicketIngMemberListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        # context = {}
        class_id = self.request.session.get('class_id', '')
        ticket_id = self.request.GET.get('ticket_id', '')
        error = None
        # member_data = []

        all_class_member_ticket_list = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb__ticket_tb',
            'member_ticket_tb__member__user'
        ).filter(Q(member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS)|Q(member_ticket_tb__state_cd=STATE_CD_HOLDING),
                 class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW, member_ticket_tb__ticket_tb_id=ticket_id,
                 member_ticket_tb__use=USE, use=USE).order_by('member_ticket_tb__member_id',
                                                              'member_ticket_tb__end_date')
        member_list = func_get_member_from_member_ticket_list(all_class_member_ticket_list, None, request.user.id)

        if error is not None:
            logger.error(self.request.user.first_name + '[' + str(self.request.user.id) + ']' + error)
            messages.error(self.request, error)

        return JsonResponse({'ticket_ing_member_list': member_list}, json_dumps_params={'ensure_ascii': True})


class GetTicketEndMemberListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        # context = {}
        class_id = self.request.session.get('class_id', '')
        ticket_id = self.request.GET.get('ticket_id', '')
        error = None

        query_member_ticket_ip_count = "select count(*) from "+MemberTicketTb._meta.db_table+" as C where C." + \
                                       MemberTicketTb._meta.get_field('member').column + \
                                       "=(select B."+MemberTicketTb._meta.get_field('member').column+" from " + \
                                       MemberTicketTb._meta.db_table+" as B where B." + \
                                       MemberTicketTb._meta.get_field('member_ticket_id').column+" =" + \
                                       " `"+ClassMemberTicketTb._meta.db_table+"`.`" + \
                                       ClassMemberTicketTb._meta.get_field('member_ticket_tb').column+"`)" \
                                       " and " + str(class_id) + \
                                       " = (select D."+ClassMemberTicketTb._meta.get_field('class_tb').column + \
                                       " from "+ClassMemberTicketTb._meta.db_table+" as D" \
                                       " where D."+ClassMemberTicketTb._meta.get_field('member_ticket_tb').column + \
                                       "=C."+ClassMemberTicketTb._meta.get_field('class_member_ticket_id').column + \
                                       " and D."+ClassMemberTicketTb._meta.get_field('class_tb').column + \
                                       "=" + str(class_id) + ")" \
                                       " and C."+MemberTicketTb._meta.get_field('state_cd').column + \
                                       "=\'IP\' and C."+MemberTicketTb._meta.get_field('ticket_tb').column + \
                                       "=" + str(ticket_id) + " and C." + \
                                       MemberTicketTb._meta.get_field('use').column + "="+str(USE)

        all_class_member_ticket_list = ClassMemberTicketTb.objects.select_related(
            'member_ticket_tb__ticket_tb',
            'member_ticket_tb__member__user'
        ).filter(~Q(member_ticket_tb__state_cd=STATE_CD_IN_PROGRESS),~Q(member_ticket_tb__state_cd=STATE_CD_HOLDING),
                 class_tb_id=class_id, auth_cd=AUTH_TYPE_VIEW,
                 member_ticket_tb__ticket_tb_id=ticket_id, member_ticket_tb__use=USE,
                 use=USE).annotate(member_ticket_ip_count=RawSQL(query_member_ticket_ip_count, [])
                                   ).filter(member_ticket_ip_count=0).order_by('member_ticket_tb__member_id',
                                                                               'member_ticket_tb__end_date')

        member_list = func_get_member_from_member_ticket_list(all_class_member_ticket_list, None, request.user.id)

        if error is not None:
            logger.error(self.request.user.first_name + '[' + str(self.request.user.id) + ']' + error)
            messages.error(self.request, error)

        return JsonResponse({'ticket_end_member_list': member_list}, json_dumps_params={'ensure_ascii': True})


class GetProgramListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        program_list = []
        error = None
        program_data = MemberClassTb.objects.select_related('class_tb', 'member'
                                                            ).filter(member_id=self.request.user.id,
                                                                     auth_cd__contains=AUTH_TYPE_VIEW,
                                                                     use=USE).order_by('-reg_dt')

        if error is None:
            for program_info in program_data:
                all_member = func_get_class_member_ing_list(program_info.class_tb.class_id, '')
                total_member_num = len(all_member)
                program_selected = 'NOT_SELECTED'
                if str(class_id) == str(program_info.class_tb.class_id):
                    program_selected = 'SELECTED'

                program_subject_type_name = program_info.class_tb.get_class_type_cd_name()
                shared_program_flag = MY_PROGRAM
                share_member_num = 0
                if str(program_info.class_tb.member.member_id) != str(request.user.id):
                    # program_subject_type_name += ' - ' + program_info.class_tb.member.name
                    shared_program_flag = SHARED_PROGRAM
                else:
                    share_member_num = MemberClassTb.objects.filter(class_tb_id=program_info.class_tb.class_id,
                                                                    auth_cd__contains=AUTH_TYPE_VIEW,
                                                                    own_cd=OWN_TYPE_SHARE).count()

                program_dict = {'program_id': program_info.class_tb.class_id,
                                'program_total_member_num': total_member_num,
                                'program_state_cd': program_info.class_tb.state_cd,
                                'program_subject_cd': program_info.class_tb.subject_cd,
                                'program_subject_cd_name': program_info.class_tb.get_subject_type_name(),
                                'program_subject_type_name': program_subject_type_name,
                                'program_upper_subject_cd': program_info.class_tb.get_upper_class_type_cd(),
                                'program_upper_subject_type_name': program_info.class_tb.get_upper_class_type_cd_name(),
                                'program_selected': program_selected,
                                'program_program_owner_id': program_info.class_tb.member_id,
                                'program_program_owner_name': program_info.class_tb.member.name,
                                'shared_program_flag': shared_program_flag,
                                'share_member_num': share_member_num,
                                'program_center_name': ''
                                }
                program_list.append(program_dict)

        if error is not None:
            messages.error(self.request, error)

        return JsonResponse({'program_data': program_list}, json_dumps_params={'ensure_ascii': True})


class GetProgramCategoryViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        program_category_dict = collections.OrderedDict()
        error = None
        program_upper_category_data = CommonCdTb.objects.filter(upper_common_cd='02', use=USE).order_by('order')
        program_sub_category_data = CommonCdTb.objects.filter(upper_common_cd='03', use=USE).order_by('order')

        if error is None:
            for program_upper_category_info in program_upper_category_data:
                try:
                    program_category_dict[program_upper_category_info.common_cd]
                except KeyError:
                    program_category_dict[program_upper_category_info.common_cd] = {'name': program_upper_category_info.common_cd_nm,
                                                                                    'sub_category': {}}
            for program_sub_category_info in program_sub_category_data:

                try:
                    program_category_dict[program_sub_category_info.group_cd]['sub_category'][program_sub_category_info.common_cd]
                except KeyError:
                    program_category_dict[program_sub_category_info.group_cd]['sub_category'][program_sub_category_info.common_cd] = {'name': program_sub_category_info.common_cd_nm}

        if error is not None:
            messages.error(self.request, error)

        return JsonResponse({'program_category_data': program_category_dict}, json_dumps_params={'ensure_ascii': True})


def add_program_info_logic(request):
    template_name = 'ajax/trainer_error_ajax.html'
    center_id = request.POST.get('center_id', '')
    subject_cd = request.POST.get('subject_cd', '')
    subject_detail_nm = request.POST.get('subject_detail_nm', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    class_hour = request.POST.get('class_hour', 60)
    start_hour_unit = request.POST.get('start_hour_unit', 1)
    class_member_num = request.POST.get('class_member_num', '')

    error = None
    class_info = None

    if subject_cd is None or subject_cd == '':
        error = '지점 종류를 설정해주세요.'

    if class_hour is None or class_hour == '':
        class_hour = 60

    if start_hour_unit is None or start_hour_unit == '':
        start_hour_unit = 1

    if start_date is None or start_date == '':
        start_date = datetime.date.today()
    if end_date is None or end_date == '':
        end_date = start_date + timezone.timedelta(days=3650)

    if subject_detail_nm is None:
        subject_detail_nm = ''

    if class_member_num is None or class_member_num == '':
        error = '수업당 최대 허용 인원을 설정해 주세요.'

    if error is None:
        try:
            with transaction.atomic():

                class_info = ClassTb(member_id=request.user.id, center_tb_id=center_id,
                                     subject_cd=subject_cd, start_date=start_date, end_date=end_date,
                                     class_hour=class_hour, start_hour_unit=start_hour_unit,
                                     subject_detail_nm=subject_detail_nm,
                                     class_member_num=int(class_member_num), state_cd=STATE_CD_IN_PROGRESS, use=USE)

                class_info.save()

                member_class_info = MemberClassTb(member_id=request.user.id, class_tb_id=class_info.class_id,
                                                  auth_cd=AUTH_TYPE_VIEW, mod_member_id=request.user.id,
                                                  own_cd=OWN_TYPE_OWNER, use=USE)
                member_class_info.save()

                one_to_one_lecture_info = LectureTb(class_tb_id=class_info.class_id, name='개인 수업',
                                                    ing_color_cd='#fbf3bd', end_color_cd='#d2d1cf',
                                                    ing_font_color_cd='#282828', end_font_color_cd='#282828',
                                                    state_cd=STATE_CD_IN_PROGRESS,
                                                    lecture_minute=60, main_trainer_id=request.user.id,
                                                    lecture_type_cd=LECTURE_TYPE_ONE_TO_ONE, member_num=1, use=USE)
                one_to_one_lecture_info.save()

                ticket_info = TicketTb(class_tb_id=class_info.class_id, name='개인 수업 수강권',
                                       state_cd=STATE_CD_IN_PROGRESS, use=USE)
                ticket_info.save()

                ticket_lecture_info = TicketLectureTb(class_tb_id=class_info.class_id,
                                                      ticket_tb_id=ticket_info.ticket_id,
                                                      lecture_tb_id=one_to_one_lecture_info.lecture_id, use=USE)
                ticket_lecture_info.save()

                group_lecture_info = LectureTb(class_tb_id=class_info.class_id, name='2:1 그룹 수업',
                                               ing_color_cd='#d8d6ff', end_color_cd='#d2d1cf',
                                               ing_font_color_cd='#282828', end_font_color_cd='#282828',
                                               state_cd=STATE_CD_IN_PROGRESS, lecture_minute=60,
                                               main_trainer_id=request.user.id,
                                               lecture_type_cd=LECTURE_TYPE_NORMAL, member_num=2, use=USE)
                group_lecture_info.save()

                ticket_group_info = TicketTb(class_tb_id=class_info.class_id, name='2:1 그룹 수업 수강권',
                                             state_cd=STATE_CD_IN_PROGRESS, use=USE)
                ticket_group_info.save()

                ticket_lecture_info = TicketLectureTb(class_tb_id=class_info.class_id,
                                                      ticket_tb_id=ticket_group_info.ticket_id,
                                                      lecture_tb_id=group_lecture_info.lecture_id, use=USE)
                ticket_lecture_info.save()

                ticket_total_info = TicketTb(class_tb_id=class_info.class_id, name='통합 수강권',
                                             state_cd=STATE_CD_IN_PROGRESS, use=USE)
                ticket_total_info.save()

                ticket_lecture_info = TicketLectureTb(class_tb_id=class_info.class_id,
                                                      ticket_tb_id=ticket_total_info.ticket_id,
                                                      lecture_tb_id=one_to_one_lecture_info.lecture_id, use=USE)
                ticket_lecture_info.save()

                ticket_lecture_info = TicketLectureTb(class_tb_id=class_info.class_id,
                                                      ticket_tb_id=ticket_total_info.ticket_id,
                                                      lecture_tb_id=group_lecture_info.lecture_id, use=USE)
                ticket_lecture_info.save()

        except ValueError:
            error = '등록 값에 문제가 있습니다.'
        except IntegrityError:
            error = '등록 값에 문제가 있습니다.'
        except TypeError:
            error = '등록 값에 문제가 있습니다.'
        except ValidationError:
            error = '등록 값에 문제가 있습니다.'
        except InternalError:
            error = '등록 값에 문제가 있습니다.'

    # if error is None:
    #     request.session['class_id'] = class_info.class_id
    #     request.session['class_hour'] = class_info.class_hour
    #     request.session['class_type_code'] = class_info.subject_cd
    #     request.session['class_type_name'] = class_info.get_class_type_cd_name()
    #     request.session['class_center_name'] = class_info.get_center_name()
    #     request.session['trainer_name'] = class_info.member.name

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    return render(request, template_name)


def delete_program_info_logic(request):
    template_name = 'ajax/trainer_error_ajax.html'

    class_id = request.POST.get('class_id', '')
    class_id_session = request.session.get('class_id', '')

    error = None
    class_info = None
    if class_id is None or class_id == '':
        error = '지점 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            class_info = MemberClassTb.objects.get(member_id=request.user.id, class_tb_id=class_id)
        except ObjectDoesNotExist:
            error = '지점 정보를 불러오지 못했습니다.'

    if error is None:
        class_info.auth_cd = AUTH_TYPE_DELETE
        class_info.save()

    if error is None:
        if str(class_id) == str(class_id_session):
            request.session['class_id'] = ''
            request.session['class_type_code'] = ''
            request.session['class_type_name'] = ''
            request.session['class_center_name'] = ''
            request.session['trainer_name'] = ''

    if error is None:
        log_data = LogTb(log_type='LC02', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info='지점', log_how='삭제', use=USE)

        log_data.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, template_name)


def update_program_info_logic(request):
    template_name = 'ajax/trainer_error_ajax.html'
    class_id = request.POST.get('class_id', '')
    class_id_session = request.session.get('class_id', '')
    subject_cd = request.POST.get('subject_cd', '')
    subject_detail_nm = request.POST.get('subject_detail_nm', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    class_hour = request.POST.get('class_hour', '')
    start_hour_unit = request.POST.get('start_hour_unit', '')
    class_member_num = request.POST.get('class_member_num', '')

    error = None
    class_info = None

    if class_id is None or class_id == '':
        error = '오류가 발생했습니다.'

    if error is None:
        try:
            class_info = ClassTb.objects.get(class_id=class_id)
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다.'

    if error is None:

        if subject_cd is not None and subject_cd != '':
            class_info.subject_cd = subject_cd

        if class_hour is not None and class_hour != '':
            class_info.class_hour = class_hour

        if start_hour_unit is not None and start_hour_unit != '':
            class_info.start_hour_unit = start_hour_unit

        if start_date is not None and start_date != '':
            class_info.start_date = start_date
        if end_date is not None and end_date != '':
            class_info.end_date = end_date

        if subject_detail_nm is not None and subject_detail_nm != '':
            class_info.subject_detail_nm = subject_detail_nm

        if class_member_num is not None and class_member_num != '':
            class_info.class_member_num = class_member_num

        class_info.save()

    if error is None:
        if str(class_id) == str(class_id_session):
            request.session['class_type_code'] = class_info.subject_cd
            request.session['class_type_name'] = class_info.get_class_type_cd_name()
            request.session['class_hour'] = class_info.class_hour

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, template_name)


def update_share_program_info_logic(request):
    trainer_id = request.POST.get('trainer_id', '')
    auth_cd = request.POST.get('auth_cd', '')
    class_id = request.POST.get('class_id', '')
    error = None
    member_info = None
    if trainer_id is None or trainer_id == '':
        error = '강사 정보를 불러오지 못했습니다.'
    if error is None:
        try:
            member_info = MemberTb.objects.get(member_id=trainer_id)
        except ObjectDoesNotExist:
            error = '강사 정보를 불러오지 못했습니다.'
    if error is None:
        if auth_cd != AUTH_TYPE_VIEW and auth_cd != AUTH_TYPE_WAIT and auth_cd != AUTH_TYPE_DELETE:
            error = '지점 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            class_member_info = MemberClassTb.objects.select_related(
                'class_tb', 'member').get(class_tb_id=class_id, member_id=trainer_id, use=USE)
            class_member_info.auth_cd = auth_cd
            class_member_info.mod_member_id = request.user.id
            class_member_info.save()
        except ObjectDoesNotExist:
            class_member_info = MemberClassTb(class_tb_id=class_id, member_id=trainer_id, auth_cd=auth_cd,
                                              mod_member_id=request.user.id, own_cd=OWN_TYPE_SHARE, use=USE)
            class_member_info.save()

        if auth_cd == AUTH_TYPE_DELETE:
            schedule_alarm_data = ScheduleAlarmTb.objects.filter(class_tb_id=class_id, member_id=trainer_id)
            schedule_alarm_data.delete()
            member_program_setting_data = SettingTb.objects.filter(class_tb_id=class_id, member_id=trainer_id)
            member_program_setting_data.delete()

        function_list = ProductFunctionAuthTb.objects.select_related(
            'function_auth_tb', 'product_tb').filter(product_tb_id=6,
                                                     use=USE).order_by('product_tb_id',
                                                                       'function_auth_tb_id',
                                                                       'auth_type_cd')
        for function_info in function_list:
            if function_info.auth_type_cd is None:
                function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd)
            else:
                function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd) \
                                             + str(function_info.auth_type_cd)
            if auth_cd != AUTH_TYPE_DELETE:
                enable_flag = request.POST.get(function_auth_type_cd_name, '')
                if enable_flag is not None and enable_flag != '':
                    try:
                        program_auth_info = ProgramAuthTb.objects.get(class_tb_id=class_id, member_id=trainer_id,
                                                                      function_auth_tb=function_info.function_auth_tb,
                                                                      auth_type_cd=function_info.auth_type_cd)
                    except ObjectDoesNotExist:
                        program_auth_info = ProgramAuthTb(class_tb_id=class_id, member_id=trainer_id,
                                                          function_auth_tb=function_info.function_auth_tb,
                                                          auth_type_cd=function_info.auth_type_cd,
                                                          use=USE)

                    program_auth_info.enable_flag = enable_flag
                    program_auth_info.save()

            else:
                try:
                    program_auth_info = ProgramAuthTb.objects.get(class_tb_id=class_id, member_id=trainer_id,
                                                                  function_auth_tb=function_info.function_auth_tb,
                                                                  auth_type_cd=function_info.auth_type_cd)
                    program_auth_info.delete()
                except ObjectDoesNotExist:
                    program_auth_info = None

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_how = '공유 신청'
        if auth_cd == AUTH_TYPE_DELETE:
            log_how = '공유 해제'
        log_data = LogTb(log_type='LP02', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name + '님 에게 \''
                                  + request.session.get('class_type_name', '') + '\' 지점',
                         log_how=log_how, log_detail='', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetShareProgramDataViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.GET.get('class_id', '')
        error = None
        member_program_auth_list = collections.OrderedDict()

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if error is None:
            member_program_data = MemberClassTb.objects.select_related(
                'class_tb', 'member').filter(class_tb_id=class_id, own_cd=OWN_TYPE_EMPLOYEE,
                                             use=USE)
            exclude_member = Q()
            for member_program_info in member_program_data:
                exclude_member |= Q(member_id=member_program_info.member_id)
            program_auth_data = ProgramAuthTb.objects.select_related('class_tb', 'member',
                                                                     'function_auth_tb').filter(class_tb_id=class_id,
                                                                                                use=USE).exclude(exclude_member)

            for program_auth_info in program_auth_data:

                if program_auth_info.auth_type_cd is None:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd)
                else:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd) \
                                                 + str(program_auth_info.auth_type_cd)

                try:
                    member_program_auth_list[program_auth_info.member_id]
                except KeyError:
                    member_program_auth_list[program_auth_info.member_id] = {}
                    member_result = func_get_trainer_info(class_id, program_auth_info.member_id)
                    member_program_auth_list[program_auth_info.member_id]['member_info'] \
                        = member_result['member_info']

                member_program_auth_list[program_auth_info.member_id][function_auth_type_cd_name] \
                    = program_auth_info.enable_flag

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        return JsonResponse(member_program_auth_list, json_dumps_params={'ensure_ascii': True})


class GetSharedProgramDataViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.GET.get('class_id', '')
        error = None
        member_program_auth_list = collections.OrderedDict()

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if error is None:
            program_auth_data = ProgramAuthTb.objects.select_related('class_tb', 'member',
                                                                     'function_auth_tb').filter(class_tb_id=class_id,
                                                                                                member_id=request.user.id,
                                                                                                use=USE)

            for program_auth_info in program_auth_data:
                if program_auth_info.auth_type_cd is None:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd)
                else:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd) \
                                                 + str(program_auth_info.auth_type_cd)

                member_program_auth_list[function_auth_type_cd_name] \
                    = program_auth_info.enable_flag

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        return JsonResponse(member_program_auth_list, json_dumps_params={'ensure_ascii': True})


def update_trainer_program_connection_info_logic(request):

    class_id = request.POST.get('class_id', '')
    program_connection_check = request.POST.get('program_connection_check', PROGRAM_SELECT)
    error = None
    member_ticket_connection_check = int(program_connection_check)
    if class_id == '':
        error = '지점 정보를 불러오지 못했습니다.[0]'

    if error is None:
        if member_ticket_connection_check == PROGRAM_LECTURE_CONNECT_DELETE:
            # 선택한 지점의 연결 대기중인 수강권 전부 삭제
            member_program_data = MemberClassTb.objects.select_related(
                'class_tb', 'member'
            ).filter(class_tb_id=class_id, member_id=request.user.id, use=USE)
            member_program_data.update(auth_cd=AUTH_TYPE_DELETE)

            schedule_alarm_data = ScheduleAlarmTb.objects.filter(class_tb_id=class_id, member_id=request.user.id)
            schedule_alarm_data.delete()
            member_program_setting_data = SettingTb.objects.filter(class_tb_id=class_id, member_id=request.user.id)
            member_program_setting_data.delete()

            class_info = None
            try:
                class_info = ClassTb.objects.get(class_id=class_id)
            except ObjectDoesNotExist:
                error = '지점 정보를 불러오지 못했습니다.[1]'

            log_data = LogTb(log_type='LP02', auth_member_id=request.user.id,
                             from_member_name=request.user.first_name,
                             class_tb_id=class_info.class_id,
                             log_info=class_info.member.name + ' 님의 \''
                                      + class_info.get_class_type_cd_name()+'\' 지점',
                             log_how='연결 신청 거절',
                             log_detail='', use=USE)
            log_data.save()

        elif member_ticket_connection_check == PROGRAM_LECTURE_CONNECT_ACCEPT:
            # 선택한 지점의 연결 대기중인 수강권 전부 연결
            member_program_data = MemberClassTb.objects.select_related(
                'class_tb', 'member'
            ).filter(class_tb_id=class_id, member_id=request.user.id, auth_cd=AUTH_TYPE_WAIT, use=USE)
            member_program_data.update(auth_cd=AUTH_TYPE_VIEW)

            class_info = None
            try:
                class_info = ClassTb.objects.get(class_id=class_id)
            except ObjectDoesNotExist:
                error = '지점 정보를 불러오지 못했습니다.[2]'

            log_data = LogTb(log_type='LP01', auth_member_id=request.user.id,
                             from_member_name=request.user.first_name,
                             class_tb_id=class_info.class_id,
                             log_info=class_info.member.name + ' 님의 \''
                                      + class_info.get_class_type_cd_name()+'\' 지점',
                             log_how='연결 신청 수락',
                             log_detail='', use=USE)
            log_data.save()

    if error is not None:
        logger.error(request.user.first_name+'['+str(request.user.id)+']'+error)
        messages.error(request, error)
    return render(request, 'ajax/trainer_error_ajax.html')


def delete_trainer_program_connection_logic(request):

    class_id = request.POST.get('class_id', '')
    error = None
    if class_id == '':
        error = '지점 정보를 불러오지 못했습니다.[0]'
    if error is None:
        # 선택한 지점의 연결 대기중인 수강권 전부 삭제
        member_program_data = MemberClassTb.objects.select_related(
            'class_tb', 'member'
        ).filter(class_tb_id=class_id, member_id=request.user.id, use=USE)
        member_program_data.update(auth_cd=AUTH_TYPE_DELETE)

        schedule_alarm_data = ScheduleAlarmTb.objects.filter(class_tb_id=class_id, member_id=request.user.id)
        schedule_alarm_data.delete()
        member_program_setting_data = SettingTb.objects.filter(class_tb_id=class_id, member_id=request.user.id)
        member_program_setting_data.delete()

        class_info = None
        try:
            class_info = ClassTb.objects.get(class_id=class_id)
        except ObjectDoesNotExist:
            error = '지점 정보를 불러오지 못했습니다.[1]'

        if error is None:
            log_data = LogTb(log_type='LP02', auth_member_id=request.user.id,
                             from_member_name=request.user.first_name,
                             class_tb_id=class_info.class_id,
                             log_info=class_info.member.name + ' 님의 \''
                                      + class_info.get_class_type_cd_name()+'\' 지점',
                             log_how='연결 해제',
                             log_detail='', use=USE)
            log_data.save()
            request.session['class_id'] = None

    if error is not None:
        logger.error(request.user.first_name+'['+str(request.user.id)+']'+error)
        messages.error(request, error)
    return render(request, 'ajax/trainer_error_ajax.html')


def refresh_trainer_page_logic(request):

    request.session['class_id'] = None

    return redirect('/')


class GetTrainerProgramConnectionListView(LoginRequiredMixin, AccessTestMixin, TemplateView):

    def get(self, request):
        member_program_auth_list = collections.OrderedDict()
        member_program_data = MemberClassTb.objects.select_related('class_tb',
                                                                   'member').filter(member_id=request.user.id,
                                                                                    auth_cd=AUTH_TYPE_WAIT)
        # 지점에 따른 권한 추가
        for member_program_info in member_program_data:
            program_auth_data = ProgramAuthTb.objects.select_related(
                'member', 'function_auth_tb').filter(class_tb_id=member_program_info.class_tb_id, use=USE)

            for program_auth_info in program_auth_data:
                if program_auth_info.auth_type_cd is None:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd)
                else:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd) \
                                                 + str(program_auth_info.auth_type_cd)

                try:
                    member_program_auth_list[member_program_info.class_tb_id]
                except KeyError:
                    member_program_auth_list[member_program_info.class_tb_id] = {}
                    member_result = func_get_trainer_info(member_program_info.class_tb_id,
                                                          member_program_info.class_tb.member_id)
                    member_program_auth_list[member_program_info.class_tb_id]['member_info'] \
                        = member_result['member_info']
                    member_program_auth_list[member_program_info.class_tb_id]['program_name'] \
                        = member_program_info.class_tb.get_class_type_cd_name()

                member_program_auth_list[member_program_info.class_tb_id]['connection_type']\
                    = member_program_info.own_cd
                member_program_auth_list[member_program_info.class_tb_id][function_auth_type_cd_name] \
                    = program_auth_info.enable_flag

        return JsonResponse(member_program_auth_list, json_dumps_params={'ensure_ascii': True})


def select_program_processing_logic(request):
    class_id = request.GET.get('class_id', '')
    next_page = request.GET.get('next_page', '')

    error = None
    class_info = None

    if class_id == '':
        error = '지점을 선택해 주세요.'

    if error is None:
        try:
            class_info = ClassTb.objects.get(class_id=class_id)
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다.'

    if error is None:
        request.session['class_id'] = class_id
        request.session['class_hour'] = class_info.class_hour
        request.session['class_type_code'] = class_info.subject_cd
        request.session['class_type_name'] = class_info.get_class_type_cd_name()
        request.session['class_center_name'] = class_info.get_center_name()
        request.session['trainer_name'] = class_info.member.name
        request.session['shared_program_flag'] = MY_PROGRAM
        if str(class_info.member_id) != str(request.user.id):
            request.session['class_type_name'] = class_info.get_class_type_cd_name() \
                                                 + '-' + class_info.member.name
            request.session['shared_program_flag'] = SHARED_PROGRAM

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    return redirect(next_page)


class GetBackgroundImgTypeListViewAjax(LoginRequiredMixin, AccessTestMixin, View):
    template_name = "ajax/trainer_common_code_ajax.html"

    def get(self, request):
        context = {}
        error = None

        context['common_cd_data'] = CommonCdTb.objects.filter(upper_common_cd='14', use=USE).order_by('order')

        if error is not None:
            messages.error(request, error)

        return render(request, self.template_name, context)


class GetBackgroundImgListViewAjax(LoginRequiredMixin, AccessTestMixin, View):
    template_name = "ajax/trainer_background_ajax.html"

    def post(self, request):
        context = {}
        class_id = request.POST.get('class_id', '')
        error = None
        background_img_data = None

        if error is None:
            if class_id is not None and class_id != '':
                background_img_data = BackgroundImgTb.objects.filter(class_tb_id=class_id,
                                                                     use=USE).order_by('-class_tb_id')
            else:
                background_img_data = BackgroundImgTb.objects.filter(class_tb__member_id=request.user.id,
                                                                     use=USE).order_by('-class_tb_id')

        if error is None:
            for background_img_info in background_img_data:
                try:
                    background_img_type_name = \
                        CommonCdTb.objects.get(common_cd=background_img_info.background_img_type_cd)
                except ObjectDoesNotExist:
                    background_img_type_name = None
                if background_img_type_name is not None:
                    background_img_info.background_img_type_name = background_img_type_name.common_cd_nm

        context['background_img_data'] = background_img_data

        if error is not None:
            messages.error(request, error)

        return render(request, self.template_name, context)


def update_background_img_info_logic(request):
    template_name = 'ajax/trainer_error_ajax.html'

    class_id = request.session.get('class_id', '')
    background_img_id = request.POST.get('background_img_id', '')
    background_img_type_cd = request.POST.get('background_img_type_cd', '')
    url = request.POST.get('url', '')

    error = None
    background_img_info = None
    if class_id is None or class_id == '':
        error = '오류가 발생했습니다.'
    if background_img_id == '' or background_img_id is None:
        try:
            background_img_info = BackgroundImgTb.objects.get(class_tb_id=class_id,
                                                              background_img_type_cd=background_img_type_cd,
                                                              use=USE)
        except ObjectDoesNotExist:
            background_img_info = BackgroundImgTb(class_tb_id=class_id,
                                                  background_img_type_cd=background_img_type_cd,
                                                  url=url, reg_info_id=request.user.id, use=USE)

        if background_img_type_cd is not None and background_img_type_cd != '':
            background_img_info.background_img_type_cd = background_img_type_cd
        if url is not None and url != '':
            background_img_info.url = url

        background_img_info.save()
    else:
        try:
            background_img_info = BackgroundImgTb.objects.get(background_img_id=background_img_id, use=USE)
        except ObjectDoesNotExist:
            error = '배경화면 정보를 불러오지 못했습니다.'

        if error is None:
            if background_img_type_cd is not None and background_img_type_cd != '':
                background_img_info.background_img_type_cd = background_img_type_cd
            if url is not None and url != '':
                background_img_info.url = url

            background_img_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, template_name)


def delete_background_img_info_logic(request):
    template_name = 'ajax/trainer_error_ajax.html'

    background_img_id = request.POST.get('background_img_id', '')

    error = None
    if background_img_id is None or background_img_id == '':
        error = '배경화면 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            background_img_info = BackgroundImgTb.objects.get(background_img_id=background_img_id)
            background_img_info.delete()
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다.'
    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, template_name)


class GetMyInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        # context = {}
        # context = super(GetTrainerInfoView, self).get_context_data(**kwargs)
        class_id = request.session.get('class_id', '')
        error = None
        # class_info = None

        # center_name = '없음'
        user_member_info = None
        class_info = None
        # off_repeat_schedule_data = None
        member_data = collections.OrderedDict()
        shared_program_flag = MY_PROGRAM

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if error is None:
            try:
                class_info = ClassTb.objects.get(class_id=class_id)

                if str(class_info.member.member_id) != str(request.user.id):
                    shared_program_flag = SHARED_PROGRAM

            except ObjectDoesNotExist:
                error = '지점 정보를 불러오지 못했습니다.'

        if error is None:
            try:
                user_member_info = MemberTb.objects.get(member_id=request.user.id)
            except ObjectDoesNotExist:
                error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            if user_member_info.profile_url is None or user_member_info.profile_url == '':
                member_profile_url = '/static/common/icon/icon_account.png'
            else:
                member_profile_url = user_member_info.profile_url
            coupon_count = CouponMemberTb.objects.filter(member_id=request.user.id, use=USE).count()
            member_data = {'member_id': request.user.id,
                           'member_user_id': user_member_info.user.username,
                           'member_name': user_member_info.name,
                           'member_phone': str(user_member_info.phone),
                           'member_email': str(user_member_info.user.email),
                           'member_sex': str(user_member_info.sex),
                           'member_birthday_dt': str(user_member_info.birthday_dt),
                           'member_profile_url': member_profile_url,
                           'member_phone_is_active': str(user_member_info.phone_is_active),
                           'member_coupon_count': coupon_count,
                           'program': {'program_id': class_id,
                                       'program_subject_type_name': class_info.get_class_type_cd_name(),
                                       'program_program_owner_id': class_info.member_id,
                                       'program_program_owner_name': class_info.member.name,
                                       'shared_program_flag': shared_program_flag
                                      }
                           }

        return JsonResponse({'trainer_info': member_data}, json_dumps_params={'ensure_ascii': True})


# 회원수정 api
def update_my_info_logic(request):
    # member_id = request.POST.get('id')
    # email = request.POST.get('email', '')
    first_name = request.POST.get('first_name', '')
    # phone = request.POST.get('phone', '')
    contents = request.POST.get('contents', '')
    country = request.POST.get('country', '')
    address = request.POST.get('address', '')
    sex = request.POST.get('sex', '')
    birthday_dt = request.POST.get('birthday', '')
    email = request.POST.get('email', '')

    error = None
    member_id = request.user.id
    user = None
    member = None
    if member_id == '':
        error = '회원 ID를 확인해 주세요.'

    if error is None:
        try:
            user = User.objects.get(id=member_id)
        except ObjectDoesNotExist:
            error = '회원 ID를 확인해 주세요.'

        try:
            member = MemberTb.objects.get(user_id=user.id)
        except ObjectDoesNotExist:
            error = '회원 ID를 확인해 주세요.'

    # input_first_name = ''
    # input_phone = ''
    # input_contents = ''
    # input_country = ''
    # input_address = ''
    # input_sex = ''
    # input_birthday_dt = ''

    if first_name is None or first_name == '':
        input_first_name = user.first_name
    else:
        input_first_name = first_name

    if contents is None or contents == '':
        input_contents = member.contents
    else:
        input_contents = contents

    if country is None or country == '':
        input_country = member.country
    else:
        input_country = country

    if address is None or address == '':
        input_address = member.address
    else:
        input_address = address

    if sex is None or sex == '':
        input_sex = member.sex
    else:
        input_sex = sex

    if birthday_dt is None or birthday_dt == '':
        input_birthday_dt = member.birthday_dt
    else:
        input_birthday_dt = birthday_dt

    # if phone is None or phone == '':
    #     input_phone = member.phone
    # else:
    #     if len(phone) != 11 and len(phone) != 10:
    #         error = '연락처를 확인해 주세요.'
    #     elif not phone.isdigit():
    #         error = '연락처를 확인해 주세요.'
    #     else:
    #         input_phone = phone
    if email is None or email == '':
        input_email = user.email
    else:
        input_email = email

    if error is None:
        try:
            with transaction.atomic():
                user.first_name = input_first_name
                # user.email = email
                user.email = input_email
                user.save()
                member.name = input_first_name
                # member.phone = input_phone
                member.contents = input_contents
                member.sex = input_sex
                if input_birthday_dt is not None and input_birthday_dt != '':
                    member.birthday_dt = input_birthday_dt
                member.country = input_country
                member.address = input_address
                member.save()

        except ValueError:
            error = '등록 값에 문제가 있습니다.'
        except IntegrityError:
            error = '등록 값에 문제가 있습니다.'
        except TypeError:
            error = '등록 값에 문제가 있습니다.'
        except ValidationError:
            error = '등록 값에 문제가 있습니다.'
        except InternalError:
            error = '등록 값에 문제가 있습니다.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 예약허용시간 setting 업데이트 api
def update_setting_push_logic(request):
    setting_to_trainee_lesson_alarm = request.POST.get('setting_to_trainee_lesson_alarm', '0')
    setting_to_shared_trainer_lesson_alarm = request.POST.get('setting_to_shared_trainer_lesson_alarm', '0')
    class_id = request.session.get('class_id', '')

    setting_type_cd_data = ['LT_PUS_TO_TRAINEE_LESSON_ALARM', 'LT_PUS_TO_SHARED_TRAINER_LESSON_ALARM']
    setting_info_data = [setting_to_trainee_lesson_alarm, setting_to_shared_trainer_lesson_alarm]

    error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_to_trainee_lesson_alarm'] = int(setting_to_trainee_lesson_alarm)
        request.session['setting_to_shared_trainer_lesson_alarm'] = int(setting_to_shared_trainer_lesson_alarm)
    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 예약허용시간 setting 업데이트 api
def update_setting_push_to_me_logic(request):
    setting_from_trainee_lesson_alarm = request.POST.get('setting_from_trainee_lesson_alarm', '1')
    setting_schedule_alarm_minute = request.POST.get('setting_schedule_alarm_minute', '-1')
    class_id = request.session.get('class_id', '')

    setting_type_cd_data = ['LT_PUS_FROM_TRAINEE_LESSON_ALARM', 'LT_PUSH_SCHEDULE_ALARM_MINUTE']
    setting_info_data = [setting_from_trainee_lesson_alarm, setting_schedule_alarm_minute]

    error = update_alarm_setting_data(class_id, request.user.id, setting_type_cd_data, setting_info_data)

    now = timezone.now()

    schedule_alarm_data = ScheduleAlarmTb.objects.select_related(
        'schedule_tb').filter(class_tb_id=class_id, alarm_dt__gte=now, member_id=request.user.id, use=USE)
    if setting_schedule_alarm_minute == '-1':
        schedule_alarm_data.delete()
    else:
        alarm_time = now + datetime.timedelta(minutes=int(setting_schedule_alarm_minute))
        alarm_time = alarm_time.strftime('%Y-%m-%d %H:%M:00')

        schedule_data = ScheduleTb.objects.filter(class_tb_id=class_id, start_dt__gte=alarm_time,
                                                  lecture_schedule_id__isnull=True,
                                                  permission_state_cd=PERMISSION_STATE_CD_APPROVE,
                                                  en_dis_type='1',
                                                  use=USE)
        for schedule_info in schedule_data:
            alarm_dt = schedule_info.start_dt - datetime.timedelta(minutes=int(setting_schedule_alarm_minute))
            # if schedule_info.push_alarm_data is not None and schedule_info.push_alarm_data != '':
            #     if '"'+str(request.user.id)+'"' in schedule_info.push_alarm_data:
            #         if ', "'+str(request.user.id)+'"' in schedule_info.push_alarm_data:
            #             schedule_info.push_alarm_data\
            #                 = schedule_info.push_alarm_data.replace(', "' + str(request.user.id) + '"', '')
            #         if '"'+str(request.user.id)+'"' in schedule_info.push_alarm_data:
            #             schedule_info.push_alarm_data\
            #                 = schedule_info.push_alarm_data.replace('"' + str(request.user.id) + '"', '')
            #
            #     push_alarm_data = json.loads(schedule_info.push_alarm_data)
            #
            #     if str(alarm_dt) in schedule_info.push_alarm_data:
            #
            #         try:
            #             push_alarm_data[str(alarm_dt)]['member_ids'].index(str(request.user.id))
            #         except ValueError:
            #             push_alarm_data[str(alarm_dt)]['member_ids'].append(str(request.user.id))
            #             schedule_info.push_alarm_data = str(push_alarm_data).replace("'", '"')
            #
            #     else:
            #         push_alarm_data[str(alarm_dt)] = {"member_ids": [str(request.user.id)]}
            #         schedule_info.push_alarm_data = str(push_alarm_data).replace("'", '"')
            # else:
            #     alarm_info = {
            #         str(alarm_dt): {
            #             "member_ids": [
            #                 str(request.user.id)
            #             ]
            #         }
            #     }
            #     schedule_info.push_alarm_data = str(alarm_info).replace("'", '"')
            # schedule_info.save()
            new_reg_test = True
            for schedule_alarm_info in schedule_alarm_data:
                if str(schedule_alarm_info.schedule_tb_id) == str(schedule_info.schedule_id):
                    schedule_alarm_info.alarm_dt = alarm_dt
                    schedule_alarm_info.alarm_minute = setting_schedule_alarm_minute
                    schedule_alarm_info.save()
                    new_reg_test = False
                    break

            if new_reg_test:
                schedule_alarm_info = ScheduleAlarmTb(class_tb_id=class_id, schedule_tb_id=schedule_info.schedule_id,
                                                      alarm_dt=alarm_dt, member_id=request.user.id,
                                                      alarm_minute=setting_schedule_alarm_minute,
                                                      use=USE)
                schedule_alarm_info.save()

    if error is None:
        request.session['setting_from_trainee_lesson_alarm'] = int(setting_from_trainee_lesson_alarm)
        request.session['setting_schedule_alarm_minute'] = setting_schedule_alarm_minute
    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 기본 setting 업데이트 api
def update_setting_calendar_setting_logic(request):
    setting_calendar_basic_select_time = request.POST.get('setting_calendar_basic_select_time', '60')
    # CALENDAR_TIME_SELECTOR_BASIC : 0 (신 PTERS 기본) / CALENDAR_TIME_SELECTOR_ORIGIN : 1 (구 PTERS 기본)
    setting_calendar_time_selector_type = request.POST.get('setting_calendar_time_selector_type',
                                                           CALENDAR_TIME_SELECTOR_BASIC)
    setting_week_start_date = request.POST.get('setting_week_start_date', 'SUN')
    setting_schedule_sign_enable = request.POST.get('setting_schedule_sign_enable', USE)
    class_id = request.session.get('class_id', '')

    if setting_calendar_basic_select_time is None or setting_calendar_basic_select_time == '':
        setting_calendar_basic_select_time = '60'
    if setting_calendar_time_selector_type is None or setting_calendar_time_selector_type == '':
        setting_calendar_time_selector_type = CALENDAR_TIME_SELECTOR_BASIC
    if setting_week_start_date is None or setting_week_start_date == '':
        setting_week_start_date = 'SUN'
    if setting_schedule_sign_enable is None or setting_schedule_sign_enable == '':
        setting_schedule_sign_enable = USE

    setting_type_cd_data = ['LT_CALENDAR_BASIC_SETTING_TIME', 'LT_CALENDAR_TIME_SELECTOR_TYPE',
                            'LT_WEEK_START_DATE', 'SCHEDULE_SIGN_ENABLE']
    setting_info_data = [setting_calendar_basic_select_time, setting_calendar_time_selector_type,
                         setting_week_start_date, setting_schedule_sign_enable]

    error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_calendar_basic_select_time'] = setting_calendar_basic_select_time
        request.session['setting_calendar_time_selector_type'] = setting_calendar_time_selector_type
        request.session['setting_week_start_date'] = setting_week_start_date
        request.session['setting_schedule_sign_enable'] = setting_schedule_sign_enable

    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 기본 setting 업데이트 api
def update_setting_work_time_logic(request):
    # setting_trainer_work_time_available = request.POST.get('setting_trainer_work_time_available', '00:00-23:59')
    setting_trainer_work_sun_time_avail = request.POST.get('setting_trainer_work_sun_time_avail', '00:00-24:00')
    setting_trainer_work_mon_time_avail = request.POST.get('setting_trainer_work_mon_time_avail', '00:00-24:00')
    setting_trainer_work_tue_time_avail = request.POST.get('setting_trainer_work_tue_time_avail', '00:00-24:00')
    setting_trainer_work_wed_time_avail = request.POST.get('setting_trainer_work_wed_time_avail', '00:00-24:00')
    setting_trainer_work_ths_time_avail = request.POST.get('setting_trainer_work_ths_time_avail', '00:00-24:00')
    setting_trainer_work_fri_time_avail = request.POST.get('setting_trainer_work_fri_time_avail', '00:00-24:00')
    setting_trainer_work_sat_time_avail = request.POST.get('setting_trainer_work_sat_time_avail', '00:00-24:00')
    setting_holiday_hide = request.POST.get('setting_holiday_hide', SHOW)
    class_id = request.session.get('class_id', '')

    if setting_trainer_work_sun_time_avail is None or setting_trainer_work_sun_time_avail == '':
        setting_trainer_work_sun_time_avail = '00:00-24:00'
    if setting_trainer_work_mon_time_avail is None or setting_trainer_work_mon_time_avail == '':
        setting_trainer_work_mon_time_avail = '00:00-24:00'
    if setting_trainer_work_tue_time_avail is None or setting_trainer_work_tue_time_avail == '':
        setting_trainer_work_tue_time_avail = '00:00-24:00'
    if setting_trainer_work_wed_time_avail is None or setting_trainer_work_wed_time_avail == '':
        setting_trainer_work_wed_time_avail = '00:00-24:00'
    if setting_trainer_work_ths_time_avail is None or setting_trainer_work_ths_time_avail == '':
        setting_trainer_work_ths_time_avail = '00:00-24:00'
    if setting_trainer_work_fri_time_avail is None or setting_trainer_work_fri_time_avail == '':
        setting_trainer_work_fri_time_avail = '00:00-24:00'
    if setting_trainer_work_sat_time_avail is None or setting_trainer_work_sat_time_avail == '':
        setting_trainer_work_sat_time_avail = '00:00-24:00'
    if setting_holiday_hide is None or setting_holiday_hide == '':
        setting_holiday_hide = SHOW
    setting_type_cd_data = ['LT_WORK_SUN_TIME_AVAIL', 'LT_WORK_MON_TIME_AVAIL',
                            'LT_WORK_TUE_TIME_AVAIL', 'LT_WORK_WED_TIME_AVAIL',
                            'LT_WORK_THS_TIME_AVAIL', 'LT_WORK_FRI_TIME_AVAIL',
                            'LT_WORK_SAT_TIME_AVAIL', 'LT_HOLIDAY_HIDE']
    setting_info_data = [setting_trainer_work_sun_time_avail, setting_trainer_work_mon_time_avail,
                         setting_trainer_work_tue_time_avail, setting_trainer_work_wed_time_avail,
                         setting_trainer_work_ths_time_avail, setting_trainer_work_fri_time_avail,
                         setting_trainer_work_sat_time_avail, setting_holiday_hide]

    error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_trainer_work_sun_time_avail'] = setting_trainer_work_sun_time_avail
        request.session['setting_trainer_work_mon_time_avail'] = setting_trainer_work_mon_time_avail
        request.session['setting_trainer_work_tue_time_avail'] = setting_trainer_work_tue_time_avail
        request.session['setting_trainer_work_wed_time_avail'] = setting_trainer_work_wed_time_avail
        request.session['setting_trainer_work_ths_time_avail'] = setting_trainer_work_ths_time_avail
        request.session['setting_trainer_work_fri_time_avail'] = setting_trainer_work_fri_time_avail
        request.session['setting_trainer_work_sat_time_avail'] = setting_trainer_work_sat_time_avail
        request.session['setting_holiday_hide'] = setting_holiday_hide

    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 기본 setting 업데이트 api
def update_setting_auto_complete_logic(request):
    setting_schedule_auto_finish = request.POST.get('setting_schedule_auto_finish', AUTO_FINISH_OFF)
    setting_member_ticket_auto_finish = request.POST.get('setting_member_ticket_auto_finish', AUTO_FINISH_OFF)
    class_id = request.session.get('class_id', '')

    if setting_schedule_auto_finish is None or setting_schedule_auto_finish == '':
        setting_schedule_auto_finish = AUTO_FINISH_OFF
    if setting_member_ticket_auto_finish is None or setting_member_ticket_auto_finish == '':
        setting_member_ticket_auto_finish = AUTO_FINISH_OFF

    setting_type_cd_data = ['LT_SCHEDULE_AUTO_FINISH', 'LT_LECTURE_AUTO_FINISH']
    setting_info_data = [setting_schedule_auto_finish, setting_member_ticket_auto_finish]
    error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_schedule_auto_finish'] = setting_schedule_auto_finish
        request.session['setting_member_ticket_auto_finish'] = setting_member_ticket_auto_finish

    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 예약허용시간 setting 업데이트 api
def update_setting_reserve_logic(request):
    # 0(UN_USE) 회원이 정원 보기 불가 /  1(USE) 회원이 정원 보기 허용
    setting_member_lecture_max_num_view_available = request.POST.get('setting_member_lecture_max_num_view_available', USE)
    setting_member_lecture_main_trainer_view_available = request.POST.get('setting_member_lecture_main_trainer_view_available', USE)
    setting_member_disable_schedule_visible = request.POST.get('setting_member_disable_schedule_visible', UN_USE)
    setting_member_reserve_time_available = request.POST.get('setting_member_reserve_time_available', '00:00-24:00')
    setting_member_reserve_enable_time = request.POST.get('setting_member_reserve_enable_time', '60')
    setting_member_reserve_cancel_time = request.POST.get('setting_member_reserve_cancel_time', '60')
    setting_member_reserve_prohibition = request.POST.get('setting_member_reserve_prohibition',
                                                          MEMBER_RESERVE_PROHIBITION_ON)
    setting_member_reserve_date_available = request.POST.get('setting_member_reserve_date_available', '7')
    # setting_member_time_duration = request.POST.get('setting_member_time_duration', '60')
    # setting_member_start_time = request.POST.get('setting_member_start_time', 'A-0')
    setting_member_private_class_auto_permission = request.POST.get('setting_member_private_class_auto_permission', USE)
    setting_member_public_class_auto_permission = request.POST.get('setting_member_public_class_auto_permission', USE)
    setting_member_public_class_wait_member_num = request.POST.get('setting_member_public_class_wait_member_num', 0)
    setting_member_wait_schedule_auto_cancel_time = request.POST.get('setting_member_wait_schedule_auto_cancel_time', 0)
    setting_single_lecture_duplicate = request.POST.get('setting_single_lecture_duplicate', UN_USE)
    class_id = request.session.get('class_id', '')

    if setting_member_lecture_max_num_view_available is None or setting_member_lecture_max_num_view_available == '':
        setting_member_lecture_max_num_view_available = USE
    if setting_member_lecture_main_trainer_view_available is None or setting_member_lecture_main_trainer_view_available == '':
        setting_member_lecture_main_trainer_view_available = USE
    if setting_member_disable_schedule_visible is None or setting_member_disable_schedule_visible == '':
        setting_member_disable_schedule_visible = UN_USE
    if setting_member_reserve_time_available is None or setting_member_reserve_time_available == '':
        setting_member_reserve_time_available = '00:00-24:00'
    if setting_member_reserve_enable_time is None or setting_member_reserve_enable_time == '':
        setting_member_reserve_enable_time = '60'
    if setting_member_reserve_cancel_time is None or setting_member_reserve_cancel_time == '':
        setting_member_reserve_cancel_time = '60'
    if setting_member_reserve_prohibition is None or setting_member_reserve_prohibition == '':
        setting_member_reserve_prohibition = MEMBER_RESERVE_PROHIBITION_ON
    if setting_member_reserve_date_available is None or setting_member_reserve_date_available == '':
        setting_member_reserve_date_available = '7'
    # if setting_member_time_duration is None or setting_member_time_duration == '':
    #     setting_member_time_duration = '60'
    # if setting_member_start_time is None or setting_member_start_time == '':
    #     setting_member_start_time = 'A-0'
    if setting_member_private_class_auto_permission is None or setting_member_private_class_auto_permission == '':
        setting_member_private_class_auto_permission = USE
    if setting_member_public_class_auto_permission is None or setting_member_public_class_auto_permission == '':
        setting_member_public_class_auto_permission = USE
    if setting_member_public_class_wait_member_num is None or setting_member_public_class_wait_member_num == '':
        setting_member_public_class_wait_member_num = 0
    if setting_member_wait_schedule_auto_cancel_time is None or setting_member_wait_schedule_auto_cancel_time == '':
        setting_member_wait_schedule_auto_cancel_time = 0
    if setting_single_lecture_duplicate is None or setting_single_lecture_duplicate == '':
        setting_single_lecture_duplicate = UN_USE

    setting_type_cd_data = ['LT_RES_MEMBER_LECTURE_MAX_NUM_VIEW', 'LT_RES_MEMBER_LECTURE_MAIN_TRAINER_VIEW',
                            'LT_RES_MEMBER_DISABLE_SCHEDULE_VISIBLE',
                            'LT_RES_01', 'LT_RES_03', 'LT_RES_05',
                            'LT_RES_CANCEL_TIME', 'LT_RES_ENABLE_TIME',
                            'LT_RES_PRIVATE_CLASS_AUTO_PERMISSION', 'LT_RES_PUBLIC_CLASS_AUTO_PERMISSION',
                            'LT_RES_PUBLIC_CLASS_WAIT_MEMBER_NUM', 'LT_RES_WAIT_SCHEDULE_AUTO_CANCEL_TIME',
                            'LT_RES_SINGLE_LECTURE_DUPLICATE']
    setting_info_data = [setting_member_lecture_max_num_view_available,
                         setting_member_lecture_main_trainer_view_available, setting_member_disable_schedule_visible,
                         setting_member_reserve_time_available, setting_member_reserve_prohibition,
                         setting_member_reserve_date_available, setting_member_reserve_cancel_time,
                         setting_member_reserve_enable_time,
                         setting_member_private_class_auto_permission, setting_member_public_class_auto_permission,
                         setting_member_public_class_wait_member_num, setting_member_wait_schedule_auto_cancel_time,
                         setting_single_lecture_duplicate]
    error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_member_lecture_max_num_view_available'] = setting_member_lecture_max_num_view_available
        request.session['setting_member_lecture_main_trainer_view_available']\
            = setting_member_lecture_main_trainer_view_available
        request.session['setting_member_disable_schedule_visible'] = setting_member_disable_schedule_visible
        request.session['setting_member_reserve_time_available'] = setting_member_reserve_time_available
        request.session['setting_member_reserve_prohibition'] = setting_member_reserve_prohibition
        request.session['setting_member_reserve_enable_time'] = setting_member_reserve_enable_time
        request.session['setting_member_reserve_cancel_time'] = setting_member_reserve_cancel_time
        request.session['setting_member_reserve_date_available'] = setting_member_reserve_date_available
        # request.session['setting_member_time_duration'] = setting_member_time_duration
        # request.session['setting_member_start_time'] = setting_member_start_time
        request.session['setting_member_private_class_auto_permission'] = setting_member_private_class_auto_permission
        request.session['setting_member_public_class_auto_permission'] = setting_member_public_class_auto_permission
        request.session['setting_member_public_class_wait_member_num'] = setting_member_public_class_wait_member_num
        request.session['setting_member_wait_schedule_auto_cancel_time'] = setting_member_wait_schedule_auto_cancel_time
        request.session['setting_single_lecture_duplicate'] = setting_single_lecture_duplicate
    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 매출 통계 관련 설정?
def update_setting_sales_logic(request):
    setting_sal_01 = request.POST.get('setting_sales_10', '')
    setting_sal_02 = request.POST.get('setting_sales_20', '')
    setting_sal_03 = request.POST.get('setting_sales_30', '')
    setting_sal_04 = request.POST.get('setting_sales_40', '')
    setting_sal_05 = request.POST.get('setting_sales_50', '')
    setting_sal_00 = request.POST.get('setting_sales', '')
    setting_sal_type = request.POST.get('setting_sales_type', '0')
    class_id = request.session.get('class_id', '')

    if setting_sal_type != '0':
        setting_sal_01 = ''
        setting_sal_02 = ''
        setting_sal_03 = ''
        setting_sal_04 = ''
        setting_sal_05 = ''

    setting_type_cd_data = ['LT_SAL_01', 'LT_SAL_02', 'LT_SAL_03', 'LT_SAL_04', 'LT_SAL_05', 'LT_SAL_00']
    setting_info_data = [setting_sal_01, setting_sal_02, setting_sal_03, setting_sal_04, setting_sal_05, setting_sal_00]
    error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 언어 setting 업데이트 api
def update_setting_language_logic(request):
    setting_member_language = request.POST.get('setting_member_language', '')
    class_id = request.session.get('class_id', '')

    error = None
    lt_lan_01 = None
    if error is None:
        if setting_member_language == '':
            setting_member_language = 'KOR'

    if error is None:
        try:
            lt_lan_01 = SettingTb.objects.get(member_id=request.user.id,  setting_type_cd='LT_LAN_01')
        except ObjectDoesNotExist:
            lt_lan_01 = SettingTb(member_id=request.user.id, setting_type_cd='LT_LAN_01', use=USE)

    if error is None:
        try:
            with transaction.atomic():
                lt_lan_01.setting_info = setting_member_language
                lt_lan_01.save()

        except ValueError:
            error = '등록 값에 문제가 있습니다.'
        except IntegrityError:
            error = '등록 값에 문제가 있습니다.'
        except TypeError:
            error = '등록 값에 문제가 있습니다.'
        except ValidationError:
            error = '등록 값에 문제가 있습니다.'
        except InternalError:
            error = '등록 값에 문제가 있습니다.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 ㅌㅔ마 setting 업데이트 api
def update_setting_theme_logic(request):
    setting_theme = request.POST.get('theme', 'light')

    if setting_theme is None or setting_theme == '':
        setting_theme = 'light'

    setting_type_cd_data = ['THEME']
    setting_info_data = [setting_theme]
    error = update_user_setting_data(request.user.id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_theme'] = setting_theme

    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 통계 잠금 설정 업데이트 api
def update_setting_access_lock_logic(request):
    setting_admin_password = request.POST.get('setting_admin_password', '0000')
    setting_trainer_attend_mode_out_lock = request.POST.get('setting_trainer_attend_mode_out_lock', str(UN_USE))
    setting_trainer_statistics_lock = request.POST.get('setting_trainer_statistics_lock', str(UN_USE))
    class_id = request.session.get('class_id', '')
    error = None
    if setting_admin_password is None or setting_admin_password == '':
        setting_admin_password = '0000'
    if len(setting_admin_password) != 4:
        error = '잠금 해제 비밀번호는 4자리만 가능합니다.'

    if setting_trainer_attend_mode_out_lock is None or setting_trainer_attend_mode_out_lock == '':
        setting_trainer_attend_mode_out_lock = str(UN_USE)
    if setting_trainer_statistics_lock is None or setting_trainer_statistics_lock == '':
        setting_trainer_statistics_lock = str(UN_USE)

    if error is None:
        setting_type_cd_data = ['LT_ADMIN_PASSWORD', 'ATTEND_MODE_OUT_LOCK', 'STATISTICS_LOCK']
        setting_info_data = [setting_admin_password, setting_trainer_attend_mode_out_lock, setting_trainer_statistics_lock]
        error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_admin_password'] = setting_trainer_statistics_lock
        request.session['setting_trainer_attend_mode_out_lock'] = setting_trainer_attend_mode_out_lock
        request.session['setting_trainer_statistics_lock'] = setting_trainer_statistics_lock

    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


class GetTrainerSettingDataView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        func_setting_data_update(request, 'trainer')
        return JsonResponse(request.session['setting_data'], json_dumps_params={'ensure_ascii': True})


class GetProgramAuthDataView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        get_function_auth_type_cd(request)
        return JsonResponse(request.session['auth_info'], json_dumps_params={'ensure_ascii': True})


class GetTrainerAuthDataView(LoginRequiredMixin, AccessTestMixin, View):
    def get(self, request):
        context = {}
        today = datetime.date.today()
        payment_data = PaymentInfoTb.objects.filter(member_id=request.user.id,
                                                    status='paid',
                                                    start_date__lte=today, end_date__gte=today,
                                                    use=USE).order_by('-payment_info_id')
        context["auth_info"] = {}
        if len(payment_data) > 0:
            payment_info = payment_data[0]
            payment_info.start_date = str(payment_info.start_date)
            payment_info.end_date = str(payment_info.end_date)

            function_list = ProductFunctionAuthTb.objects.select_related(
                'function_auth_tb', 'product_tb').filter(product_tb_id=payment_info.product_tb_id,
                                                         use=USE).order_by('product_tb_id',
                                                                           'function_auth_tb_id',
                                                                           'auth_type_cd')
            for function_info in function_list:
                auth_info = {}
                if function_info.auth_type_cd is None:
                    function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd)
                else:
                    function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd) \
                                                 + str(function_info.auth_type_cd)

                auth_info["active"] = 1
                auth_info["limit_num"] = function_info.counts
                auth_info["limit_type"] = function_info.product_tb.name
                context["auth_info"][function_auth_type_cd_name] = auth_info

        else:
            function_list = ProductFunctionAuthTb.objects.select_related(
                'function_auth_tb', 'product_tb').filter(product_tb_id=6, use=USE).order_by('product_tb_id',
                                                                                            'function_auth_tb_id',
                                                                                            'auth_type_cd')

            for function_info in function_list:
                auth_info = {}
                if function_info.auth_type_cd is None:
                    function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd)
                else:
                    function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd) \
                                                 + str(function_info.auth_type_cd)
                auth_info["active"] = 1
                auth_info["limit_num"] = function_info.counts
                auth_info["limit_type"] = str(function_info.product_tb.name)
                context["auth_info"][function_auth_type_cd_name] = auth_info
        return JsonResponse(context["auth_info"], json_dumps_params={'ensure_ascii': True})


# log 삭제
def alarm_delete_logic(request):
    log_size = request.POST.get('log_id_size')
    delete_log_id = request.POST.getlist('log_id_array[]', '')

    error = None
    if log_size == '0':
        error = '알람이 없습니다.'

    if error is None:

        for i in range(0, int(log_size)):
            try:
                log_info = LogTb.objects.get(log_id=delete_log_id[i])
            except ObjectDoesNotExist:
                error = '알람 정보를 불러오지 못했습니다.'
            if error is None:
                log_info.use = 0
                log_info.save()

    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    return render(request, 'ajax/trainer_error_ajax.html')


class GetNoticeInfoView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'ajax/notice_info_ajax.html'

    def get_context_data(self, **kwargs):
        context = super(GetNoticeInfoView, self).get_context_data(**kwargs)
        # class_id = self.request.session.get('class_id', '')

        board_list = BoardTb.objects.filter(board_type_cd='NOTICE',
                                            to_member_type_cd='ALL', use=USE).order_by('mod_dt')
        board_list |= BoardTb.objects.filter(board_type_cd='NOTICE',
                                             to_member_type_cd='TRAINEE', use=USE).order_by('mod_dt')
        board_list.order_by('mod_dt')
        for board_info in board_list:
            board_info.hits += 1
            board_info.save()

        context['board_list'] = board_list

        return context


class GetProgramBoardIngListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        program_board_dict = collections.OrderedDict()
        class_id = request.session.get('class_id')
        program_board_list = ProgramBoardTb.objects.filter(class_tb=class_id, auth_type_cd=AUTH_TYPE_VIEW, use=USE
                                                           ).order_by('-reg_dt')

        for program_board_info in program_board_list:
            program_board_dict[program_board_info.program_board_id] = {
                'program_board_id': program_board_info.program_board_id,
                'program_board_type_cd': program_board_info.board_type_cd,
                'program_board_member_id': program_board_info.member_id,
                'program_board_class_id': program_board_info.class_tb_id,
                'program_board_to_member_type_cd': program_board_info.to_member_type_cd,
                'program_board_comment_available': program_board_info.comment_available,
                'program_board_anonymous_available': program_board_info.anonymous_available,
                'program_board_important_flag': program_board_info.important_flag,
                'program_board_auth_type_cd': program_board_info.auth_type_cd,
                'program_board_push_status': program_board_info.push_status,
                'program_board_order': program_board_info.order,
                'program_board_name': program_board_info.name,
                'program_board_note': program_board_info.note,
                'program_board_use': program_board_info.use}
        return JsonResponse(program_board_dict, json_dumps_params={'ensure_ascii': True})


class GetProgramBoardEndListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        program_board_dict = collections.OrderedDict()
        class_id = request.session.get('class_id')
        program_board_list = ProgramBoardTb.objects.filter(class_tb=class_id, auth_type_cd=AUTH_TYPE_VIEW, use=UN_USE
                                                           ).order_by('-reg_dt')

        for program_board_info in program_board_list:
            program_board_dict[program_board_info.program_board_id] = {
                'program_board_id': program_board_info.program_board_id,
                'program_board_type_cd': program_board_info.board_type_cd,
                'program_board_member_id': program_board_info.member_id,
                'program_board_class_id': program_board_info.class_tb_id,
                'program_board_to_member_type_cd': program_board_info.to_member_type_cd,
                'program_board_comment_available': program_board_info.comment_available,
                'program_board_anonymous_available': program_board_info.anonymous_available,
                'program_board_important_flag': program_board_info.important_flag,
                'program_board_auth_type_cd': program_board_info.auth_type_cd,
                'program_board_push_status': program_board_info.push_status,
                'program_board_order': program_board_info.order,
                'program_board_name': program_board_info.name,
                'program_board_note': program_board_info.note,
                'program_board_use': program_board_info.use}
        return JsonResponse(program_board_dict, json_dumps_params={'ensure_ascii': True})


class GetProgramBoardInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        program_board_dict = collections.OrderedDict()
        program_board_id = request.GET.get('program_board_id')
        class_id = request.session.get('class_id')
        error = None
        if program_board_id is None or program_board_id == '':
            error = '게시판 정보를 불러오지 못했습니다.'
        if error is None:
            try:
                program_board_info = ProgramBoardTb.objects.get(program_board_id=program_board_id,
                                                                class_tb=class_id, auth_type_cd=AUTH_TYPE_VIEW)
                program_board_dict = {
                    'program_board_id': program_board_info.program_board_id,
                    'program_board_type_cd': program_board_info.board_type_cd,
                    'program_board_member_id': program_board_info.member_id,
                    'program_board_class_id': program_board_info.class_tb_id,
                    'program_board_to_member_type_cd': program_board_info.to_member_type_cd,
                    'program_board_comment_available': program_board_info.comment_available,
                    'program_board_anonymous_available': program_board_info.anonymous_available,
                    'program_board_important_flag': program_board_info.important_flag,
                    'program_board_auth_type_cd': program_board_info.auth_type_cd,
                    'program_board_push_status': program_board_info.push_status,
                    'program_board_order': program_board_info.order,
                    'program_board_name': program_board_info.name,
                    'program_board_note': program_board_info.note,
                    'program_board_use': program_board_info.use}
            except ObjectDoesNotExist:
                error = '게시판 정보를 불러오지 못했습니다.'
        if error is not None:
            context = {'messageArray': error}
            return JsonResponse(context, json_dumps_params={'ensure_ascii': True})

        return JsonResponse(program_board_dict, json_dumps_params={'ensure_ascii': True})


class GetProgramBoardAllView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        notice_data_dict = collections.OrderedDict()
        notice_data = NoticeTb.objects.filter().order_by('-reg_dt')

        for notice_info in notice_data:
            notice_data_dict[notice_info.notice_id] = {'notice_id': notice_info.notice_id,
                                                       'notice_type_cd': notice_info.notice_type_cd,
                                                       'notice_title': notice_info.title,
                                                       'notice_contents': notice_info.contents,
                                                       'notice_to_member_type_cd': notice_info.to_member_type_cd,
                                                       'notice_hits': notice_info.hits,
                                                       'notice_mod_dt': notice_info.mod_dt,
                                                       'notice_reg_dt': notice_info.reg_dt,
                                                       'notice_use': notice_info.use}

        return JsonResponse(notice_data_dict, json_dumps_params={'ensure_ascii': True})


def add_program_board_info_logic(request):

    # NOTICE : 공지사항, BOARD : 게시판
    board_type_cd = request.POST.get('board_type_cd', BOARD_TYPE_CD_NOTICE)
    to_member_type_cd = request.POST.get('to_member_type_cd', 'trainee')
    comment_available = request.POST.get('comment_available', UN_USE)
    anonymous_available = request.POST.get('anonymous_available', UN_USE)
    important_flag = request.POST.get('important_flag', UN_USE)
    auth_type_cd = request.POST.get('auth_type_cd', AUTH_TYPE_VIEW)
    push_status = request.POST.get('push_status', UN_USE)
    name = request.POST.get('name', '')
    note = request.POST.get('note', '')
    use = request.POST.get('use', USE)
    class_id = request.session.get('class_id')

    context = {}
    error = None
    if board_type_cd == '' or board_type_cd is None:
        error = '게시판 종류를 선택해주세요.'

    if error is None:
        program_board_info = ProgramBoardTb(board_type_cd=board_type_cd, class_tb_id=class_id,
                                            member_id=request.user.id, to_member_type_cd=to_member_type_cd,
                                            comment_available=comment_available, auth_type_cd=auth_type_cd,
                                            anonymous_available=anonymous_available, important_flag=important_flag,
                                            push_status=push_status, name=name, note=note, use=use)
        program_board_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        # messages.error(request, error)
        context['messageArray'] = error

    return render(request, 'ajax/trainer_error_ajax.html')


def update_program_board_info_logic(request):

    # NOTICE : 공지사항, BOARD : 게시판
    program_board_id = request.POST.get('program_board_id')
    board_type_cd = request.POST.get('board_type_cd', '')
    to_member_type_cd = request.POST.get('to_member_type_cd', 'trainee')
    comment_available = request.POST.get('comment_available', '')
    anonymous_available = request.POST.get('anonymous_available', '')
    important_flag = request.POST.get('important_flag', '')
    auth_type_cd = request.POST.get('auth_type_cd', '')
    push_status = request.POST.get('push_status', '')
    name = request.POST.get('name', '')
    note = request.POST.get('note', '')
    use = request.POST.get('use', '')
    class_id = request.session.get('class_id')

    context = {}
    error = None
    program_board_info = None

    if program_board_id is None or program_board_id == '':
        error = '게시판 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            program_board_info = ProgramBoardTb.objects.get(program_board_id=program_board_id, class_tb=class_id)
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다. [0]'

    if error is None:
        if board_type_cd != '' and board_type_cd is not None:
            program_board_info.board_type_cd = board_type_cd
        if to_member_type_cd != '' and to_member_type_cd is not None:
            program_board_info.to_member_type_cd = to_member_type_cd
        if comment_available != '' and comment_available is not None:
            program_board_info.comment_available = comment_available
        if anonymous_available != '' and anonymous_available is not None:
            program_board_info.anonymous_available = anonymous_available
        if important_flag != '' and important_flag is not None:
            program_board_info.important_flag = important_flag
        if auth_type_cd != '' and auth_type_cd is not None:
            program_board_info.auth_type_cd = auth_type_cd
        if push_status != '' and push_status is not None:
            program_board_info.push_status = push_status
        if name != '' and name is not None:
            program_board_info.name = name
        if note != '' and note is not None:
            program_board_info.note = note
        if use != '' and use is not None:
            program_board_info.use = use
        program_board_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        # messages.error(request, error)
        context['messageArray'] = error

    return render(request, 'ajax/trainer_error_ajax.html')


def delete_program_board_info_logic(request):

    notice_id = request.POST.get('notice_id')
    member_type_cd = request.session.get('group_name')

    context = {}
    error = None
    notice_info = None

    if member_type_cd != 'admin':
        error = '관리자만 접근 가능합니다.'

    if notice_id is None or notice_id == '':
        error = '변경할 게시글을 선택해주세요.'

    if error is None:
        try:
            notice_info = NoticeTb.objects.get(notice_id=notice_id)
        except ObjectDoesNotExist:
            error = '게시글을 불러오지 못했습니다.'

    if error is None:
        notice_info.delete()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        # messages.error(request, error)
        context['messageArray'] = error

    return render(request, 'ajax/trainer_error_ajax.html')


class GetProgramNoticeAllView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = request.session.get('class_id')
        program_notice_data_dict = collections.OrderedDict()

        program_notice_data = ProgramNoticeTb.objects.select_related(
            'member').filter(class_tb_id=class_id).order_by('-reg_dt')

        for program_notice_info in program_notice_data:
            program_notice_data_dict[program_notice_info.program_notice_id] = {
                'program_notice_id': program_notice_info.program_notice_id,
                'program_notice_type_cd': program_notice_info.notice_type_cd,
                'program_notice_title': program_notice_info.title,
                'program_notice_contents': program_notice_info.contents,
                'program_notice_to_member_type_cd': program_notice_info.to_member_type_cd,
                'program_notice_hits': program_notice_info.hits,
                'program_notice_mod_dt': str(program_notice_info.mod_dt),
                'program_notice_reg_dt': str(program_notice_info.reg_dt),
                'program_notice_reg_member_name': str(program_notice_info.member.name),
                'program_notice_use': program_notice_info.use
            }

        return JsonResponse(program_notice_data_dict, json_dumps_params={'ensure_ascii': True})


class AddProgramNoticeInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def post(self, request):
        # NOTICE : 공지사항, SYS_USAGE : 사용법 , FAQ : 자주 묻는 질문
        notice_type_cd = request.POST.get('notice_type_cd', '')
        title = request.POST.get('title', '')
        contents = request.POST.get('contents', '')
        to_member_type_cd = request.POST.get('to_member_type_cd')
        push_use = request.POST.get('push_use', str(UN_USE))
        use = request.POST.get('use', USE)
        class_id = request.session.get('class_id')
        class_name = request.session.get('class_type_name', '')
        context = {}
        error = None

        if notice_type_cd == '' or notice_type_cd is None:
            error = '공지 유형을 선택해주세요.'

        if error is None:
            program_notice_info = ProgramNoticeTb(member_id=request.user.id, notice_type_cd=notice_type_cd,
                                                  class_tb_id=class_id, title=title, contents=contents,
                                                  to_member_type_cd=to_member_type_cd, use=use)
            program_notice_info.save()
        if error is None:
            if str(use) == str(USE) and str(push_use) == str(USE):
                error = func_send_push_notice(to_member_type_cd, class_id,
                                              '[공지사항] - ' + class_name, title)
        if error is not None:
            logger.error('[' + str(request.user.id) + ']' + error)
            # messages.error(request, error)
            context['messageArray'] = error

        return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


class UpdateProgramNoticeInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def post(self, request):
        program_notice_id = request.POST.get('program_notice_id')
        notice_type_cd = request.POST.get('notice_type_cd', '')
        title = request.POST.get('title', '')
        contents = request.POST.get('contents', '')
        to_member_type_cd = request.POST.get('to_member_type_cd')
        push_use = request.POST.get('push_use', str(UN_USE))
        use = request.POST.get('use', USE)
        class_id = request.session.get('class_id')
        class_name = request.session.get('class_type_name', '')
        context = {}
        error = None
        program_notice_info = None

        if program_notice_id is None or program_notice_id == '':
            error = '변경할 공지사항을 선택해주세요.'

        if notice_type_cd == '' or notice_type_cd is None:
            error = '공지 유형을 선택해주세요.'

        if error is None:
            try:
                program_notice_info = ProgramNoticeTb.objects.get(program_notice_id=program_notice_id)
            except ObjectDoesNotExist:
                error = '공지사항을 불러오지 못했습니다.'

        if error is None:
            program_notice_info.member_id = request.user.id
            program_notice_info.notice_type_cd = notice_type_cd
            program_notice_info.title = title
            program_notice_info.contents = contents
            program_notice_info.to_member_type_cd = to_member_type_cd
            program_notice_info.mod_dt = timezone.now()
            program_notice_info.use = use
            program_notice_info.save()

        if error is None:
            if str(use) == str(USE) and str(push_use) == str(USE):
                func_send_push_notice(to_member_type_cd, class_id, '[공지사항] - ' + class_name, title)

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            # messages.error(request, error)
            context['messageArray'] = error

        return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


class DeleteProgramNoticeInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def post(self, request):
        program_notice_id = request.POST.get('program_notice_id')

        context = {}
        error = None
        program_notice_info = None

        if program_notice_id is None or program_notice_id == '':
            error = '변경할 공지사항을 선택해주세요.'

        if error is None:
            try:
                program_notice_info = ProgramNoticeTb.objects.get(program_notice_id=program_notice_id)
            except ObjectDoesNotExist:
                error = '공지사항을 불러오지 못했습니다.'

        if error is None:
            program_notice_info.delete()

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            # messages.error(request, error)
            context['messageArray'] = error

        return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


# 상품 추가
def add_shop_info_logic(request):
    class_id = request.session.get('class_id', '')
    name = request.POST.get('name', '')
    note = request.POST.get('note', '')
    price = request.POST.get('price', 0)
    error = None

    try:
        price = int(price)
    except ValueError:
        error = '상품 금액은 숫자만 입력 가능합니다.'

    if name == '' or name is None:
        error = '상품명을 입력해주세요. '

    if error is None:
        try:
            with transaction.atomic():

                shop_info = ShopTb(class_tb_id=class_id, name=name, price=price, note=note, use=USE)
                shop_info.save()

        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id, from_member_name=request.user.first_name,
                         class_tb_id=class_id, log_info=name+' 상품', log_how='추가', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 상품 삭제
def delete_shop_info_logic(request):
    class_id = request.session.get('class_id', '')
    shop_id = request.POST.get('shop_id', '')
    error = None
    shop_info = None
    now = timezone.now()

    try:
        shop_info = ShopTb.objects.get(class_tb_id=class_id, shop_id=shop_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        try:
            with transaction.atomic():
                shop_info.state_cd = STATE_CD_FINISH
                shop_info.use = UN_USE
                shop_info.save()

        except ValueError:
            error = '오류가 발생했습니다. [1]'
        except IntegrityError:
            error = '오류가 발생했습니다. [2]'
        except TypeError:
            error = '오류가 발생했습니다. [3]'
        except ValidationError:
            error = '오류가 발생했습니다. [4]'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id, from_member_name=request.user.first_name,
                         class_tb_id=class_id, log_info=shop_info.name+' 상품', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 상품 수정
def update_shop_info_logic(request):
    class_id = request.session.get('class_id', '')
    shop_id = request.POST.get('shop_id', '')
    name = request.POST.get('name', '')
    note = request.POST.get('note', '')
    price = request.POST.get('price', '')
    error = None
    shop_info = None

    try:
        shop_info = ShopTb.objects.get(class_tb_id=class_id, shop_id=shop_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        if name != '' and name is not None:
            shop_info.name = name
        if note != '' and note is not None:
            shop_info.note = note
        if price != '' and price is not None:
            shop_info.price = price
        shop_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id, from_member_name=request.user.first_name,
                         class_tb_id=class_id, log_info=shop_info.name+' 상품', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetShopInfoViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id')
        shop_id = request.GET.get('shop_id')
        error = None
        shop_dict = {}
        try:
            shop_info = ShopTb.objects.get(shop_id=shop_id, use=USE)
        except ObjectDoesNotExist:
            error = '상품 정보를 가져오지 못했습니다.'

        if error is None:
            shop_dict = {'shop_id': str(shop_info.shop_id),
                         'shop_name': shop_info.name,
                         'shop_price': shop_info.price,
                         'shop_note': shop_info.note,
                         'shop_mod_dt': str(shop_info.mod_dt),
                         'shop_reg_dt': str(shop_info.reg_dt)
                         }

        return JsonResponse(shop_dict, json_dumps_params={'ensure_ascii': True})


class GetShopIngListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_sort = request.GET.get('sort_val', SORT_MEMBER_NAME)
        sort_order_by = request.GET.get('sort_order_by', SORT_ASC)
        keyword = request.GET.get('keyword', '')

        shop_data = ShopTb.objects.filter(class_tb_id=class_id, state_cd=STATE_CD_IN_PROGRESS, use=USE).order_by('name')
        shop_list = []
        for shop_info in shop_data:
            shop_dict = {'shop_id':shop_info.shop_id,
                         'shop_name': shop_info.name,
                         'shop_price': shop_info.price,
                         'shop_note': shop_info.note}
            shop_list.append(shop_dict)

        return JsonResponse({'current_shop_data': shop_list}, json_dumps_params={'ensure_ascii': True})


# 수강권/상품 결제 내역 추가
def add_member_shop_info_logic(request):
    class_id = request.session.get('class_id', '')
    member_id = request.POST.get('member_id', None)
    shop_id = request.POST.get('shop_id', None)
    name = request.POST.get('name', '')
    price = request.POST.get('price', 0)
    payment_price = request.POST.get('payment_price', 0)
    pay_method = request.POST.get('pay_method')
    note = request.POST.get('note', '')
    state_cd = request.POST.get('state_cd')

    today = datetime.date.today()
    start_date = today
    end_date = None
    error = None

    if state_cd == STATE_CD_FINISH:
        end_date = today

    if member_id == '' or member_id is None:
        error = '회원 정보를 불러오지 못했습니다.'
    if shop_id == '' or shop_id is None:
        error = '상품 정보를 불러오지 못했습니다.'

    if price == '':
        price = 0
    else:
        try:
            price = int(price)
        except ValueError:
            error = '상품 가격은 숫자만 입력 가능합니다.'

    if payment_price == '':
        payment_price = 0
    else:
        try:
            payment_price = int(payment_price)
        except ValueError:
            error = '납부 금액은 숫자만 입력 가능합니다.'

    if error is None:
        try:
            with transaction.atomic():
                member_shop_info = MemberShopTb(class_tb_id=class_id, member_id=member_id, shop_tb_id=shop_id,
                                                price=price, payment_price=payment_price, pay_method=pay_method,
                                                start_date=start_date, end_date=end_date, note=note, state_cd=state_cd,
                                                use=USE)
                member_shop_info.save()
                if payment_price > 0:
                    member_payment_history_info = MemberPaymentHistoryTb(class_tb_id=class_id,
                                                                         member_id=member_id,
                                                                         member_shop_tb_id=member_shop_info.member_shop_id,
                                                                         payment_price=payment_price,
                                                                         refund_price=0,
                                                                         pay_method=pay_method,
                                                                         pay_date=start_date, note=note, use=USE)
                    member_payment_history_info.save()

        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id, from_member_name=request.user.first_name,
                         class_tb_id=class_id, log_info=name + ' 상품 결제 내역', log_how='추가', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetMemberShopHistoryViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_id = request.GET.get('member_id')
        day = request.GET.get('day')
        page = request.GET.get('page', 1)
        sort = request.GET.get('sort_val', SORT_SHOP_START_DATE)
        today = datetime.date.today()
        member_shop_list = []
        ordered_member_shop_dict = collections.OrderedDict()
        error = None

        if member_id == '':
            error = '회원 정보를 불러오지 못했습니다.'
        # context = {}
        if error is None:
            try:
                check_date = today - datetime.timedelta(days=int(day))
                member_shop_data = MemberShopTb.objects.select_related(
                    'shop_tb').filter(class_tb_id=class_id, member_id=member_id,
                                      start_date__gte=check_date, use=USE).order_by('-start_date', '-reg_dt')
            except ValueError:
                member_shop_data = MemberShopTb.objects.select_related(
                    'shop_tb').filter(class_tb_id=class_id, member_id=member_id, use=USE).order_by('-start_date',
                                                                                                   '-reg_dt')

            paginator = Paginator(member_shop_data, SCHEDULE_PAGINATION_COUNTER)
            try:
                member_shop_data = paginator.page(page)
            except EmptyPage:
                member_shop_data = []
            member_shop_idx = paginator.count - SCHEDULE_PAGINATION_COUNTER * (int(page) - 1)

            ordered_member_shop_dict['max_page'] = paginator.num_pages
            ordered_member_shop_dict['this_page'] = page

        if error is None:
            for member_shop_info in member_shop_data:
                shop_tb = member_shop_info.shop_tb
                shop_id = ''
                shop_name = ''
                shop_price = 0
                shop_note = ''
                if shop_tb is not None :
                    shop_id = member_shop_info.shop_tb.shop_id
                    shop_name = member_shop_info.shop_tb.name
                    shop_price = member_shop_info.price
                    shop_note = member_shop_info.shop_tb.note
                member_shop_dict = {'member_shop_idx': str(member_shop_idx),
                                    'member_shop_id': member_shop_info.member_shop_id,
                                    'member_id': member_shop_info.member_id,
                                    'payment_price': member_shop_info.payment_price,
                                    'refund_price': member_shop_info.refund_price,
                                    'start_date': str(member_shop_info.start_date),
                                    'end_date': str(member_shop_info.end_date),
                                    'note': member_shop_info.note,
                                    'state_cd': member_shop_info.state_cd,
                                    'shop_id': shop_id,
                                    'shop_name': shop_name,
                                    'shop_price': shop_price,
                                    'shop_note': shop_note
                                    }
                member_shop_list.append(member_shop_dict)
                member_shop_idx -= 1
            ordered_member_shop_dict['member_shop_list'] = member_shop_list

        return JsonResponse(ordered_member_shop_dict, json_dumps_params={'ensure_ascii': True})


# 상품 구매 내역 일괄 삭제
def delete_member_shop_data_logic(request):
    class_id = request.session.get('class_id', '')
    member_shop_id = request.POST.get('member_shop_id', '')
    error = None
    member_shop_info = None
    member_name = ''
    shop_name = ''
    try:
        member_shop_info = MemberShopTb.objects.get(member_shop_id=member_shop_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        try:
            with transaction.atomic():
                member_name = member_shop_info.member.name
                shop_name = member_shop_info.shop_tb.name
                member_shop_info.delete()

        except ValueError:
            error = '오류가 발생했습니다. [1]'
        except IntegrityError:
            error = '오류가 발생했습니다. [2]'
        except TypeError:
            error = '오류가 발생했습니다. [3]'
        except ValidationError:
            error = '오류가 발생했습니다. [4]'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id, from_member_name=request.user.first_name,
                         class_tb_id=class_id, log_info=member_name + '님의 '
                                                        + shop_name+' 상품 구매 내역', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 수강권/상품 결제 내역 추가
def add_member_payment_history_info_logic(request):
    class_id = request.session.get('class_id', '')

    member_shop_id = request.POST.get('member_shop_id', None)
    member_ticket_id = request.POST.get('member_ticket_id', None)
    member_id = request.POST.get('member_id', None)
    payment_price = request.POST.get('payment_price', 0)
    refund_price = request.POST.get('refund_price', 0)
    pay_date = request.POST.get('pay_date')
    pay_method = request.POST.get('pay_method')
    note = request.POST.get('note')
    log_message = ''
    error = None

    if member_shop_id == '':
        member_shop_id = None
    if member_ticket_id == '':
        member_ticket_id = None

    if payment_price == '':
        payment_price = 0
    else:
        try:
            payment_price = int(payment_price)
        except ValueError:
            error = '납부 금액은 숫자만 입력 가능합니다.'

    if refund_price == '':
        refund_price = 0
    else:
        try:
            refund_price = int(refund_price)
        except ValueError:
            error = '환불 금액은 숫자만 입력 가능합니다.'

    if member_shop_id is None and member_ticket_id is None:
        error = '오류가 발생했습니다.[0]'

    if error is None:
        try:
            with transaction.atomic():
                if member_ticket_id is not None:
                    try:
                        member_ticket_info = MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
                        if member_ticket_info.state_cd == STATE_CD_REFUND:
                            error = '이미 환불된 구매 내역입니다.'
                        else:
                            if payment_price > 0:
                                member_ticket_info.payment_price += payment_price
                                member_ticket_info.save()
                            if refund_price > 0:
                                member_ticket_info.refund_price = refund_price
                                member_ticket_info.refund_date = pay_date
                                member_ticket_info.state_cd = STATE_CD_REFUND
                                member_ticket_info.save()
                        log_message = member_ticket_info.member.name + '님 ' + member_ticket_info.ticket_tb.name + ' 수강권'
                    except ObjectDoesNotExist:
                        error = '수강권 정보를 불러오지 못했습니다.'

                if member_shop_id is not None:
                    try:
                        member_shop_info = MemberShopTb.objects.get(member_shop_id=member_shop_id)
                        if member_shop_info.state_cd == STATE_CD_REFUND:
                            error = '이미 환불된 구매 내역입니다.'
                        else:
                            if payment_price > 0:
                                member_shop_info.payment_price += payment_price
                                if member_shop_info.payment_price >= member_shop_info.price:
                                    member_shop_info.state_cd = STATE_CD_FINISH
                                    member_shop_info.end_date = pay_date
                                else:
                                    member_shop_info.state_cd = STATE_CD_IN_PROGRESS
                                member_shop_info.save()
                            if refund_price > 0:
                                member_shop_info.refund_price = refund_price
                                member_shop_info.state_cd = STATE_CD_REFUND
                                member_shop_info.end_date = pay_date
                                member_shop_info.save()
                        log_message = member_shop_info.member.name + '님 ' + member_shop_info.shop_tb.name + '상품'
                    except ObjectDoesNotExist:
                        error = '상품 구매 정보를 불러오지 못했습니다.'

                if error is None:
                    member_payment_history_info = MemberPaymentHistoryTb(class_tb_id=class_id,
                                                                         member_id=member_id,
                                                                         member_ticket_tb_id=member_ticket_id,
                                                                         member_shop_tb_id=member_shop_id,
                                                                         payment_price=payment_price,
                                                                         refund_price=refund_price,
                                                                         pay_method=pay_method,
                                                                         pay_date=pay_date, note=note, use=USE)
                    member_payment_history_info.save()

        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id, from_member_name=request.user.first_name,
                         class_tb_id=class_id, log_info=log_message + ' 결제 내역', log_how='추가', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetMemberPaymentHistoryViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        member_shop_id = request.GET.get('member_shop_id', None)
        member_ticket_id = request.GET.get('member_ticket_id', None)
        member_payment_list = []
        member_payment_data = []
        ordered_member_payment_dict = collections.OrderedDict()
        error = None

        if member_shop_id == '':
            member_shop_id = None
        if member_ticket_id == '':
            member_ticket_id = None

        if member_ticket_id is not None:
            member_payment_data = MemberPaymentHistoryTb.objects.select_related(
                'member_ticket_tb').filter(class_tb_id=class_id, member_ticket_tb_id=member_ticket_id,
                                           use=USE).order_by('-pay_date', '-reg_dt')

        if member_shop_id is not None:
            member_payment_data = MemberPaymentHistoryTb.objects.select_related(
                'member_shop_tb').filter(class_tb_id=class_id, member_shop_tb_id=member_shop_id,
                                         use=USE).order_by('-pay_date', '-reg_dt')

        if error is None:
            member_payment_history_idx = len(member_payment_data)
            for member_payment_info in member_payment_data:
                member_payment_dict = {'member_payment_history_idx': member_payment_history_idx,
                                       'member_payment_history_id': member_payment_info.member_payment_history_id,
                                       'payment_price': member_payment_info.payment_price,
                                       'refund_price': member_payment_info.refund_price,
                                       'pay_date': str(member_payment_info.pay_date),
                                       'pay_method': member_payment_info.pay_method,
                                       'note': str(member_payment_info.note)
                                       }
                member_payment_list.append(member_payment_dict)
                member_payment_history_idx -= 1
            ordered_member_payment_dict['member_payment_list'] = member_payment_list
        return JsonResponse(ordered_member_payment_dict, json_dumps_params={'ensure_ascii': True})


# 상품 결제 내역 삭제
def delete_member_payment_history_info_logic(request):
    class_id = request.session.get('class_id', '')
    member_payment_history_id = request.POST.get('member_payment_history_id', '')
    error = None
    member_payment_history_info = None
    member_name = ''
    shop_name = ''
    try:
        member_payment_history_info = MemberPaymentHistoryTb.objects.get(member_payment_history_id=member_payment_history_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        try:
            with transaction.atomic():
                member_name = member_payment_history_info.member.name
                member_shop_tb = member_payment_history_info.member_shop_tb
                member_ticket_tb = member_payment_history_info.member_ticket_tb
                if member_shop_tb is not None and member_shop_tb != '':
                    shop_name = member_shop_tb.shop_tb.name
                    sum_payment_price = MemberPaymentHistoryTb.objects.filter(member_shop_tb_id=member_shop_tb.member_shop_id,
                                                                              use=USE).aggregate(sum_payment_price=Sum('payment_price'))
                    sum_refund_price = MemberPaymentHistoryTb.objects.filter(member_shop_tb_id=member_shop_tb.member_shop_id,
                                                                             use=USE).aggregate(sum_refund_price=Sum('refund_price'))

                    member_shop_tb.payment_price = sum_payment_price['sum_payment_price']\
                                                   - member_payment_history_info.payment_price
                    member_shop_tb.refund_price = sum_refund_price['sum_refund_price']\
                                                  - member_payment_history_info.refund_price
                    if member_shop_tb.payment_price >= member_shop_tb.price:
                        member_shop_tb.state_cd = STATE_CD_FINISH
                    if member_shop_tb.payment_price == 0:
                        member_shop_tb.state_cd = STATE_CD_NOT_PROGRESS
                    else:
                        member_shop_tb.state_cd = STATE_CD_IN_PROGRESS

                    if member_shop_tb.refund_price > 0:
                        member_shop_tb.state_cd = STATE_CD_REFUND

                    member_shop_tb.save()
                if member_ticket_tb is not None and member_ticket_tb != '':
                    shop_name = member_ticket_tb.ticket_tb.name
                    sum_payment_price = MemberPaymentHistoryTb.objects.filter(member_ticket_tb_id=member_ticket_tb.member_ticket_id,
                                                                              use=USE).aggregate(sum_payment_price=Sum('payment_price'))
                    sum_refund_price = MemberPaymentHistoryTb.objects.filter(member_ticket_tb_id=member_ticket_tb.member_ticket_id,
                                                                             use=USE).aggregate(sum_refund_price=Sum('refund_price'))

                    member_ticket_tb.payment_price = sum_payment_price['sum_payment_price']\
                                                     - member_payment_history_info.payment_price
                    member_ticket_tb.refund_price = sum_refund_price['sum_refund_price']\
                                                    - member_payment_history_info.refund_price

                    if member_payment_history_info.refund_price > 0:
                        schedule_data_count = ScheduleTb.objects.filter(member_ticket_tb_id=member_ticket_tb.member_ticket_id).count()
                        schedule_data_finish_count = ScheduleTb.objects.filter(Q(state_cd=STATE_CD_FINISH)
                                                                               | Q(state_cd=STATE_CD_ABSENCE),
                                                                               member_ticket_tb_id=member_ticket_tb.member_ticket_id).count()

                        member_ticket_avail_count = 0
                        member_ticket_rem_count = 0
                        if member_ticket_tb.member_ticket_reg_count >= schedule_data_count:
                            member_ticket_avail_count = member_ticket_tb.member_ticket_reg_count - schedule_data_count

                        if member_ticket_tb.member_ticket_reg_count >= schedule_data_finish_count:
                            member_ticket_rem_count = member_ticket_tb.member_ticket_reg_count - schedule_data_finish_count
                        member_ticket_tb.member_ticket_avail_count = member_ticket_avail_count
                        member_ticket_tb.member_ticket_rem_count = member_ticket_rem_count
                        member_ticket_tb.refund_price = 0
                        member_ticket_tb.refund_date = None
                        member_ticket_tb.state_cd = STATE_CD_IN_PROGRESS
                    member_ticket_tb.save()

                member_payment_history_info.delete()

        except ValueError:
            error = '오류가 발생했습니다. [1]'
        except IntegrityError:
            error = '오류가 발생했습니다. [2]'
        except TypeError:
            error = '오류가 발생했습니다. [3]'
        except ValidationError:
            error = '오류가 발생했습니다. [4]'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id, from_member_name=request.user.first_name,
                         class_tb_id=class_id, log_info=member_name + '님의 ' + shop_name+' 결제 내역',
                         log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 출석 체크 완료 기능
def attend_mode_check_logic(request):
    class_id = request.session.get('class_id', '')
    phone_number = request.POST.get('phone_number', '')
    schedule_id = request.POST.get('schedule_id', '')
    error = None
    member_ticket_id = None

    try:
        schedule_info = ScheduleTb.objects.get(schedule_id=schedule_id)
    except ObjectDoesNotExist:
        schedule_info = None

    member_data = ClassMemberTicketTb.objects.filter(class_tb_id=class_id, member_ticket_tb__member__phone=phone_number,
                                                     auth_cd=AUTH_TYPE_VIEW, use=USE)
    member_id_list = []
    for member_info in member_data:
        member_id_test = 0
        for member_id_element in member_id_list:
            if member_id_element == member_info.member_ticket_tb.member_id:
                member_id_test += 1
                break
        if member_id_test == 0:
            member_id_list.append(member_info.member_ticket_tb.member_id)

    if len(member_id_list) > 1:
        error = '중복되는 휴대폰 번호 2개 이상 존재합니다. 강사님에게 문의해주세요.'

    if error is None:
        if len(member_data) > 0:
            member_id = member_data[0].member_ticket_tb.member_id
            if schedule_info.member_ticket_tb is None or schedule_info.member_ticket_tb == '':
                try:
                    lecture_schedule_info = ScheduleTb.objects.get(lecture_schedule_id=schedule_id,
                                                                   lecture_tb_id=schedule_info.lecture_tb_id,
                                                                   member_ticket_tb__member_id=member_id)
                    if lecture_schedule_info.state_cd == STATE_CD_FINISH:
                        error = '이미 출석 처리된 수업입니다.'

                except ObjectDoesNotExist:
                    member_ticket_result = func_get_lecture_member_ticket_id(class_id, schedule_info.lecture_tb_id,
                                                                             member_id)

                    if member_ticket_result['error'] is not None:
                        error = member_ticket_result['error']
                    else:
                        member_ticket_id = member_ticket_result['member_ticket_id']

                    # if error is None:
                    #     if member_ticket_id is None or member_ticket_id == '':
                    #         member_ticket_id = func_get_member_ticket_id_old(class_id, member_id)

                    if error is None:
                        if member_ticket_id is None or member_ticket_id == '':
                            error = '선택하신 수업의 예약 가능 횟수를 확인해주세요.'
                        else:
                            try:
                                MemberTicketTb.objects.get(member_ticket_id=member_ticket_id)
                            except ObjectDoesNotExist:
                                error = '수강권 정보를 불러오지 못했습니다.'
            else:
                if schedule_info.state_cd == STATE_CD_FINISH:
                    error = '이미 출석 처리된 수업입니다.'
                if error is None:
                    if schedule_info.member_ticket_tb.member_id != member_id:
                        error = '휴대폰 번호가 일치하지 않습니다.'
        else:
            error = '휴대폰 번호를 확인해주세요.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html', {'member_ticket_id': member_ticket_id})


# 출석 체크 완료 기능
def attend_mode_finish_logic(request):
    member_id = request.POST.get('member_id')
    schedule_id = request.POST.get('schedule_id')
    class_id = request.session.get('class_id', '')
    next_page = '/trainer/attend_mode/'
    error = None
    try:
        schedule_info = ScheduleTb.objects.get(schedule_id=schedule_id)
    except ObjectDoesNotExist:
        schedule_info = None

    if member_id is not None:
        try:
            with transaction.atomic():
                if schedule_info.member_ticket_tb is None or schedule_info.member_ticket_tb == '':
                    try:
                        lecture_schedule_info = ScheduleTb.objects.get(lecture_schedule_id=schedule_id,
                                                                       lecture_tb_id=schedule_info.lecture_tb_id,
                                                                       member_ticket_tb__member_id=member_id)
                        if lecture_schedule_info.state_cd == STATE_CD_FINISH:
                            error = '이미 출석 처리된 수업입니다.'

                        if error is None:
                            lecture_schedule_info.state_cd = STATE_CD_FINISH
                            lecture_schedule_info.save()
                            error = func_refresh_member_ticket_count(class_id,
                                                                     lecture_schedule_info.member_ticket_tb_id)

                    except ObjectDoesNotExist:
                        member_ticket_id = None
                        member_ticket_result = func_get_lecture_member_ticket_id(class_id, schedule_info.lecture_tb_id,
                                                                                 member_id)

                        if member_ticket_result['error'] is not None:
                            error = member_ticket_result['error']
                        else:
                            member_ticket_id = member_ticket_result['member_ticket_id']

                        # if error is None:
                        #     if member_ticket_id is None or member_ticket_id == '':
                        #         member_ticket_id = func_get_member_ticket_id_old(class_id, member_id)

                        if error is None:
                            if member_ticket_id is None or member_ticket_id == '':
                                error = '선택하신 수업의 예약 가능 횟수가 없습니다.'
                        if error is None:
                            schedule_result = func_add_schedule(class_id, member_ticket_id, None,
                                                                schedule_info.lecture_tb, schedule_id,
                                                                schedule_info.start_dt, schedule_info.end_dt,
                                                                schedule_info.note, schedule_info.private_note,
                                                                ON_SCHEDULE_TYPE,
                                                                member_id, PERMISSION_STATE_CD_APPROVE,
                                                                STATE_CD_FINISH, UN_USE, SCHEDULE_DUPLICATION_ENABLE,
                                                                schedule_info.trainer.member_id)
                            error = schedule_result['error']

                else:
                    if schedule_info.member_ticket_tb.member_id == member_id:
                        if schedule_info.state_cd == STATE_CD_FINISH:
                            error = '이미 출석 처리된 수업입니다.'
                        if error is None:
                            schedule_info.state_cd = STATE_CD_FINISH
                            schedule_info.save()
                            error = func_refresh_member_ticket_count(class_id,
                                                                     schedule_info.member_ticket_tb.member_ticket_id)

        except TypeError:
            error = error
        except ValueError:
            error = error
        except IntegrityError:
            error = error
        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
        next_page = '/trainer/attend_mode_detail/'

    return redirect(next_page)


# 강사 출석 기능 setting 업데이트 api
def update_attend_mode_setting_logic(request):
    setting_admin_password = request.POST.get('setting_admin_password', '0000')
    setting_attend_class_prev_display_time = request.POST.get('setting_attend_class_prev_display_time', '5')
    setting_attend_class_after_display_time = request.POST.get('setting_attend_class_after_display_time', '5')
    setting_attend_mode_max_num_view_available = request.POST.get('setting_attend_mode_max_num_view_available', USE)

    class_id = request.session.get('class_id', '')
    # next_page = request.POST.get('next_page', '/trainer/attend_mode/')

    if setting_admin_password is None or setting_admin_password == '':
        setting_admin_password = '0000'
    if setting_attend_class_prev_display_time is None or setting_attend_class_prev_display_time == '':
        setting_attend_class_prev_display_time = '5'
    if setting_attend_class_after_display_time is None or setting_attend_class_after_display_time == '':
        setting_attend_class_after_display_time = '5'
    if setting_attend_mode_max_num_view_available is None or setting_attend_mode_max_num_view_available == '':
        setting_attend_mode_max_num_view_available = USE
    setting_type_cd_data = ['LT_ADMIN_PASSWORD', 'LT_ATTEND_CLASS_PREV_DISPLAY_TIME',
                            'LT_ATTEND_CLASS_AFTER_DISPLAY_TIME', 'LT_ATTEND_CLASS_MAX_NUM_VIEW']
    setting_info_data = [setting_admin_password, setting_attend_class_prev_display_time,
                         setting_attend_class_after_display_time, setting_attend_mode_max_num_view_available]
    error = update_program_setting_data(class_id, setting_type_cd_data, setting_info_data)

    if error is None:
        request.session['setting_admin_password'] = setting_admin_password
        request.session['setting_attend_class_prev_display_time'] = setting_attend_class_prev_display_time
        request.session['setting_attend_class_after_display_time'] = setting_attend_class_after_display_time
        request.session['setting_attend_mode_max_num_view_available'] = setting_attend_mode_max_num_view_available
    else:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


def check_admin_password_logic(request):
    setting_admin_password = request.POST.get('setting_admin_password', '')
    class_id = request.session.get('class_id', '')

    error = None

    try:
        admin_password = SettingTb.objects.get(class_tb_id=class_id,
                                               setting_type_cd='LT_ADMIN_PASSWORD').setting_info
    except ObjectDoesNotExist:
        admin_password = '0000'

    if admin_password != setting_admin_password:
        error = '잠금 해제 비밀번호가 일치하지 않습니다.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)

    return render(request, 'ajax/trainer_error_ajax.html')


class GetAttendModeScheduleView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'ajax/attend_schedule_ajax.html'

    def get_context_data(self, **kwargs):
        context = super(GetAttendModeScheduleView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        setting_attend_class_prev_display_time = 0
        setting_attend_class_after_display_time = 0
        current_time = timezone.now()
        setting_data = SettingTb.objects.filter(class_tb_id=class_id, use=USE)

        for setting_info in setting_data:
            if setting_info.setting_type_cd == 'LT_ATTEND_CLASS_PREV_DISPLAY_TIME':
                setting_attend_class_prev_display_time = int(setting_info.setting_info)
            if setting_info.setting_type_cd == 'LT_ATTEND_CLASS_AFTER_DISPLAY_TIME':
                setting_attend_class_after_display_time = int(setting_info.setting_info)

        start_date = current_time + datetime.timedelta(minutes=int(setting_attend_class_prev_display_time))
        end_date = current_time - datetime.timedelta(minutes=int(setting_attend_class_after_display_time))
        context = func_get_trainer_attend_schedule(context, class_id, start_date, end_date, current_time)

        return context


# 리뉴얼
class PopupCalendarPlanView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_plan_view.html'

    def get_context_data(self, **kwargs):
        context = super(PopupCalendarPlanView, self).get_context_data(**kwargs)
        return context


class PopupCalendarPlanAdd(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_plan_add.html'

    def get_context_data(self, **kwargs):
        context = super(PopupCalendarPlanAdd, self).get_context_data(**kwargs)
        return context


class PopupMemberAdd(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_member_add.html'

    def get_context_data(self, **kwargs):
        context = super(PopupMemberAdd, self).get_context_data(**kwargs)
        return context


class PopupMemberView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_member_view.html'

    def get_context_data(self, **kwargs):
        context = super(PopupMemberView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        member_id = self.request.GET.get('member_id')
        member_result = func_get_member_info(class_id, self.request.user.id, member_id)
        member_ticket_list = dict(func_get_member_ticket_list(class_id, member_id))
        member_lecture_list = dict(func_get_member_lecture_list(class_id, member_id))
        context['member_info'] = member_result['member_info']
        context['member_lecture_list'] = member_lecture_list
        context['member_ticket_list'] = member_ticket_list
        return context


class PopupMemberSimpleView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_member_simple_view.html'

    def get_context_data(self, **kwargs):
        context = super(PopupMemberSimpleView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        member_id = self.request.GET.get('member_id')
        member_result = func_get_member_info(class_id, self.request.user.id, member_id)
        member_ticket_list = dict(func_get_member_ticket_list(class_id, member_id))
        member_lecture_list = dict(func_get_member_lecture_list(class_id, member_id))
        context['member_info'] = member_result['member_info']
        context['member_lecture_list'] = member_lecture_list
        context['member_ticket_list'] = member_ticket_list
        return context


class PopupMemberEdit(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_member_edit.html'

    def get_context_data(self, **kwargs):
        context = super(PopupMemberEdit, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        member_id = self.request.GET.get('member_id')
        member_result = func_get_member_info(class_id, self.request.user.id, member_id)
        context['member_info'] = member_result['member_info']
        return context


class PopupLectureAdd(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_lecture_add.html'

    def get_context_data(self, **kwargs):
        context = super(PopupLectureAdd, self).get_context_data(**kwargs)
        return context


class PopupLectureView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_lecture_view.html'

    def get_context_data(self, **kwargs):
        context = super(PopupLectureView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        lecture_id = self.request.GET.get('lecture_id')
        context['lecture_info'] = func_get_lecture_info(class_id, lecture_id, self.request.user.id)
        return context


class PopupLectureSimpleView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_lecture_simple_view.html'

    def get_context_data(self, **kwargs):
        context = super(PopupLectureSimpleView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        lecture_id = self.request.GET.get('lecture_id')
        context['lecture_info'] = func_get_lecture_info(class_id, lecture_id, self.request.user.id)
        return context


class PopupLectureEdit(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_lecture_edit.html'

    def get_context_data(self, **kwargs):
        context = super(PopupLectureEdit, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        lecture_id = self.request.GET.get('lecture_id')
        context['lecture_info'] = func_get_lecture_info(class_id, lecture_id, self.request.user.id)
        return context


class PopupTicketAdd(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_ticket_add.html'

    def get_context_data(self, **kwargs):
        context = super(PopupTicketAdd, self).get_context_data(**kwargs)
        return context


class PopupTicketView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_ticket_view.html'

    def get_context_data(self, **kwargs):
        context = super(PopupTicketView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        ticket_id = self.request.GET.get('ticket_id')

        context['ticket_info'] = func_get_ticket_info(class_id, ticket_id, self.request.user.id)
        return context

class PopupTicketSimpleView(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_ticket_simple_view.html'

    def get_context_data(self, **kwargs):
        context = super(PopupTicketSimpleView, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        ticket_id = self.request.GET.get('ticket_id')

        context['ticket_info'] = func_get_ticket_info(class_id, ticket_id, self.request.user.id)
        return context


class PopupTicketEdit(LoginRequiredMixin, AccessTestMixin, TemplateView):
    template_name = 'popup/trainer_popup_ticket_edit.html'

    def get_context_data(self, **kwargs):
        context = super(PopupTicketEdit, self).get_context_data(**kwargs)
        class_id = self.request.session.get('class_id')
        ticket_id = self.request.GET.get('ticket_id')

        context['ticket_info'] = func_get_ticket_info(class_id, ticket_id, self.request.user.id)
        return context


class PopupMemberSelect(TemplateView):
    template_name = 'popup/trainer_popup_member_select.html'

    def get_context_data(self, **kwargs):
        context = super(PopupMemberSelect, self).get_context_data(**kwargs)
        return context


class PopupLectureSelect(TemplateView):
    template_name = 'popup/trainer_popup_lecture_select.html'

    def get_context_data(self, **kwargs):
        context = super(PopupLectureSelect, self).get_context_data(**kwargs)
        return context


class PopupTicketSelect(TemplateView):
    template_name = 'popup/trainer_popup_ticket_select.html'

    def get_context_data(self, **kwargs):
        context = super(PopupTicketSelect, self).get_context_data(**kwargs)
        return context


class PopupColorSelect(TemplateView):
    template_name = 'popup/trainer_popup_color_select.html'

    def get_context_data(self, **kwargs):
        context = super(PopupColorSelect, self).get_context_data(**kwargs)
        return context


def update_trainer_board_content_img_logic(request):
    error_message = None
    img_url = None
    context = {}
    if request.method == 'POST':
        # 대표 이미지 설정
        try:
            img_url = func_upload_board_content_image_logic(request.FILES['content_img_file'],
                                                            request.POST.get('content_img_file_name'),
                                                            request.POST.get('board_type_cd'),
                                                            request.user.id, 'trainer')
        except MultiValueDictKeyError:
            img_url = None
    else:
        error_message = '잘못된 요청입니다.'

    if img_url is None:
        error_message = '이미지 업로드중 오류가 발생했습니다.'

    if error_message is not None:
        messages.error(request, error_message)
        context['messageArray'] = error_message
    else:
        context['img_url'] = img_url
    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


# 프로필 사진 수정
def update_member_profile_img_logic(request):
    error = None
    member_info = None
    img_url = None
    member_id = request.POST.get('member_id', '')
    class_id = request.session.get('class_id', '')
    # group_name = request.session.get('group_name', 'trainer')
    group_name = 'trainee'

    try:
        member_info = MemberTb.objects.get(member_id=member_id)
    except ObjectDoesNotExist:
        error = '회원 정보를 불러오지 못했습니다.'

    if error is None:
        if member_info.user.is_active:
            error = '회원가입이 완료된 회원의 프로필 이미지는 등록/수정이 불가능합니다.'
        if str(member_info.reg_info) != str(request.user.id):
            error = '직접 등록한 회원만 프로필 이미지만 등록/수정이 가능합니다.'

    if error is None:
        if member_info.profile_url is not None and member_info.profile_url != '':
            error = func_delete_profile_image_logic(member_info.profile_url)

    if error is None:
        max_range = 9999999999
        random_file_name = str(random.randrange(0, max_range)).zfill(len(str(max_range)))
        if request.method == 'POST':
            try:
                img_url = func_upload_profile_image_logic(request.POST.get('profile_img_file'),
                                                          str(member_id)+'/'+str(random_file_name), group_name)
                if img_url is None:
                    error = '프로필 이미지 변경에 실패했습니다.[1]'
            except MultiValueDictKeyError:
                error = '프로필 이미지 변경에 실패했습니다.[2]'

    if error is None:
        member_info.profile_url = img_url
        member_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name+' 회원 프로필', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 프로필 사진 삭제
def delete_member_profile_img_logic(request):
    error = None
    member_info = None
    member_id = request.POST.get('member_id', '')
    class_id = request.session.get('class_id', '')

    try:
        member_info = MemberTb.objects.get(member_id=member_id)
    except ObjectDoesNotExist:
        error = '회원 정보를 불러오지 못했습니다.'

    if error is None:
        if member_info.user.is_active:
            error = '회원가입이 완료된 회원의 프로필 이미지는 삭제가 불가능합니다.'
        if str(member_info.reg_info) != str(request.user.id):
            error = '직접 등록한 회원만 프로필 이미지 삭제 가능합니다.'

    if error is None:
        if member_info.profile_url is not None and member_info.profile_url != '':
            error = func_delete_profile_image_logic(member_info.profile_url)

    if error is None:
        member_info.profile_url = '/static/common/icon/icon_account.png'
        member_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name+' 회원 프로필', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 프로필 사진 수정
def update_trainer_profile_img_logic(request):
    error = None
    member_info = None
    img_url = None
    member_id = request.POST.get('trainer_id', '')
    class_id = request.session.get('class_id', '')
    # group_name = request.session.get('group_name', 'trainer')
    group_name = 'trainer'

    try:
        member_info = MemberTb.objects.get(member_id=member_id)
    except ObjectDoesNotExist:
        error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        if member_info.user.is_active:
            error = '회원가입이 완료된 강사의 프로필 이미지는 등록/수정이 불가능합니다.'
        if str(member_info.reg_info) != str(request.user.id):
            error = '직접 등록한 강사만 프로필 이미지만 등록/수정이 가능합니다.'

    if error is None:
        if member_info.profile_url is not None and member_info.profile_url != '':
            error = func_delete_profile_image_logic(member_info.profile_url)

    if error is None:
        max_range = 9999999999
        random_file_name = str(random.randrange(0, max_range)).zfill(len(str(max_range)))
        if request.method == 'POST':
            try:
                img_url = func_upload_profile_image_logic(request.POST.get('profile_img_file'),
                                                          str(member_id)+'/'+str(random_file_name), group_name)
                if img_url is None:
                    error = '프로필 이미지 변경에 실패했습니다.[1]'
            except MultiValueDictKeyError:
                error = '프로필 이미지 변경에 실패했습니다.[2]'

    if error is None:
        member_info.profile_url = img_url
        member_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name+' 강사 프로필', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 프로필 사진 삭제
def delete_trainer_profile_img_logic(request):
    error = None
    member_info = None
    member_id = request.POST.get('trainer_id', '')
    class_id = request.session.get('class_id', '')

    try:
        member_info = MemberTb.objects.get(member_id=member_id)
    except ObjectDoesNotExist:
        error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        if member_info.user.is_active:
            error = '회원가입이 완료된 강사의 프로필 이미지는 삭제가 불가능합니다.'
        if str(member_info.reg_info) != str(request.user.id):
            error = '직접 등록한 강사만 프로필 이미지 삭제 가능합니다.'

    if error is None:
        if member_info.profile_url is not None and member_info.profile_url != '':
            error = func_delete_profile_image_logic(member_info.profile_url)

    if error is None:
        member_info.profile_url = '/static/common/icon/icon_account.png'
        member_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name+' 강사 프로필', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetTrainerIngListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        trainer_sort = request.GET.get('sort_val', SORT_TRAINER_NAME)
        sort_order_by = request.GET.get('sort_order_by', SORT_ASC)
        keyword = request.GET.get('keyword', '')
        current_trainer_data = []
        error = None
        try:
            class_info = ClassTb.objects.get(class_id=class_id)
            class_trainer_info = class_info.member
        except ObjectDoesNotExist:
            error = '지점 정보를 불러오지 못했습니다.'

        if error is None:
            current_trainer_data = func_get_trainer_ing_list(class_id, request.user.id, keyword)
            finish_trainer_num = len(func_get_class_trainer_end_list(class_id, keyword))
            sort_info = int(trainer_sort)
            trainer_data = {'trainer_id': class_trainer_info.member_id,
                            'trainer_user_id': class_trainer_info.user.username,
                            'trainer_name': class_trainer_info.name,
                            'trainer_phone': str(class_trainer_info.phone),
                            'trainer_email': str(class_trainer_info.user.email),
                            'trainer_sex': str(class_trainer_info.sex),
                            'trainer_profile_url': class_trainer_info.profile_url,
                            'trainer_birthday_dt': str(class_trainer_info.birthday_dt),
                            'trainer_connection_type': AUTH_TYPE_VIEW}
            if sort_info == SORT_TRAINER_NAME:
                current_trainer_data = sorted(current_trainer_data, key=lambda elem: elem['trainer_name'],
                                              reverse=int(sort_order_by))
                current_trainer_data.insert(0, trainer_data)

        # context['total_member_num'] = len(member_data)
        # if page != 0:
        #     paginator = Paginator(member_data, 20)  # Show 20 contacts per page
        #     try:
        #         member_data = paginator.page(page)
        #     except EmptyPage:
        #         member_data = None
        #
        # context['member_data'] = member_data
        # end_dt = timezone.now()

        return JsonResponse({'current_trainer_data': current_trainer_data, 'finish_trainer_num': finish_trainer_num,
                             'error': error},
                            json_dumps_params={'ensure_ascii': True})


class GetTrainerIngListConnectViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        trainer_sort = request.GET.get('sort_val', SORT_TRAINER_NAME)
        sort_order_by = request.GET.get('sort_order_by', SORT_ASC)
        keyword = request.GET.get('keyword', '')
        current_trainer_data = []
        error = None
        try:
            class_info = ClassTb.objects.get(class_id=class_id)
            class_trainer_info = class_info.member
        except ObjectDoesNotExist:
            error = '지점 정보를 불러오지 못했습니다.'

        if error is None:
            current_trainer_data = func_get_trainer_ing_list_connect(class_id, request.user.id, keyword)
            finish_trainer_num = len(func_get_class_trainer_end_list(class_id, keyword))
            sort_info = int(trainer_sort)
            trainer_data = {'trainer_id': class_trainer_info.member_id,
                            'trainer_user_id': class_trainer_info.user.username,
                            'trainer_name': class_trainer_info.name,
                            'trainer_phone': str(class_trainer_info.phone),
                            'trainer_email': str(class_trainer_info.user.email),
                            'trainer_sex': str(class_trainer_info.sex),
                            'trainer_profile_url': class_trainer_info.profile_url,
                            'trainer_birthday_dt': str(class_trainer_info.birthday_dt),
                            'trainer_connection_type': AUTH_TYPE_VIEW}
            if sort_info == SORT_TRAINER_NAME:
                current_trainer_data = sorted(current_trainer_data, key=lambda elem: elem['trainer_name'],
                                              reverse=int(sort_order_by))
                current_trainer_data.insert(0, trainer_data)

        return JsonResponse({'current_trainer_data': current_trainer_data, 'finish_trainer_num': finish_trainer_num,
                             'error': error},
                            json_dumps_params={'ensure_ascii': True})


class GetTrainerEndListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        # start_dt = timezone.now()
        class_id = self.request.session.get('class_id', '')
        # page = request.GET.get('page', 0)
        trainer_sort = request.GET.get('sort_val', SORT_TRAINER_NAME)
        sort_order_by = request.GET.get('sort_order_by', SORT_ASC)
        keyword = request.GET.get('keyword', '')

        finish_trainer_data = func_get_trainer_end_list(class_id, request.user.id, keyword)
        current_trainer_num = len(func_get_class_trainer_ing_list(class_id, keyword))

        sort_info = int(trainer_sort)
        if sort_info == SORT_TRAINER_NAME:
            finish_trainer_data = sorted(finish_trainer_data, key=lambda elem: elem['trainer_name'],
                                         reverse=int(sort_order_by))

        # context['total_member_num'] = len(member_data)
        # if page != 0:
        #     paginator = Paginator(member_data, 20)  # Show 20 contacts per page
        #     try:
        #         member_data = paginator.page(page)
        #     except EmptyPage:
        #         member_data = None
        #
        # end_dt = timezone.now()
        return JsonResponse({'finish_trainer_data': finish_trainer_data, 'current_trainer_num': current_trainer_num},
                            json_dumps_params={'ensure_ascii': True})


def add_trainer_program_info_logic(request):
    trainer_id = request.POST.get('trainer_id', '')
    auth_cd = request.POST.get('auth_cd', '')
    class_id = request.session.get('class_id', '')

    error = None
    member_info = None
    if trainer_id is None or trainer_id == '':
        error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        if str(request.user.id) == str(trainer_id):
            error = '본인은 등록할수 없습니다.'

    if error is None:
        try:
            member_info = MemberTb.objects.get(member_id=trainer_id)
        except ObjectDoesNotExist:
            error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            class_member_info = MemberClassTb.objects.select_related(
                'class_tb', 'member').get(class_tb_id=class_id, member_id=trainer_id)
            class_member_info.auth_cd = auth_cd
            class_member_info.own_cd = OWN_TYPE_EMPLOYEE
            class_member_info.mod_member_id = request.user.id
            class_member_info.use=USE
            class_member_info.save()
        except ObjectDoesNotExist:
            class_member_info = MemberClassTb(class_tb_id=class_id, member_id=trainer_id, auth_cd=auth_cd,
                                              mod_member_id=request.user.id, own_cd=OWN_TYPE_EMPLOYEE, use=USE)
            class_member_info.save()

        function_list = ProductFunctionAuthTb.objects.select_related(
            'function_auth_tb', 'product_tb').filter(product_tb_id=6,
                                                     use=USE).order_by('product_tb_id',
                                                                       'function_auth_tb_id',
                                                                       'auth_type_cd')
        for function_info in function_list:
            if function_info.auth_type_cd is None:
                function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd)
            else:
                function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd) \
                                             + str(function_info.auth_type_cd)

            enable_flag = request.POST.get(function_auth_type_cd_name, '')
            if enable_flag is not None and enable_flag != '':
                try:
                    program_auth_info = ProgramAuthTb.objects.get(class_tb_id=class_id, member_id=trainer_id,
                                                                  function_auth_tb=function_info.function_auth_tb,
                                                                  auth_type_cd=function_info.auth_type_cd)
                except ObjectDoesNotExist:
                    program_auth_info = ProgramAuthTb(class_tb_id=class_id, member_id=trainer_id,
                                                      function_auth_tb=function_info.function_auth_tb,
                                                      auth_type_cd=function_info.auth_type_cd,
                                                      use=USE)

                program_auth_info.enable_flag = enable_flag
                program_auth_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_how = '강사 추가'
        log_data = LogTb(log_type='LP02', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name + '님 에게 \''
                                  + request.session.get('class_type_name', '') + '\' 지점',
                         log_how=log_how, log_detail='', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


def update_trainer_program_info_logic(request):
    trainer_id = request.POST.get('trainer_id', '')
    # auth_cd = request.POST.get('auth_cd', '')
    class_id = request.session.get('class_id', '')

    error = None
    member_info = None
    if trainer_id is None or trainer_id == '':
        error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        if str(request.user.id) == str(trainer_id):
            error = '본인은 등록할수 없습니다.'
    if error is None:
        try:
            member_info = MemberTb.objects.get(member_id=trainer_id)
        except ObjectDoesNotExist:
            error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            class_member_info = MemberClassTb.objects.select_related(
                'class_tb', 'member').get(class_tb_id=class_id, member_id=trainer_id, use=USE)
            # class_member_info.auth_cd = auth_cd
            class_member_info.own_cd = OWN_TYPE_EMPLOYEE
            class_member_info.mod_member_id = request.user.id
            class_member_info.save()
        except ObjectDoesNotExist:
            error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        function_list = ProductFunctionAuthTb.objects.select_related(
            'function_auth_tb', 'product_tb').filter(product_tb_id=6,
                                                     use=USE).order_by('product_tb_id',
                                                                       'function_auth_tb_id',
                                                                       'auth_type_cd')
        for function_info in function_list:
            if function_info.auth_type_cd is None:
                function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd)
            else:
                function_auth_type_cd_name = str(function_info.function_auth_tb.function_auth_type_cd) \
                                             + str(function_info.auth_type_cd)

            enable_flag = request.POST.get(function_auth_type_cd_name, '')
            if enable_flag is not None and enable_flag != '':
                try:
                    program_auth_info = ProgramAuthTb.objects.get(class_tb_id=class_id, member_id=trainer_id,
                                                                  function_auth_tb=function_info.function_auth_tb,
                                                                  auth_type_cd=function_info.auth_type_cd)
                except ObjectDoesNotExist:
                    program_auth_info = ProgramAuthTb(class_tb_id=class_id, member_id=trainer_id,
                                                      function_auth_tb=function_info.function_auth_tb,
                                                      auth_type_cd=function_info.auth_type_cd,
                                                      use=USE)

                program_auth_info.enable_flag = enable_flag
                program_auth_info.save()

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_how = '강사 권한 변경'
        log_data = LogTb(log_type='LP02', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name + '님 에게 \''
                                  + request.session.get('class_type_name', '') + '\' 지점',
                         log_how=log_how, log_detail='', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetTrainerProgramDataViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        trainer_id = request.GET.get('trainer_id', '')
        error = None
        member_program_auth_list = collections.OrderedDict()

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if error is None:
            program_auth_data = ProgramAuthTb.objects.select_related('class_tb', 'member',
                                                                     'function_auth_tb').filter(class_tb_id=class_id,
                                                                                                member_id=trainer_id,
                                                                                                use=USE)

            for program_auth_info in program_auth_data:
                if program_auth_info.auth_type_cd is None:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd)
                else:
                    function_auth_type_cd_name = str(program_auth_info.function_auth_tb.function_auth_type_cd) \
                                                 + str(program_auth_info.auth_type_cd)

                try:
                    member_program_auth_list[program_auth_info.member_id]
                except KeyError:
                    member_program_auth_list[program_auth_info.member_id] = {}
                    member_result = func_get_trainer_info(class_id, program_auth_info.member_id)
                    member_program_auth_list[program_auth_info.member_id]['member_info'] \
                        = member_result['member_info']

                member_program_auth_list[program_auth_info.member_id][function_auth_type_cd_name] \
                    = program_auth_info.enable_flag

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        return JsonResponse(member_program_auth_list, json_dumps_params={'ensure_ascii': True})


class GetTrainerInfoView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        # context = {}
        # context = super(GetTrainerInfoView, self).get_context_data(**kwargs)
        trainer_id = request.GET.get('trainer_id', '')
        class_id = request.session.get('class_id', '')
        error = None
        # class_info = None

        # center_name = '없음'
        user_member_info = None
        class_info = None
        # off_repeat_schedule_data = None
        trainer_data = collections.OrderedDict()

        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'
        if trainer_id is None or trainer_id == '':
            error = '오류가 발생했습니다.[2]'

        if error is None:
            try:
                class_info = MemberClassTb.objects.get(class_tb_id=class_id, member_id=trainer_id)
            except ObjectDoesNotExist:
                error = '지점 정보를 불러오지 못했습니다.'

        if error is None:
            try:
                user_member_info = MemberTb.objects.get(member_id=trainer_id)
            except ObjectDoesNotExist:
                error = '회원 정보를 불러오지 못했습니다.'

        if error is None:
            connection_check = 0
            if class_info.auth_cd == AUTH_TYPE_VIEW:
                connection_check = 2
            elif class_info.auth_cd == AUTH_TYPE_WAIT:
                connection_check = 1
            if user_member_info.reg_info is None or str(user_member_info.reg_info) != str(request.user.id):
                # 연결이 안되어 있는 경우 회원 정보 표시 안함
                if connection_check != 2:
                    user_member_info.sex = ''
                    user_member_info.birthday_dt = ''
                    if user_member_info.phone is None or user_member_info.phone == '':
                        user_member_info.phone = ''
                    else:
                        user_member_info.phone = '*******' + user_member_info.phone[7:]
                    user_member_info.user.email = ''
                    user_member_info.profile_url = '/static/common/icon/icon_account.png'
            if user_member_info.profile_url is None or user_member_info.profile_url == '':
                user_member_info.profile_url = '/static/common/icon/icon_account.png'

            trainer_data = {'trainer_id': request.user.id,
                            'trainer_user_id': user_member_info.user.username,
                            'trainer_name': user_member_info.name,
                            'trainer_phone': str(user_member_info.phone),
                            'trainer_email': str(user_member_info.user.email),
                            'trainer_sex': str(user_member_info.sex),
                            'trainer_birthday_dt': str(user_member_info.birthday_dt),
                            'trainer_profile_url': user_member_info.profile_url,
                            'trainer_phone_is_active': str(user_member_info.phone_is_active),
                            'trainer_connection_check': connection_check,
                            'trainer_is_active': str(user_member_info.user.is_active),
                            'program': {'program_id': class_id,
                                        'program_subject_type_name': class_info.class_tb.get_class_type_cd_name(),
                                        'program_program_owner_id': class_info.class_tb.member_id,
                                        'program_program_owner_name': class_info.class_tb.member.name,
                                      }
                           }

        return JsonResponse(trainer_data, json_dumps_params={'ensure_ascii': True})


class GetTrainerLectureListViewAjax(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        trainer_id = request.GET.get('trainer_id', '')
        lecture_data = LectureTb.objects.select_related(
            'main_trainer').filter(class_tb_id=class_id, state_cd=STATE_CD_IN_PROGRESS,
                                   main_trainer_id=trainer_id, use=USE).order_by('lecture_id')

        trainer_lecture_list = []
        # 수업과 연관되어있는 수강권 정보 셋팅
        for lecture_info in lecture_data:
            lecture_id = str(lecture_info.lecture_id)
            main_trainer_id = lecture_info.main_trainer_id
            main_trainer_name = lecture_info.main_trainer.name
            lecture_data_dict = {'lecture_id': lecture_id,
                                 'lecture_name': lecture_info.name,
                                 'lecture_note': lecture_info.note,
                                 'lecture_max_num': lecture_info.member_num,
                                 'lecture_ing_color_cd': lecture_info.ing_color_cd,
                                 'lecture_ing_font_color_cd': lecture_info.ing_font_color_cd,
                                 'lecture_end_color_cd': lecture_info.end_color_cd,
                                 'lecture_end_font_color_cd': lecture_info.end_font_color_cd,
                                 'lecture_minute': lecture_info.lecture_minute,
                                 'lecture_type_cd': lecture_info.lecture_type_cd,
                                 'lecture_mod_dt': str(lecture_info.mod_dt),
                                 'lecture_reg_dt': str(lecture_info.reg_dt),
                                 'lecture_state_cd': lecture_info.state_cd,
                                 'main_trainer_id': main_trainer_id,
                                 'main_trainer_name': main_trainer_name}
            trainer_lecture_list.append(lecture_data_dict)

        return JsonResponse({'trainer_lecture_data': trainer_lecture_list},
                             json_dumps_params={'ensure_ascii': True})


# 회원수정
def update_trainer_info_logic(request):
    trainer_id = request.POST.get('trainer_id')
    first_name = request.POST.get('first_name', '')
    phone = request.POST.get('phone', '')
    sex = request.POST.get('sex', '')
    birthday_dt = request.POST.get('birthday', '')
    class_id = request.session.get('class_id', '')
    error = None
    member = None

    if trainer_id == '':
        error = '강사 ID를 확인해 주세요.'

    if error is None:
        try:
            member = MemberTb.objects.get(member_id=trainer_id)
        except ObjectDoesNotExist:
            error = '강사 ID를 확인해 주세요.'
    if error is None:
        if member.user.is_active:
            error = '강사 정보를 수정할수 없습니다.'
        if str(request.user.id) != str(member.reg_info):
            error = '다른 강사가 등록한 강사는 수정할수 없습니다.'
    if error is None:
        try:
            with transaction.atomic():
                # 회원의 이름을 변경하는 경우 자동으로 ID 변경되도록 설정
                if member.user.first_name != first_name:
                    username = first_name
                    i = 0
                    count = MemberTb.objects.filter(name=username).count()
                    max_range = (100 * (10 ** len(str(count)))) - 1
                    for i in range(0, 100):
                        username += str(random.randrange(0, max_range)).zfill(len(str(max_range)))
                        try:
                            User.objects.get(username=username)
                        except ObjectDoesNotExist:
                            break

                    if i == 100:
                        error = 'ID 변경에 실패했습니다. 다시 시도해주세요.'
                        raise InternalError

                    member.user.username = username
                    member.user.first_name = first_name
                    member.user.save()

                member.name = first_name
                member.phone = phone
                member.sex = sex
                if birthday_dt is not None and birthday_dt != '':
                    member.birthday_dt = birthday_dt
                member.save()

        except ValueError:
            error = '등록 값에 문제가 있습니다.'
        except IntegrityError:
            error = '등록 값에 문제가 있습니다.'
        except TypeError:
            error = '등록 값에 문제가 있습니다.'
        except ValidationError:
            error = '등록 값에 문제가 있습니다.'
        except InternalError:
            error = error

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member.name+' 강사 정보', log_how='변경', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


def update_trainer_connection_info_logic(request):
    trainer_id = request.POST.get('trainer_id', '')
    auth_cd = request.POST.get('trainer_auth_cd', '')
    class_id = request.session.get('class_id', '')

    error = None
    member_info = None
    if trainer_id is None or trainer_id == '':
        error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        if str(request.user.id) == str(trainer_id):
            error = '본인은 등록할수 없습니다.'

    if error is None:
        try:
            member_info = MemberTb.objects.get(member_id=trainer_id)
        except ObjectDoesNotExist:
            error = '강사 정보를 불러오지 못했습니다.'

    if error is None:
        try:
            class_member_info = MemberClassTb.objects.select_related(
                'class_tb', 'member').get(class_tb_id=class_id, member_id=trainer_id, use=USE)
            class_member_info.auth_cd = auth_cd
            class_member_info.own_cd = OWN_TYPE_EMPLOYEE
            class_member_info.mod_member_id = request.user.id
            class_member_info.save()
        except ObjectDoesNotExist:
            class_member_info = MemberClassTb(class_tb_id=class_id, member_id=trainer_id, auth_cd=auth_cd,
                                              mod_member_id=request.user.id, own_cd=OWN_TYPE_EMPLOYEE, use=USE)
            class_member_info.save()

    if error is None:
        if auth_cd == AUTH_TYPE_DELETE:
            lecture_data = LectureTb.objects.filter(class_tb_id=class_id, main_trainer_id=trainer_id, use=USE)
            schedule_data = ScheduleTb.objects.filter(class_tb_id=class_id, trainer_id=trainer_id,
                                                      end_dt__gt=timezone.now(), state_cd=STATE_CD_NOT_PROGRESS)
            repeat_schedule_data = RepeatScheduleTb.objects.filter(class_tb_id=class_id,
                                                                   end_date__gt=datetime.date.today(),
                                                                   repeat_trainer_id=trainer_id)

            try:
                with transaction.atomic():
                    # 등록된 일정 정리
                    if len(schedule_data) > 0:
                        # 예약된 일정 삭제
                        schedule_data.update(trainer_id=request.user.id)
                    if len(repeat_schedule_data) > 0:
                        # 반복일정 삭제
                        repeat_schedule_data.update(repeat_trainer_id=request.user.id)

                    lecture_data.update(main_trainer_id=request.user.id)

            except ValueError:
                error = '등록 값에 문제가 있습니다.'
            except IntegrityError:
                error = '등록 값에 문제가 있습니다.'
            except TypeError:
                error = '등록 값의 형태가 문제 있습니다.'
            except ValidationError:
                error = '등록 값의 형태가 문제 있습니다'
            except InternalError:
                error = '등록 값에 문제가 있습니다.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_how = '강사 연결 요청'
        if auth_cd == AUTH_TYPE_DELETE:
            log_how = '강사 연결 해제'
        log_data = LogTb(log_type='LP02', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member_info.name + '님 에게 \''
                                  + request.session.get('class_type_name', '') + '\' 지점',
                         log_how=log_how, log_detail='', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


# 강사 삭제
def delete_trainer_info_logic(request):
    class_id = request.session.get('class_id', '')
    trainer_id = request.POST.get('trainer_id')
    user_id = request.user.id
    error = None
    member_class_data = None
    program_auth_data = None
    lecture_data = None
    member = None
    if trainer_id == '':
        error = '강사 ID를 확인해 주세요.'

    if error is None:
        try:
            member = MemberTb.objects.get(member_id=trainer_id)
        except ObjectDoesNotExist:
            error = '강사 ID를 확인해 주세요.'

    if error is None:
        user_id = request.user.id
        # try:
        #     class_info = ClassTb.objects.get(class_id=class_id)
        #     user_id = class_info.member.member_id
        # except ObjectDoesNotExist:
        #     user_id = request.user.id

    if error is None:
        if user_id == trainer_id:
            error = '지점 소유자는 삭제할수 없습니다.'

    if error is None:
        program_auth_data = ProgramAuthTb.objects.filter(class_tb_id=class_id, member_id=trainer_id)
        member_class_data = MemberClassTb.objects.filter(class_tb_id=class_id, member_id=trainer_id,
                                                         own_cd=OWN_TYPE_EMPLOYEE)

        lecture_data = LectureTb.objects.filter(class_tb_id=class_id, main_trainer_id=trainer_id, use=USE)

    if error is None:

        schedule_data = ScheduleTb.objects.filter(class_tb_id=class_id, trainer_id=trainer_id,
                                                  end_dt__gt=timezone.now(), state_cd=STATE_CD_NOT_PROGRESS)

        repeat_schedule_data = RepeatScheduleTb.objects.filter(class_tb_id=class_id, end_date__gt=datetime.date.today(),
                                                               repeat_trainer_id=trainer_id)

        try:
            with transaction.atomic():
                # 등록된 일정 정리
                if len(schedule_data) > 0:
                    # 예약된 일정 삭제
                    schedule_data.update(trainer_id=user_id)
                if len(repeat_schedule_data) > 0:
                    # 반복일정 삭제
                    repeat_schedule_data.update(repeat_trainer_id=user_id)

                program_auth_data.delete()
                member_class_data.delete()
                lecture_data.update(main_trainer_id=user_id)

                if member.user.is_active == 0 and str(member.reg_info) == str(request.user.id):
                    member.contents = member.user.username + ':' + str(trainer_id)
                    member.use = UN_USE
                    member.save()

        except ValueError:
            error = '등록 값에 문제가 있습니다.'
        except IntegrityError:
            error = '등록 값에 문제가 있습니다.'
        except TypeError:
            error = '등록 값의 형태가 문제 있습니다.'
        except ValidationError:
            error = '등록 값의 형태가 문제 있습니다'
        except InternalError:
            error = '등록 값에 문제가 있습니다.'

    if error is not None:
        logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info=member.name+' 강사 정보', log_how='삭제', use=USE)
        log_data.save()

    return render(request, 'ajax/trainer_error_ajax.html')


class GetTrainerClosedDateListView(LoginRequiredMixin, AccessTestMixin, View):
    def get(self, request):
        class_id = request.session.get('class_id', '')
        trainer_id = request.GET.get('trainer_id', '')
        error = None
        today = datetime.date.today()
        trainer_closed_list = []

        if trainer_id is None or trainer_id == '':
            error = '회원 정보를 불러오지 못했습니다.'

        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)

        return JsonResponse({'trainer_closed_list': trainer_closed_list}, json_dumps_params={'ensure_ascii': True})


def add_trainer_closed_date_logic(request):
    class_id = request.session.get('class_id', '')
    member_id = request.POST.get('member_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    repeat_week_type = request.POST.get('week_info', '')
    title = request.POST.get('title', '')
    contents = request.POST.get('contents', '')
    is_member_view = request.POST.get('is_member_view', '1')

    error = None
    schedule_closed_id = ''
    start_date_info = None
    end_date_info = None
    repeat_schedule_date_list = []

    if member_id is None:
        member_id = ''
    if repeat_week_type == '':
        error = '요일을 선택해주세요.'
    if start_date == '':
        error = '시작 날짜를 선택해 주세요.'
    elif end_date == '':
        error = '종료 날짜를 선택해 주세요.'

    if error is None:
        try:
            start_date_info = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date_info = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            if (start_date_info - end_date_info) > datetime.timedelta(days=365):
                error = '1년까지만 불가 일정을 등록할수 있습니다.'
        except ValueError:
            error = '날짜 오류가 발생했습니다.[0]'
        except IntegrityError:
            error = '날짜 오류가 발생했습니다.[1]'
        except TypeError:
            error = '날짜 오류가 발생했습니다.[2]'

    if error is None:
        if member_id != '':
            try:
                MemberTb.objects.get(member_id=member_id)
            except ObjectDoesNotExist:
                error = '회원 정보를 불러오지 못했습니다.'

    # 등록할 날짜 list 가져오기
    if error is None:
        week_order = ['SUN', 'MON', 'TUE', 'WED', 'THS', 'FRI', 'SAT']
        week_order = {key: i for i, key in enumerate(week_order)}
        week_data = repeat_week_type.split('/')
        week_data = sorted(week_data, key=lambda week_info: week_order.get(week_info))
        repeat_week_type = "/".join(week_data)
        repeat_schedule_date_list = func_get_repeat_schedule_date_list(repeat_week_type,
                                                                       start_date_info,
                                                                       end_date_info)
        if len(repeat_schedule_date_list) == 0:
            error = '등록할 수 있는 일정이 없습니다.'

    if error is None:
        try:
            with transaction.atomic():
                schedule_closed_info = ScheduleClosedTb(class_tb_id=class_id, member_id=member_id,
                                                        start_date=start_date_info, end_date=end_date_info,
                                                        title=title, contents=contents,
                                                        week_info=repeat_week_type,
                                                        is_member_view=is_member_view, use=USE)

                schedule_closed_info.save()

                schedule_closed_id = schedule_closed_info.schedule_closed_id

                # 반복일정 데이터 등록
                for repeat_schedule_date_info in repeat_schedule_date_list:
                    repeat_schedule_info_date = str(repeat_schedule_date_info).split(' ')[0]

                    closed_date = datetime.datetime.strptime(repeat_schedule_info_date, '%Y-%m-%d')

                    schedule_closed_day = ScheduleClosedDayTb(schedule_closed_tb_id=schedule_closed_id,
                                                              closed_date=closed_date, use=USE)
                    schedule_closed_day.save()

        except ValueError:
            error = '오류가 발생했습니다. [1]'
        except IntegrityError:
            error = '오류가 발생했습니다. [2]'
        except TypeError:
            error = '오류가 발생했습니다. [3]'
        except ValidationError:
            error = '오류가 발생했습니다. [4]'
        except InternalError:
            error = '오류가 발생했습니다. [5]'

    if error is not None:
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info='불가 일정', log_how='추가', use=USE)
        log_data.save()

    return JsonResponse({'schedule_closed_id': str(schedule_closed_id)}, json_dumps_params={'ensure_ascii': True})


def update_trainer_closed_date_logic(request):
    class_id = request.session.get('class_id', '')
    schedule_closed_id = request.POST.get('schedule_closed_id', '')
    member_id = request.POST.get('member_id', '')
    start_date = request.POST.get('start_date', '')
    end_date = request.POST.get('end_date', '')
    week_info = request.POST.get('week_info', '')
    title = request.POST.get('title', '')
    contents = request.POST.get('contents', '')
    is_member_view = request.POST.get('is_member_view', '')

    error = None
    schedule_closed_info = None
    start_date_info = None
    end_date_info = None

    if member_id is None:
        member_id = ''

    try:
        schedule_closed_info = ScheduleClosedTb.objects.get(schedule_closed_id=schedule_closed_id)
    except ObjectDoesNotExist:
        error = '오류가 발생했습니다. [0]'

    if error is None:
        if week_info == '':
            week_info = schedule_closed_info.week_info
        if start_date == '':
            start_date = str(schedule_closed_info.start_date)
        if end_date == '':
            end_date = str(schedule_closed_info.end_date)
        if title == '':
            title = schedule_closed_info.title
        if contents == '':
            contents = schedule_closed_info.contents
        if is_member_view == '':
            is_member_view = schedule_closed_info.is_member_view

    if error is None:
        try:
            start_date_info = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date_info = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            if (start_date_info - end_date_info) > datetime.timedelta(days=365):
                error = '1년까지만 불가 일정을 등록할수 있습니다.'
        except ValueError:
            error = '날짜 오류가 발생했습니다.[0]'
        except IntegrityError:
            error = '날짜 오류가 발생했습니다.[1]'
        except TypeError:
            error = '날짜 오류가 발생했습니다.[2]'

    if error is None:
        if member_id != '':
            try:
                MemberTb.objects.get(member_id=member_id)
            except ObjectDoesNotExist:
                error = '회원 정보를 불러오지 못했습니다.'

    try:
        with transaction.atomic():
            schedule_closed_info.week_info = week_info
            schedule_closed_info.start_date = start_date_info
            schedule_closed_info.end_date = end_date_info
            schedule_closed_info.title = title
            schedule_closed_info.contents = contents
            schedule_closed_info.is_member_view = is_member_view

            schedule_closed_info.save()
            schedule_closed_id = schedule_closed_info.schedule_closed_id
    except ValueError:
        error = '오류가 발생했습니다. [1]'
    except IntegrityError:
        error = '오류가 발생했습니다. [2]'
    except TypeError:
        error = '오류가 발생했습니다. [3]'
    except ValidationError:
        error = '오류가 발생했습니다. [4]'
    except InternalError:
        error = '오류가 발생했습니다. [5]'

    if error is not None:
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info='불가 일정', log_how='수정', use=USE)
        log_data.save()

    return JsonResponse({'schedule_closed_id': str(schedule_closed_id)}, json_dumps_params={'ensure_ascii': True})


def delete_trainer_closed_date_logic(request):
    class_id = request.session.get('class_id', '')
    schedule_closed_id = request.POST.get('schedule_closed_id', '')
    error = None

    if schedule_closed_id == '':
        error = '삭제할 불가 일정을 선택해주세요.'

    if error is None:
        try:
            schedule_close_info = ScheduleClosedTb.objects.get(schedule_closed_id=schedule_closed_id)
            schedule_close_info.delete()
        except ObjectDoesNotExist:
            error = '오류가 발생했습니다. [0]'

    if error is not None:
        messages.error(request, error)
    else:
        log_data = LogTb(log_type='LC01', auth_member_id=request.user.id,
                         from_member_name=request.user.first_name,
                         class_tb_id=class_id,
                         log_info='불가 일정', log_how='삭제', use=USE)
        log_data.save()

    return JsonResponse({'schedule_closed_id': str(schedule_closed_id)}, json_dumps_params={'ensure_ascii': True})



class GetTrainerMemberTicketPriceBugListView(LoginRequiredMixin, AccessTestMixin, View):

    def get(self, request):
        class_id = self.request.session.get('class_id', '')
        error = None
        member_ticket_list = collections.OrderedDict()
        bug_check = 0
        if class_id is None or class_id == '':
            error = '오류가 발생했습니다.'

        if error is None:
            try:
                bug_info = BugMemberTicketPriceTb.objects.get(class_tb_id=class_id)
                bug_check = bug_info.bug_check
            except ObjectDoesNotExist:
                bug_check = 0
        if error is None:
            if bug_check == 0:
                member_ticket_data = ClassMemberTicketTb.objects.select_related(
                    'member_ticket_tb__member',
                    'member_ticket_tb__ticket_tb').filter(class_tb_id=class_id,
                                                          auth_cd=AUTH_TYPE_VIEW,
                                                          member_ticket_tb__use=USE,
                                                          member_ticket_tb__price=0,
                                                          member_ticket_tb__reg_dt__gte='2020-02-02',
                                                          member_ticket_tb__reg_dt__lt='2020-02-09',
                                                          use=USE).order_by('-member_ticket_tb__start_date',
                                                                            '-member_ticket_tb__reg_dt')
                for member_ticket_info in member_ticket_data:
                    member_ticket_tb = member_ticket_info.member_ticket_tb
                    ticket_tb = member_ticket_tb.ticket_tb
                    ticket_id = str(ticket_tb.ticket_id)
                    if '\r\n' in member_ticket_tb.note:
                        member_ticket_tb.note = member_ticket_tb.note.replace('\r\n', ' ')
                    if ticket_id != '3198' and ticket_id != '3238' and ticket_id != '3315' and ticket_id != '4334' \
                            and ticket_id != '4337' and ticket_id != '4339' and ticket_id != '4351' \
                            and ticket_id != '4374' and ticket_id != '4379' and ticket_id != '4383'and ticket_id != '4386' \
                            and ticket_id != '4397' and ticket_id != '4417' and ticket_id != '4434':
                        member_ticket_info = {'member_name': str(member_ticket_tb.member.name),
                                              'member_id': str(member_ticket_tb.member.member_id),
                                              'member_ticket_id': str(member_ticket_tb.member_ticket_id),
                                              'member_ticket_name': ticket_tb.name,
                                              'member_ticket_state_cd': member_ticket_tb.state_cd,
                                              'member_ticket_reg_count': member_ticket_tb.member_ticket_reg_count,
                                              'member_ticket_rem_count': member_ticket_tb.member_ticket_rem_count,
                                              'member_ticket_avail_count': member_ticket_tb.member_ticket_avail_count,
                                              'member_ticket_start_date': str(member_ticket_tb.start_date),
                                              'member_ticket_end_date': str(member_ticket_tb.end_date),
                                              'member_ticket_price': member_ticket_tb.price,
                                              'member_ticket_pay_method': member_ticket_tb.pay_method,
                                              'member_ticket_refund_date': str(member_ticket_tb.refund_date),
                                              'member_ticket_refund_price': member_ticket_tb.refund_price,
                                              'member_ticket_note': str(member_ticket_tb.note),
                                              'member_ticket_reg_dt': str(member_ticket_tb.reg_dt)
                                              }
                        member_ticket_list[str(member_ticket_tb.member_ticket_id)] = member_ticket_info
        if error is not None:
            logger.error(request.user.first_name + '[' + str(request.user.id) + ']' + error)
            messages.error(request, error)
        return JsonResponse(member_ticket_list, json_dumps_params={'ensure_ascii': True})


def add_trainer_member_ticket_price_bug_check_logic(request):
    context = {}

    class_id = request.session.get('class_id', '')
    error = None
    if class_id is None or class_id == '':
        error = '오류가 발생했습니다.'

    if error is None:
        try:
            bug_info = BugMemberTicketPriceTb.objects.get(class_tb_id=class_id)
            bug_info.bug_check = 1
            bug_info.save()
        except ObjectDoesNotExist:
            bug_info = BugMemberTicketPriceTb(member_id=request.user.id,
                                              class_tb_id=class_id, bug_check=1, use=USE)
            bug_info.save()

    if error is not None:
        messages.error(request, error)
        context['messageArray'] = error
    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def delete_trainer_member_ticket_price_bug_check_logic(request):
    context = {}

    class_id = request.session.get('class_id', '')
    error = None
    if class_id is None or class_id == '':
        error = '오류가 발생했습니다.'

    if error is None:
        try:
            bug_info = BugMemberTicketPriceTb.objects.get(class_tb_id=class_id)
            bug_info.delete()
        except ObjectDoesNotExist:
            error = None

    if error is not None:
        messages.error(request, error)
        context['messageArray'] = error
    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def holding_test_logic(request):
    context = {}
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)

    holding_data = MemberClosedDateHistoryTb.objects.filter(start_date__lte=today,
                                                            end_date__gte=yesterday, use=USE).order_by('start_date')

    for holding_info in holding_data:
        member_ticket_tb = holding_info.member_ticket_tb
        if holding_info.start_date <= today <= holding_info.end_date:
            # holding 기간에 속한 경우 일시정지 처리
            member_ticket_tb.state_cd = STATE_CD_HOLDING

        elif today > holding_info.end_date:
            # holding 기간이 지난 경우 재개 처리
            if member_ticket_tb.state_cd == STATE_CD_HOLDING:
                member_ticket_tb.state_cd = STATE_CD_IN_PROGRESS

        member_ticket_tb.save()

    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def member_setting_test_logic(request):
    context = {}
    today = datetime.date.today()

    schedule_alarm_data = ScheduleAlarmTb.objects.select_related('member', 'schedule_tb__member_ticket_tb').filter()
    for schedule_alarm_info in schedule_alarm_data:
        member = schedule_alarm_info.member
        member_id = schedule_alarm_info.member.member_id
        schedule_tb = schedule_alarm_info.schedule_tb
        member_ticket_tb = ''
        member_ticket_tb_member_id = ''
        if schedule_tb is not None and schedule_tb != '':
            member_ticket_tb = schedule_alarm_info.schedule_tb.member_ticket_tb

        if member_ticket_tb is not None and member_ticket_tb != '':
            member_ticket_tb_member_id = member_ticket_tb.member_id
        if str(member_id) != member_ticket_tb_member_id:
            if str(member.user.groups.all()[0]) == 'trainee':
                print('alarm_id::'+str(schedule_alarm_info.schedule_alarm_id))
                # schedule_alarm_info.delete()

    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def update_schedule_trainer_test(request):
    context = {}
    schedule_data = ScheduleTb.objects.select_related('class_tb__member').filter(trainer=None,
                                                                                 en_dis_type=ON_SCHEDULE_TYPE)

    for schedule_info in schedule_data:
        schedule_info.trainer_id = schedule_info.class_tb.member_id
        schedule_info.save()
    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def update_repeat_schedule_trainer_test(request):
    context = {}
    repeat_schedule_data = RepeatScheduleTb.objects.select_related(
        'class_tb__member').filter(en_dis_type=ON_SCHEDULE_TYPE)

    for repeat_schedule_info in repeat_schedule_data:
        repeat_schedule_info.repeat_trainer_id = repeat_schedule_info.class_tb.member_id
        repeat_schedule_info.save()

    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def update_lecture_trainer_test(request):
    context = {}
    lecture_data = LectureTb.objects.select_related('class_tb__member').filter()

    for lecture_info in lecture_data:
        lecture_info.main_trainer_id = lecture_info.class_tb.member_id
        lecture_info.save()

    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def update_member_ticket_payment_test(request):
    context = {}
    member_ticket_data = MemberTicketTb.objects.select_related('ticket_tb__class_tb', 'member').filter()

    for member_ticket_info in member_ticket_data:
        member_payment_history_info = MemberPaymentHistoryTb(class_tb_id=member_ticket_info.ticket_tb.class_tb_id,
                                                             member_id=member_ticket_info.member_id,
                                                             member_ticket_tb_id=member_ticket_info.member_ticket_id,
                                                             member_shop_tb_id=None,
                                                             payment_price=member_ticket_info.price,
                                                             refund_price=0,
                                                             pay_method=member_ticket_info.pay_method,
                                                             pay_date=member_ticket_info.start_date,
                                                             note='', use=USE)
        member_payment_history_info.save()
        if member_ticket_info.refund_price > 0:
            member_payment_history_info = MemberPaymentHistoryTb(class_tb_id=member_ticket_info.ticket_tb.class_tb_id,
                                                                 member_id=member_ticket_info.member_id,
                                                                 member_ticket_tb_id=member_ticket_info.member_ticket_id,
                                                                 member_shop_tb_id=None,
                                                                 payment_price=0,
                                                                 refund_price=member_ticket_info.refund_price,
                                                                 pay_method=member_ticket_info.pay_method,
                                                                 pay_date=member_ticket_info.refund_date,
                                                                 note='', use=USE)
            member_payment_history_info.save()

    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})


def update_member_ticket_payment_test1(request):
    context = {}
    member_ticket_data = MemberTicketTb.objects.select_related('ticket_tb__class_tb', 'member').filter()

    for member_ticket_info in member_ticket_data:
        member_ticket_info.payment_price = member_ticket_info.price
        member_ticket_info.save()

    return JsonResponse(context, json_dumps_params={'ensure_ascii': True})
